#!/usr/bin/env python3
"""
Chhattisgarh SLBC Data Extractor v2
====================================

Extends extract_chhattisgarh.py to:
  1. Cover older meetings (78–89) downloaded from slbcchhattisgarh.com
  2. Handle nested folder structures used by older zips
     (e.g. data-tables/88/88 DATA TABLE/, data-tables/89/89 Dec. 2022 Tables/)
  3. Add an agenda-PDF fallback path (pdftotext -layout) for quick validation
     of a single meeting when Excel data-tables are missing.

Source URL: https://slbcchhattisgarh.com (Convenor: State Bank of India)
The "data-table" ZIP per meeting is the authoritative source; the agenda PDF
is used only as a fallback / sanity-check.

Outputs are identical to extract_chhattisgarh.py: per-quarter CSVs,
chhattisgarh_complete.json, chhattisgarh_fi_timeseries.json/csv.
"""

import os
import re
import sys
import glob
import json
import csv
import subprocess
import warnings

import openpyxl

# Reuse all extraction primitives from v1
from extract_chhattisgarh import (
    CHHATTISGARH_DISTRICTS,
    DISTRICT_ALIASES,
    DISTRICT_PATTERNS,
    normalize_district,
    parse_number,
    quarter_from_month,
    detect_quarter_from_header,
    extract_deposit_advance,
    extract_branch_network,
    extract_atm_network,
    extract_acp_achievement,
    get_fy,
    save_quarterly_csvs,
    build_complete_json,
    build_timeseries,
    MEETING_QUARTERS as MEETING_QUARTERS_V1,
)

warnings.filterwarnings('ignore')

# Extended meeting-to-quarter mapping (from slbcchhattisgarh.com/slbc_meetings.php).
# Indented carry-over from v1 + older meetings 50..89.
MEETING_QUARTERS_EXT = {
    # Older meetings — only used as quarter labels when a meeting actually
    # surfaces district-wise data. Many of these will have only bank-wise
    # Excel sheets in their data-table zip.
    50:  ('2013-03', 'March 2013',     '31-03-2013'),
    54:  ('2014-03', 'March 2014',     '31-03-2014'),
    58:  ('2015-03', 'March 2015',     '31-03-2015'),
    62:  ('2016-03', 'March 2016',     '31-03-2016'),
    66:  ('2017-03', 'March 2017',     '31-03-2017'),
    70:  ('2018-03', 'March 2018',     '31-03-2018'),
    71:  ('2018-06', 'June 2018',      '30-06-2018'),
    72:  ('2018-09', 'September 2018', '30-09-2018'),
    74:  ('2019-03', 'March 2019',     '31-03-2019'),
    76:  ('2019-09', 'September 2019', '30-09-2019'),
    77:  ('2019-12', 'December 2019',  '31-12-2019'),
    78:  ('2020-03', 'March 2020',     '31-03-2020'),
    79:  ('2020-06', 'June 2020',      '30-06-2020'),
    80:  ('2020-09', 'September 2020', '30-09-2020'),
    81:  ('2020-12', 'December 2020',  '31-12-2020'),
    82:  ('2021-03', 'March 2021',     '31-03-2021'),
    83:  ('2021-06', 'June 2021',      '30-06-2021'),
    84:  ('2021-09', 'September 2021', '30-09-2021'),
    85:  ('2021-12', 'December 2021',  '31-12-2021'),
    86:  ('2022-03', 'March 2022',     '31-03-2022'),
    87:  ('2022-06', 'June 2022',      '30-06-2022'),
    88:  ('2022-09', 'September 2022', '30-09-2022'),
    89:  ('2022-12', 'December 2022',  '31-12-2022'),
}
MEETING_QUARTERS = {**MEETING_QUARTERS_EXT, **MEETING_QUARTERS_V1}


# ---- Directory descent for older meetings ----

def _resolve_data_dir(data_dir):
    """Older zips contain a single nested folder. Descend into it if so."""
    if not os.path.isdir(data_dir):
        return data_dir
    entries = [e for e in os.listdir(data_dir) if not e.startswith('.')]

    # Existing v1 convention: a 'TABLE' subfolder (used by meeting 98)
    if 'TABLE' in entries and os.path.isdir(os.path.join(data_dir, 'TABLE')):
        return os.path.join(data_dir, 'TABLE')

    # Single subfolder shape (e.g. '88 DATA TABLE', '89 Dec. 2022  Tables',
    # '84/data table update', '86/data table')
    if len(entries) == 1 and os.path.isdir(os.path.join(data_dir, entries[0])):
        return os.path.join(data_dir, entries[0])

    return data_dir


# ---- File finding (looser than v1, to match older naming) ----

def _find_xlsx(data_dir, *substrs):
    """Find an .xlsx file whose lowercase basename contains all the substrs."""
    for f in glob.glob(os.path.join(data_dir, '*.xlsx')):
        bn = os.path.basename(f).lower()
        if all(s.lower() in bn for s in substrs):
            return f
    return None


def _find_xlsx_any(data_dir, substrs_list):
    """Try each substring tuple in order; return first match."""
    for substrs in substrs_list:
        f = _find_xlsx(data_dir, *substrs)
        if f:
            return f
    return None


# ---- Per-meeting Excel processor (refactor of v1's process_meeting) ----

def process_meeting_excel(meeting_num, data_dir):
    """Run all extractors against a meeting's data-tables folder."""
    data_dir = _resolve_data_dir(data_dir)
    print(f"\n{'='*60}\nProcessing Meeting {meeting_num}  (dir: {data_dir})\n{'='*60}")

    if not os.path.isdir(data_dir):
        print(f"  Data dir not found: {data_dir}")
        return None

    xlsx_files = glob.glob(os.path.join(data_dir, '*.xlsx'))
    if not xlsx_files:
        print(f"  No .xlsx files in {data_dir} (PDFs/scans only)")
        return None

    results = []
    quarter_info = None

    # 1. Deposit/Advance/CD Ratio — handle both v1 ("Table_1 Deposit Advance")
    #    and older ("Table_1 part 1.xlsx", "OUTPUT TABLE 1 B.xlsx") layouts.
    dep_file = _find_xlsx_any(data_dir, [
        ('deposit', 'advance'),
        ('advance', 'deposit'),
        ('table 1', 'advance'),
        ('table_1', 'part 1'),
        ('table 1', 'part 1'),
        ('output table 1', 'b'),    # Meeting 89's "OUTPUT TABLE 1 B.xlsx"
    ])
    # Fallback: standalone "Table 1"/"Table_1" file w/o Branch/ATM/part2
    if not dep_file:
        for f in xlsx_files:
            bn = os.path.basename(f).lower()
            if re.match(r'^table[ _]1\b', bn) and not any(
                bad in bn for bad in ('table 11', 'table_11', 'branch', 'atm', 'part 2', 'part_2')
            ):
                dep_file = f
                break
    if dep_file:
        try:
            wb = openpyxl.load_workbook(dep_file, data_only=True)
            t = extract_deposit_advance(wb, meeting_num)
            wb.close()
            if t:
                quarter_info = t.pop('quarter_info', None)
                results.append(t)
                print(f"  [OK] credit_deposit_ratio: {len(t['districts'])} districts ({os.path.basename(dep_file)})")
            else:
                print(f"  [SKIP] credit_deposit_ratio: no district sheet in {os.path.basename(dep_file)}")
        except Exception as e:
            print(f"  [ERR] credit_deposit_ratio: {e}")

    # 2. Branch & ATM network — v1 file pattern ("branch atm") OR older
    #    "Table_1 part 2.xlsx" / "OUTPUT TABLE 1 A.xlsx" (89) which carry
    #    the DISTRIST WISE BR / DISTRICT WISE ATM sheets.
    br_file = _find_xlsx_any(data_dir, [
        ('branch', 'atm'),
        ('table_1', 'part 2'),
        ('table 1', 'part 2'),
        ('output table 1', 'a'),    # Meeting 89's "OUTPUT TABLE 1 A.xlsx"
    ])
    if br_file:
        try:
            wb = openpyxl.load_workbook(br_file, data_only=True)
            t = extract_branch_network(wb, meeting_num)
            if t:
                results.append(t)
                print(f"  [OK] branch_network: {len(t['districts'])} districts ({os.path.basename(br_file)})")
            t = extract_atm_network(wb, meeting_num)
            if t:
                results.append(t)
                print(f"  [OK] atm_network: {len(t['districts'])} districts")
            wb.close()
        except Exception as e:
            print(f"  [ERR] branch/atm: {e}")

    # 3. ACP achievement (Table 4)
    acp_file = _find_xlsx_any(data_dir, [
        ('table 4',), ('output table 4',), ('table_4',),
    ])
    if acp_file:
        try:
            wb = openpyxl.load_workbook(acp_file, data_only=True)
            t = extract_acp_achievement(wb, meeting_num)
            wb.close()
            if t:
                results.append(t)
                print(f"  [OK] acp_target_achievement: {len(t['districts'])} districts ({os.path.basename(acp_file)})")
            else:
                print(f"  [SKIP] acp_target_achievement: no district sheet")
        except Exception as e:
            print(f"  [ERR] acp_target_achievement: {e}")

    if not results:
        print(f"  No district-wise data extracted from meeting {meeting_num}")
        return None

    if meeting_num in MEETING_QUARTERS:
        qk, period, as_on = MEETING_QUARTERS[meeting_num]
    elif quarter_info:
        qk, period, as_on = quarter_info
    else:
        print(f"  WARNING: no quarter mapping for meeting {meeting_num}")
        return None

    print(f"  -> Quarter: {period}  (as on {as_on}), {len(results)} tables")
    return {'quarter_key': qk, 'period': period, 'as_on_date': as_on, 'tables': results}


# ---- Agenda PDF fallback (validation only) ----

def extract_agenda_pdf_cd_ratio(pdf_path, meeting_num=None):
    """Best-effort extraction of district-wise CD ratio from an agenda PDF
    using pdftotext -layout. Only meant as a sanity-check / validation source
    for meetings where Excel data-tables are unavailable.

    Returns a single 'credit_deposit_ratio' table dict in v1 schema, with the
    columns reduced to (cd_ratio_prev_quarter, overall_cd_ratio) only — the
    PDF agendas typically print just Sept vs Dec columns, not full deposits.
    """
    if not os.path.exists(pdf_path):
        print(f"  PDF not found: {pdf_path}")
        return None

    try:
        out = subprocess.run(
            ['pdftotext', '-layout', pdf_path, '-'],
            capture_output=True, text=True, check=True,
        )
    except Exception as e:
        print(f"  pdftotext failed: {e}")
        return None

    text = out.stdout
    lines = text.split('\n')

    # Find the district CD ratio block. It is a numbered table 1..33 with the
    # district name in UPPERCASE and the last numeric value being the current
    # quarter's CD ratio.
    rows = []
    upper_d = {d.upper(): d for d in CHHATTISGARH_DISTRICTS}
    for d, canon in list(upper_d.items()):
        upper_d[d.replace('-', ' ')] = canon
    upper_d['GARIYABAND'] = 'Gariaband'
    upper_d['KORIYA'] = 'Korea'

    for line in lines:
        m = re.match(r'^\s*(\d{1,2})\s+([A-Z][A-Z \-\.&\(\)]+?)\s+'
                     r'(-?\d+\.\d+)\s+(-?\d+\.\d+)(?:\s+(-?\d+\.\d+))?\s*$',
                     line)
        if not m:
            continue
        sno = int(m.group(1))
        if sno < 1 or sno > 33:
            continue
        d_raw = m.group(2).strip()
        district = upper_d.get(d_raw)
        if not district:
            # Try fuzzy match against canonical list
            for u, canon in upper_d.items():
                if d_raw.startswith(u) or u in d_raw:
                    district = canon
                    break
        if not district:
            continue
        prev_q = m.group(3)
        curr   = m.group(4)
        rows.append((district, prev_q, curr))

    if not rows:
        return None

    print(f"  PDF CD-ratio rows recovered: {len(rows)}")
    headers = ['overall_cd_ratio', 'cd_ratio_prev_quarter']
    districts = {}
    for d, prev_q, curr in rows:
        # Deduplicate (PDFs sometimes repeat the table on subsequent pages)
        if d in districts:
            continue
        districts[d] = {
            'overall_cd_ratio': curr,
            'cd_ratio_prev_quarter': prev_q,
        }
    return {
        'category': 'credit_deposit_ratio',
        'headers': headers,
        'districts': districts,
        'is_dict': True,
        '_source': 'agenda_pdf',
    }


# ---- Main ----

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("Usage:")
        print("  python3 extract_chhattisgarh_v2.py --all")
        print("  python3 extract_chhattisgarh_v2.py 88 89 90 ...")
        print("  python3 extract_chhattisgarh_v2.py --pdf 101th-agenda.pdf  (PDF-only validation)")
        print("  python3 extract_chhattisgarh_v2.py --output-dir DIR --all")
        sys.exit(1)

    # Parse args
    meetings = []
    output_dir = '../../public/slbc-data/chhattisgarh'
    data_base = 'data-tables'
    process_all = False
    pdf_path = None

    i = 1
    while i < len(sys.argv):
        a = sys.argv[i]
        if a == '--output-dir' and i + 1 < len(sys.argv):
            output_dir = sys.argv[i + 1]; i += 2
        elif a == '--data-dir' and i + 1 < len(sys.argv):
            data_base = sys.argv[i + 1]; i += 2
        elif a == '--all':
            process_all = True; i += 1
        elif a == '--pdf' and i + 1 < len(sys.argv):
            pdf_path = sys.argv[i + 1]; i += 2
        else:
            try:
                meetings.append(int(a))
            except ValueError:
                print(f"warning: ignoring {a!r}")
            i += 1

    # PDF-only validation mode
    if pdf_path:
        print(f"\n=== PDF agenda validation mode: {pdf_path} ===")
        table = extract_agenda_pdf_cd_ratio(pdf_path)
        if not table:
            print("No district-wise CD ratio found in PDF.")
            sys.exit(1)
        print(f"Categories: {table['category']}, districts: {len(table['districts'])}")
        for d in sorted(table['districts'].keys())[:5]:
            print(f"  {d}: {table['districts'][d]}")
        sys.exit(0)

    if process_all:
        for d in sorted(glob.glob(os.path.join(data_base, '*'))):
            if os.path.isdir(d):
                try:
                    meetings.append(int(os.path.basename(d)))
                except ValueError:
                    pass

    if not meetings:
        print("No meetings to process!"); sys.exit(1)

    meetings = sorted(set(meetings))
    print(f"Processing {len(meetings)} meeting(s): {meetings}")
    print(f"Output directory: {output_dir}")

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'quarterly'), exist_ok=True)

    all_quarters = {}
    for n in meetings:
        ddir = os.path.join(data_base, str(n))
        if not os.path.isdir(ddir):
            print(f"warning: data-tables/{n} not found, skipping")
            continue
        try:
            res = process_meeting_excel(n, ddir)
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"  ERROR meeting {n}: {e}")
            res = None
        if not res:
            continue

        qk = res['quarter_key']
        save_quarterly_csvs(res['tables'], qk, output_dir)

        if qk in all_quarters:
            existing_cats = {t['category'] for t in all_quarters[qk]['tables']}
            for t in res['tables']:
                # Replace rather than append on duplicate category
                all_quarters[qk]['tables'] = [
                    x for x in all_quarters[qk]['tables']
                    if x['category'] != t['category']
                ]
                all_quarters[qk]['tables'].append(t)
        else:
            all_quarters[qk] = {
                'period': res['period'],
                'as_on_date': res['as_on_date'],
                'tables': res['tables'],
            }

    if not all_quarters:
        print("\nNo data extracted from any meeting!"); sys.exit(1)

    complete = build_complete_json(all_quarters, output_dir)
    build_timeseries(complete, output_dir)

    # Summary
    print(f"\n{'='*60}\nEXTRACTION SUMMARY\n{'='*60}")
    print(f"Meetings processed : {len(meetings)}")
    print(f"Quarters extracted : {len(all_quarters)}")
    cats_seen = set()
    for qk in sorted(all_quarters.keys()):
        qd = all_quarters[qk]
        cats = sorted({t['category'] for t in qd['tables']})
        cats_seen.update(cats)
        print(f"  {qk}  ({qd['period']:>17}): {len(cats)} tables  {', '.join(cats)}")
    print(f"Distinct categories: {len(cats_seen)}  {', '.join(sorted(cats_seen))}")


if __name__ == '__main__':
    main()
