#!/usr/bin/env python3
"""
Chhattisgarh SLBC Data Extractor
Extracts district-wise banking data from Chhattisgarh SLBC Excel data-table files.
Data source: http://slbcchhattisgarh.com/ (Convenor: State Bank of India)

Each SLBC meeting provides a ZIP of Excel files containing:
- Table_1 Deposit Advance: District-wise deposits, advances, CD ratio (sheet: Districtwise de adv / t1 a-1)
- Table_1 Branch ATM: District-wise branch network (sheet: DISTRIST WISE BR) & ATM network (sheet: DISTRICT WISE ATM)
- OUTPUT TABLE 4: District-wise ACP achievement (sheet: *Districtwise* / *districtwise*)

Outputs: quarterly CSVs, chhattisgarh_complete.json, chhattisgarh_fi_timeseries.json/csv.
"""

import json
import csv
import os
import re
import sys
import glob
import warnings
from pathlib import Path
from collections import OrderedDict

import openpyxl

warnings.filterwarnings('ignore')

# Chhattisgarh's 33 districts (canonical names — title case)
CHHATTISGARH_DISTRICTS = [
    "Balod", "Baloda Bazar", "Balrampur", "Bastar", "Bemetara",
    "Bijapur", "Bilaspur", "Dantewada", "Dhamtari", "Durg",
    "Gariaband", "Gaurela-Pendra-Marwahi", "Janjgir-Champa", "Jashpur",
    "Kabirdham", "Kanker", "Khairagarh-Chhuikhadan-Gandai", "Kondagaon",
    "Korba", "Korea", "Mahasamund", "Manendragarh-Chirmiri-Bharatpur",
    "Mohla-Manpur-Ambagarh Chowki", "Mungeli", "Narayanpur", "Raigarh",
    "Raipur", "Rajnandgaon", "Sakti", "Sarangarh-Bilaigarh", "Sukma",
    "Surajpur", "Surguja"
]

# Fuzzy matching for district names
DISTRICT_ALIASES = {}
for d in CHHATTISGARH_DISTRICTS:
    DISTRICT_ALIASES[d.upper()] = d
    DISTRICT_ALIASES[d] = d

DISTRICT_ALIASES.update({
    # Common variations in data
    "BALODA BAZAR": "Baloda Bazar",
    "BALODABAZAR": "Baloda Bazar",
    "BALODA-BAZAR": "Baloda Bazar",
    "GARIYABAND": "Gariaband",
    "GARIABAND": "Gariaband",
    "KABIRDHAM": "Kabirdham",
    "KAWARDHA": "Kabirdham",
    "KAWARDHA (KABIRDHAM)": "Kabirdham",
    "KOREA": "Korea",
    "KORIYA": "Korea",
    "DANTEWADA": "Dantewada",
    "DANTEWARA": "Dantewada",
    "JANJGIR-CHAMPA": "Janjgir-Champa",
    "JANJGIR CHAMPA": "Janjgir-Champa",
    "KHAIRAGARH CHHUIKHADAN-GANDAI": "Khairagarh-Chhuikhadan-Gandai",
    "KHAIRAGARH-CHHUIKHADAN-GANDAI": "Khairagarh-Chhuikhadan-Gandai",
    "KHAIRAGARH CHHUIKHADAN GANDAI": "Khairagarh-Chhuikhadan-Gandai",
    "MANENDRAGARH-CHIRMIRI BHARATPUR": "Manendragarh-Chirmiri-Bharatpur",
    "MANENDRAGARH-CHIRMIRI-BHARATPUR": "Manendragarh-Chirmiri-Bharatpur",
    "MANENDRAGARH CHIRMIRI BHARATPUR": "Manendragarh-Chirmiri-Bharatpur",
    "MANENDRAGARH-CHIRMIRI BHARATPUR (MCB)": "Manendragarh-Chirmiri-Bharatpur",
    "MOHLA-MANPUR AMBAGARH CHOUKI": "Mohla-Manpur-Ambagarh Chowki",
    "MOHLA-MANPUR-AMBAGARH CHOWKI": "Mohla-Manpur-Ambagarh Chowki",
    "MOHLA-MANPUR-AMBAGARH CHOUKI": "Mohla-Manpur-Ambagarh Chowki",
    "MOHLA MANPUR AMBAGARH CHOUKI": "Mohla-Manpur-Ambagarh Chowki",
    "MOHLA-MANPUR AMBAGARH CHOWKI": "Mohla-Manpur-Ambagarh Chowki",
    "GAURELA-PENDRA-MARWAHI": "Gaurela-Pendra-Marwahi",
    "GAURELA PENDRA MARWAHI": "Gaurela-Pendra-Marwahi",
    "GPM": "Gaurela-Pendra-Marwahi",
    "SARANGARH-BILAIGARH": "Sarangarh-Bilaigarh",
    "SARANGARH BILAIGARH": "Sarangarh-Bilaigarh",
    "MUNGELI": "Mungeli",
    "SUKMA": "Sukma",
    "SAKTI": "Sakti",
})

DISTRICT_PATTERNS = sorted(DISTRICT_ALIASES.keys(), key=len, reverse=True)

# Known meeting-to-quarter mapping (from agenda PDFs)
MEETING_QUARTERS = {
    90:  ('2023-03', 'March 2023',     '31-03-2023'),
    91:  ('2023-06', 'June 2023',      '30-06-2023'),
    92:  ('2023-09', 'September 2023', '30-09-2023'),
    93:  ('2023-12', 'December 2023',  '31-12-2023'),
    94:  ('2024-03', 'March 2024',     '31-03-2024'),
    95:  ('2024-06', 'June 2024',      '30-06-2024'),
    96:  ('2024-09', 'September 2024', '30-09-2024'),
    97:  ('2024-12', 'December 2024',  '31-12-2024'),
    98:  ('2025-03', 'March 2025',     '31-03-2025'),
    99:  ('2025-06', 'June 2025',      '30-06-2025'),
    100: ('2025-09', 'September 2025', '30-09-2025'),
    101: ('2025-12', 'December 2025',  '31-12-2025'),
}


def normalize_district(name):
    """Normalize a district name to canonical form."""
    if not name:
        return None
    name = str(name).strip()
    # Remove serial number prefix
    name = re.sub(r'^\d+[\.\)\s]*', '', name).strip()
    # Remove extra whitespace
    name = re.sub(r'\s+', ' ', name)

    if not name or len(name) < 3:
        return None

    # Skip non-district values
    upper = name.upper()
    if any(skip in upper for skip in [
        'GRAND TOTAL', 'TOTAL', 'STATE', 'BANK', 'S NO', 'SNO',
        'DISTRICT NAME', 'DISTRICT WISE', 'SL NO', 'SLNO'
    ]):
        return None

    # Try exact match
    if name in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[name]
    if upper in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[upper]

    # Try fuzzy: longest prefix match
    for pattern in DISTRICT_PATTERNS:
        if upper == pattern or upper.startswith(pattern + ' ') or upper.startswith(pattern + '-'):
            return DISTRICT_ALIASES[pattern]

    return None


def parse_number(val):
    """Clean a number value from Excel cell."""
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        # Round floats to reasonable precision
        if isinstance(val, float):
            return str(round(val, 4))
        return str(val)
    val = str(val).strip()
    if val in ('', '-', 'NA', 'N/A', 'NIL', 'Nil', 'nil', '--', '---', 'None'):
        return ''
    # Remove commas
    val = val.replace(',', '')
    # Remove percentage sign
    val = val.replace('%', '').strip()
    return val


def to_snake_case(name):
    """Convert a header name to snake_case."""
    if not name:
        return ''
    name = name.replace('%', 'pct').replace('/', '_').replace('-', '_')
    name = name.replace('(', '').replace(')', '').replace('.', '')
    name = name.replace("'", '').replace('"', '').replace(':', '')
    name = name.replace('&', 'and')
    name = re.sub(r'\s+', '_', name)
    name = re.sub(r'_+', '_', name)
    name = name.strip('_').lower()
    if len(name) > 80:
        name = name[:80].rstrip('_')
    return name


# ---- Quarter/date detection ----

def detect_quarter_from_header(text):
    """Detect quarter from header text like 'MARCH 2025', 'DEC. 2024', 'SEPT. 2023'."""
    text = text.upper().strip()

    month_map = {
        'JANUARY': ('01', '31'), 'JAN': ('01', '31'),
        'FEBRUARY': ('02', '28'), 'FEB': ('02', '28'),
        'MARCH': ('03', '31'), 'MAR': ('03', '31'),
        'APRIL': ('04', '30'), 'APR': ('04', '30'),
        'MAY': ('05', '31'),
        'JUNE': ('06', '30'), 'JUN': ('06', '30'),
        'JULY': ('07', '31'), 'JUL': ('07', '31'),
        'AUGUST': ('08', '31'), 'AUG': ('08', '31'),
        'SEPTEMBER': ('09', '30'), 'SEPT': ('09', '30'), 'SEP': ('09', '30'),
        'OCTOBER': ('10', '31'), 'OCT': ('10', '31'),
        'NOVEMBER': ('11', '30'), 'NOV': ('11', '30'),
        'DECEMBER': ('12', '31'), 'DEC': ('12', '31'),
    }

    for month_name, (month_num, last_day) in month_map.items():
        pattern = rf'\b{month_name}\.?\s*[,\s]*(\d{{4}})\b'
        m = re.search(pattern, text)
        if m:
            year = m.group(1)
            return month_num, year, last_day

    return None, None, None


def quarter_from_month(month_num, year):
    """Convert month number and year to quarter key, period name, and as_on_date."""
    month = int(month_num)
    year = int(year)

    month_names = {
        1: 'March', 2: 'March', 3: 'March',
        4: 'June', 5: 'June', 6: 'June',
        7: 'September', 8: 'September', 9: 'September',
        10: 'December', 11: 'December', 12: 'December',
    }
    quarter_months = {
        1: '03', 2: '03', 3: '03',
        4: '06', 5: '06', 6: '06',
        7: '09', 8: '09', 9: '09',
        10: '12', 11: '12', 12: '12',
    }
    quarter_days = {
        '03': '31', '06': '30', '09': '30', '12': '31'
    }

    q_month = quarter_months[month]
    q_name = month_names[month]
    q_day = quarter_days[q_month]

    quarter_key = f"{year}-{q_month}"
    period_name = f"{q_name} {year}"
    as_on_date = f"{q_day}-{q_month}-{year}"

    return quarter_key, period_name, as_on_date


def get_fy(quarter_key):
    """Get financial year from quarter key like '2025-09'."""
    year, month = quarter_key.split('-')
    year = int(year)
    month = int(month)
    if month <= 3:
        return f"{year-1}-{str(year)[2:]}"
    else:
        return f"{year}-{str(year+1)[2:]}"


# ---- Sheet extraction functions ----

def find_district_sheet(wb, patterns):
    """Find a sheet matching any of the given name patterns (case-insensitive)."""
    for sname in wb.sheetnames:
        for pat in patterns:
            if pat.lower() in sname.lower():
                return sname
    return None


def extract_district_rows(ws, district_col=1, data_start_row=None):
    """Extract district rows from a worksheet, returning {district: [row_values]}."""
    districts = {}

    # Find where data starts (first row with a known district)
    if data_start_row is None:
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if len(row) > district_col:
                d = normalize_district(str(row[district_col] or ''))
                if d:
                    data_start_row = i
                    break
        if data_start_row is None:
            return {}

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if len(row) <= district_col:
            continue
        d = normalize_district(str(row[district_col] or ''))
        if not d:
            continue
        if d in districts:
            continue
        # Collect all value columns after the district name column
        values = [parse_number(row[c]) if c < len(row) else '' for c in range(district_col + 1, len(row))]
        districts[d] = values

    return districts


def extract_deposit_advance(wb, meeting_num):
    """Extract district-wise deposits, advances, CD ratio from Table 1(A-1)."""
    # Find the district sheet
    sheet_name = find_district_sheet(wb, ['districtwise', 'district wise', 't1 a-1', 't1a-1'])
    if not sheet_name:
        return None

    ws = wb[sheet_name]

    # Detect quarter from header rows
    quarter_info = None
    for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
        for cell in row[:5]:
            if cell:
                month, year, _ = detect_quarter_from_header(str(cell))
                if month and year:
                    quarter_info = quarter_from_month(month, year)
                    break
        if quarter_info:
            break

    # Extract district data
    districts = extract_district_rows(ws, district_col=1)
    if not districts:
        return None

    # The columns are:
    # Deposits: prev_year, prev_quarter, current, growth_abs, growth_pct
    # Advances: prev_year, prev_quarter, current, growth_abs, growth_pct
    # CD Ratio: prev_year, prev_quarter, current, benchmark, growth_pct
    # Total cols = ~15 (varies slightly)

    # Take only the "current" values for deposits, advances, CD ratio
    # We'll store all columns with descriptive headers
    max_cols = max(len(v) for v in districts.values()) if districts else 0

    # Columns layout (15 cols): Deposits(5) + Advances(5) + CD Ratio(5)
    # Each group: prev_fy, prev_quarter, current, growth_abs, growth_pct
    # CD Ratio last group: prev_fy, prev_quarter, current, benchmark, growth_pct

    # Build standardized output per district using known column positions
    # Standardized fields match other states: total_deposit, total_advance, overall_cd_ratio
    std_headers = [
        'total_deposit', 'total_advance', 'overall_cd_ratio',
        'deposits_prev_fy', 'deposits_prev_quarter',
        'deposits_growth_abs', 'deposits_growth_pct',
        'advances_prev_fy', 'advances_prev_quarter',
        'advances_growth_abs', 'advances_growth_pct',
        'cd_ratio_prev_fy', 'cd_ratio_prev_quarter',
        'cd_ratio_benchmark',
    ]

    std_districts = {}
    for d, vals in districts.items():
        row = {}
        if max_cols >= 15:
            row = {
                'total_deposit': vals[2] if len(vals) > 2 else '',       # current deposits
                'total_advance': vals[7] if len(vals) > 7 else '',       # current advances
                'overall_cd_ratio': vals[12] if len(vals) > 12 else '',  # current CD ratio
                'deposits_prev_fy': vals[0] if len(vals) > 0 else '',
                'deposits_prev_quarter': vals[1] if len(vals) > 1 else '',
                'deposits_growth_abs': vals[3] if len(vals) > 3 else '',
                'deposits_growth_pct': vals[4] if len(vals) > 4 else '',
                'advances_prev_fy': vals[5] if len(vals) > 5 else '',
                'advances_prev_quarter': vals[6] if len(vals) > 6 else '',
                'advances_growth_abs': vals[8] if len(vals) > 8 else '',
                'advances_growth_pct': vals[9] if len(vals) > 9 else '',
                'cd_ratio_prev_fy': vals[10] if len(vals) > 10 else '',
                'cd_ratio_prev_quarter': vals[11] if len(vals) > 11 else '',
                'cd_ratio_benchmark': vals[13] if len(vals) > 13 else '',
            }
        elif max_cols >= 12:
            row = {
                'total_deposit': vals[2] if len(vals) > 2 else '',
                'total_advance': vals[7] if len(vals) > 7 else '',
                'overall_cd_ratio': vals[11] if len(vals) > 11 else '',
                'deposits_prev_fy': vals[0] if len(vals) > 0 else '',
                'deposits_prev_quarter': vals[1] if len(vals) > 1 else '',
                'deposits_growth_abs': vals[3] if len(vals) > 3 else '',
                'deposits_growth_pct': vals[4] if len(vals) > 4 else '',
                'advances_prev_fy': vals[5] if len(vals) > 5 else '',
                'advances_prev_quarter': vals[6] if len(vals) > 6 else '',
                'advances_growth_abs': vals[8] if len(vals) > 8 else '',
                'advances_growth_pct': vals[9] if len(vals) > 9 else '',
                'cd_ratio_prev_fy': vals[10] if len(vals) > 10 else '',
                'cd_ratio_prev_quarter': '',
                'cd_ratio_benchmark': '',
            }
        else:
            # Fallback: best guess
            row = {
                'total_deposit': vals[2] if len(vals) > 2 else '',
                'total_advance': vals[7] if len(vals) > 7 else '' if max_cols > 7 else '',
                'overall_cd_ratio': '',
            }
            for i, v in enumerate(vals):
                row[f'col_{i+1}'] = v
            std_headers = list(row.keys())

        std_districts[d] = row

    return {
        'category': 'credit_deposit_ratio',
        'headers': std_headers,
        'districts': std_districts,
        'quarter_info': quarter_info,
        'is_dict': True,  # Flag: districts values are dicts, not lists
    }


def extract_branch_network(wb, meeting_num):
    """Extract district-wise branch network from Table 1(N-1)."""
    sheet_name = find_district_sheet(wb, ['DISTRIST WISE BR', 'DISTRICT WISE BR'])
    if not sheet_name:
        return None

    ws = wb[sheet_name]
    districts = extract_district_rows(ws, district_col=1)
    if not districts:
        return None

    max_cols = max(len(v) for v in districts.values()) if districts else 0

    # Columns: Rural(5), Semi-Urban(5), Urban(5), Total(5) = 20 cols
    # Each group: prev_fy, prev_q, current, growth_abs, growth_pct
    # Standardized fields: branch_rural, branch_semi_urban, branch_urban, total_branch
    std_headers = [
        'total_branch', 'branch_rural', 'branch_semi_urban', 'branch_urban',
        'branch_rural_prev_fy', 'branch_rural_prev_q', 'branch_rural_growth_abs', 'branch_rural_growth_pct',
        'branch_semi_urban_prev_fy', 'branch_semi_urban_prev_q', 'branch_semi_urban_growth_abs', 'branch_semi_urban_growth_pct',
        'branch_urban_prev_fy', 'branch_urban_prev_q', 'branch_urban_growth_abs', 'branch_urban_growth_pct',
        'branch_total_prev_fy', 'branch_total_prev_q', 'branch_total_growth_abs', 'branch_total_growth_pct',
    ]

    std_districts = {}
    for d, vals in districts.items():
        if max_cols >= 20:
            std_districts[d] = {
                'total_branch': vals[17] if len(vals) > 17 else '',        # Total current
                'branch_rural': vals[2] if len(vals) > 2 else '',          # Rural current
                'branch_semi_urban': vals[7] if len(vals) > 7 else '',     # Semi-Urban current
                'branch_urban': vals[12] if len(vals) > 12 else '',        # Urban current
                'branch_rural_prev_fy': vals[0] if len(vals) > 0 else '',
                'branch_rural_prev_q': vals[1] if len(vals) > 1 else '',
                'branch_rural_growth_abs': vals[3] if len(vals) > 3 else '',
                'branch_rural_growth_pct': vals[4] if len(vals) > 4 else '',
                'branch_semi_urban_prev_fy': vals[5] if len(vals) > 5 else '',
                'branch_semi_urban_prev_q': vals[6] if len(vals) > 6 else '',
                'branch_semi_urban_growth_abs': vals[8] if len(vals) > 8 else '',
                'branch_semi_urban_growth_pct': vals[9] if len(vals) > 9 else '',
                'branch_urban_prev_fy': vals[10] if len(vals) > 10 else '',
                'branch_urban_prev_q': vals[11] if len(vals) > 11 else '',
                'branch_urban_growth_abs': vals[13] if len(vals) > 13 else '',
                'branch_urban_growth_pct': vals[14] if len(vals) > 14 else '',
                'branch_total_prev_fy': vals[15] if len(vals) > 15 else '',
                'branch_total_prev_q': vals[16] if len(vals) > 16 else '',
                'branch_total_growth_abs': vals[18] if len(vals) > 18 else '',
                'branch_total_growth_pct': vals[19] if len(vals) > 19 else '',
            }
        else:
            row = {}
            for i, v in enumerate(vals):
                row[f'col_{i+1}'] = v
            std_districts[d] = row
            std_headers = list(row.keys())

    return {
        'category': 'branch_network',
        'headers': std_headers,
        'districts': std_districts,
        'is_dict': True,
    }


def extract_atm_network(wb, meeting_num):
    """Extract district-wise ATM network from Table 1(O-1)."""
    sheet_name = find_district_sheet(wb, ['DISTRICT WISE ATM'])
    if not sheet_name:
        return None

    ws = wb[sheet_name]
    districts = extract_district_rows(ws, district_col=1)
    if not districts:
        return None

    max_cols = max(len(v) for v in districts.values()) if districts else 0

    # Columns: Rural(5), Semi-Urban(5), Urban(5), Total(5) = 20 cols
    std_headers = [
        'total_atm', 'atm_rural', 'atm_semi_urban', 'atm_urban',
        'atm_rural_prev_fy', 'atm_rural_prev_q', 'atm_rural_growth_abs', 'atm_rural_growth_pct',
        'atm_semi_urban_prev_fy', 'atm_semi_urban_prev_q', 'atm_semi_urban_growth_abs', 'atm_semi_urban_growth_pct',
        'atm_urban_prev_fy', 'atm_urban_prev_q', 'atm_urban_growth_abs', 'atm_urban_growth_pct',
        'atm_total_prev_fy', 'atm_total_prev_q', 'atm_total_growth_abs', 'atm_total_growth_pct',
    ]

    std_districts = {}
    for d, vals in districts.items():
        if max_cols >= 20:
            std_districts[d] = {
                'total_atm': vals[17] if len(vals) > 17 else '',
                'atm_rural': vals[2] if len(vals) > 2 else '',
                'atm_semi_urban': vals[7] if len(vals) > 7 else '',
                'atm_urban': vals[12] if len(vals) > 12 else '',
                'atm_rural_prev_fy': vals[0] if len(vals) > 0 else '',
                'atm_rural_prev_q': vals[1] if len(vals) > 1 else '',
                'atm_rural_growth_abs': vals[3] if len(vals) > 3 else '',
                'atm_rural_growth_pct': vals[4] if len(vals) > 4 else '',
                'atm_semi_urban_prev_fy': vals[5] if len(vals) > 5 else '',
                'atm_semi_urban_prev_q': vals[6] if len(vals) > 6 else '',
                'atm_semi_urban_growth_abs': vals[8] if len(vals) > 8 else '',
                'atm_semi_urban_growth_pct': vals[9] if len(vals) > 9 else '',
                'atm_urban_prev_fy': vals[10] if len(vals) > 10 else '',
                'atm_urban_prev_q': vals[11] if len(vals) > 11 else '',
                'atm_urban_growth_abs': vals[13] if len(vals) > 13 else '',
                'atm_urban_growth_pct': vals[14] if len(vals) > 14 else '',
                'atm_total_prev_fy': vals[15] if len(vals) > 15 else '',
                'atm_total_prev_q': vals[16] if len(vals) > 16 else '',
                'atm_total_growth_abs': vals[18] if len(vals) > 18 else '',
                'atm_total_growth_pct': vals[19] if len(vals) > 19 else '',
            }
        else:
            row = {}
            for i, v in enumerate(vals):
                row[f'col_{i+1}'] = v
            std_districts[d] = row
            std_headers = list(row.keys())

    return {
        'category': 'atm_network',
        'headers': std_headers,
        'districts': std_districts,
        'is_dict': True,
    }


def extract_acp_achievement(wb, meeting_num):
    """Extract district-wise ACP achievement from Table 4(C-1)."""
    sheet_name = find_district_sheet(wb, ['districtwise', 'district wise', 'c-1', 'c1'])
    if not sheet_name:
        return None

    ws = wb[sheet_name]
    districts = extract_district_rows(ws, district_col=1)
    if not districts:
        return None

    max_cols = max(len(v) for v in districts.values()) if districts else 0

    # Standard: AGRL(3: commitment, achievement, %ach), MSME(3), OTHER PSA(3), TOTAL(3) = 12 cols
    # Use "target" instead of "commitment" to match other states
    std_headers = [
        'acp_agri_target', 'acp_agri_achievement', 'acp_agri_pct',
        'acp_msme_target', 'acp_msme_achievement', 'acp_msme_pct',
        'acp_other_psa_target', 'acp_other_psa_achievement', 'acp_other_psa_pct',
        'acp_total_target', 'acp_total_achievement', 'acp_total_pct',
    ]

    std_districts = {}
    for d, vals in districts.items():
        if max_cols >= 12:
            std_districts[d] = {
                'acp_agri_target': vals[0] if len(vals) > 0 else '',
                'acp_agri_achievement': vals[1] if len(vals) > 1 else '',
                'acp_agri_pct': vals[2] if len(vals) > 2 else '',
                'acp_msme_target': vals[3] if len(vals) > 3 else '',
                'acp_msme_achievement': vals[4] if len(vals) > 4 else '',
                'acp_msme_pct': vals[5] if len(vals) > 5 else '',
                'acp_other_psa_target': vals[6] if len(vals) > 6 else '',
                'acp_other_psa_achievement': vals[7] if len(vals) > 7 else '',
                'acp_other_psa_pct': vals[8] if len(vals) > 8 else '',
                'acp_total_target': vals[9] if len(vals) > 9 else '',
                'acp_total_achievement': vals[10] if len(vals) > 10 else '',
                'acp_total_pct': vals[11] if len(vals) > 11 else '',
            }
        else:
            row = {}
            for i, v in enumerate(vals):
                row[f'col_{i+1}'] = v
            std_districts[d] = row
            std_headers = list(row.keys())

    return {
        'category': 'acp_target_achievement',
        'headers': std_headers,
        'districts': std_districts,
        'is_dict': True,
    }


# ---- File finding helpers ----

def find_excel_file(data_dir, patterns):
    """Find an Excel file matching any pattern (case-insensitive) in the data directory."""
    all_files = glob.glob(os.path.join(data_dir, '*.xlsx'))
    for f in all_files:
        basename = os.path.basename(f).lower()
        for pat in patterns:
            if pat.lower() in basename:
                return f
    return None


def process_meeting(meeting_num, data_dir):
    """Process a single meeting's data files and return extracted tables."""
    print(f"\n{'='*60}")
    print(f"Processing Meeting {meeting_num}")
    print(f"{'='*60}")

    # Handle meeting 98 which has a nested TABLE subdirectory
    if os.path.isdir(os.path.join(data_dir, 'TABLE')):
        data_dir = os.path.join(data_dir, 'TABLE')

    results = []
    quarter_info = None

    # 1. Deposit/Advance/CD Ratio
    dep_file = find_excel_file(data_dir, ['deposit advance', 'deposit_advance', 'advance deposit', 'TABLE 1 Advance'])
    if not dep_file:
        # Try broader match: "Table 1" but exclude "Table 11"
        all_xlsx = glob.glob(os.path.join(data_dir, '*.xlsx'))
        for f in all_xlsx:
            bn = os.path.basename(f).lower()
            if re.match(r'^table[ _]1\b', bn) and 'table 11' not in bn and 'table_11' not in bn and 'branch' not in bn and 'atm' not in bn:
                dep_file = f
                break
    if dep_file:
        try:
            wb = openpyxl.load_workbook(dep_file, data_only=True)
            table = extract_deposit_advance(wb, meeting_num)
            wb.close()
            if table:
                quarter_info = table.pop('quarter_info', None)
                results.append(table)
                print(f"  [OK] credit_deposit_ratio: {len(table['districts'])} districts, {len(table['headers'])} cols")
            else:
                print(f"  [SKIP] Deposit/Advance: no district data found")
        except Exception as e:
            print(f"  [ERR] Deposit/Advance: {e}")

    # 2. Branch Network & ATM Network
    br_file = find_excel_file(data_dir, ['branch atm', 'branch_atm'])
    if br_file:
        try:
            wb = openpyxl.load_workbook(br_file, data_only=True)

            # Branch network
            table = extract_branch_network(wb, meeting_num)
            if table:
                results.append(table)
                print(f"  [OK] branch_network: {len(table['districts'])} districts, {len(table['headers'])} cols")
            else:
                print(f"  [SKIP] Branch network: no district data found")

            # ATM network
            table = extract_atm_network(wb, meeting_num)
            if table:
                results.append(table)
                print(f"  [OK] atm_network: {len(table['districts'])} districts, {len(table['headers'])} cols")
            else:
                print(f"  [SKIP] ATM network: no district data found")

            wb.close()
        except Exception as e:
            print(f"  [ERR] Branch/ATM: {e}")

    # 3. ACP Achievement
    acp_file = find_excel_file(data_dir, ['table 4', 'output table 4', 'table_4', 'acpachievement'])
    if acp_file:
        try:
            wb = openpyxl.load_workbook(acp_file, data_only=True)
            table = extract_acp_achievement(wb, meeting_num)
            wb.close()
            if table:
                results.append(table)
                print(f"  [OK] acp_target_achievement: {len(table['districts'])} districts, {len(table['headers'])} cols")
            else:
                print(f"  [SKIP] ACP Achievement: no district data found")
        except Exception as e:
            print(f"  [ERR] ACP Achievement: {e}")

    if not results:
        print(f"  No district-wise data extracted!")
        return None

    # Use hardcoded meeting-to-quarter mapping (most reliable)
    if meeting_num in MEETING_QUARTERS:
        quarter_key, period_name, as_on_date = MEETING_QUARTERS[meeting_num]
    elif quarter_info:
        quarter_key, period_name, as_on_date = quarter_info
    else:
        # Try from filenames as last resort
        for f in glob.glob(os.path.join(data_dir, '*.xlsx')):
            basename = os.path.basename(f).upper()
            month, year, _ = detect_quarter_from_header(basename)
            if month and year:
                quarter_key, period_name, as_on_date = quarter_from_month(month, year)
                break
        else:
            print(f"  WARNING: Could not detect quarter for meeting {meeting_num}")
            return None
    print(f"\n  Quarter: {period_name} (as on {as_on_date})")
    print(f"  Extracted {len(results)} tables")

    return {
        'quarter_key': quarter_key,
        'period': period_name,
        'as_on_date': as_on_date,
        'tables': results,
    }


def save_quarterly_csvs(tables, quarter_key, output_dir):
    """Save extracted tables as quarterly CSVs."""
    quarter_dir = os.path.join(output_dir, 'quarterly', quarter_key)
    os.makedirs(quarter_dir, exist_ok=True)

    saved = 0
    for table in tables:
        category = table['category']
        districts = table['districts']
        headers = table['headers']

        if not districts:
            continue

        csv_path = os.path.join(quarter_dir, f"{category}.csv")
        with open(csv_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['district'] + headers)

            for district in CHHATTISGARH_DISTRICTS:
                if district in districts:
                    ddata = districts[district]
                    if isinstance(ddata, dict):
                        row = [district] + [ddata.get(h, '') for h in headers]
                    else:
                        row = [district] + list(ddata)
                        while len(row) < len(headers) + 1:
                            row.append('')
                        row = row[:len(headers) + 1]
                    writer.writerow(row)

        saved += 1

    print(f"  Saved {saved} CSVs to {quarter_dir}")
    return quarter_dir


def build_complete_json(all_quarters, output_dir):
    """Build and save chhattisgarh_complete.json."""
    complete = {
        "source": "SLBC Chhattisgarh",
        "state": "Chhattisgarh",
        "convenor": "State Bank of India",
        "description": "Complete district-wise banking & financial inclusion data",
        "amount_unit": "Rs. Crore",
        "quarters": {}
    }

    for qk in sorted(all_quarters.keys()):
        qdata = all_quarters[qk]
        fy = get_fy(qk)

        quarter_obj = {
            "period": qdata['period'],
            "as_on_date": qdata['as_on_date'],
            "fy": fy,
            "tables": {}
        }

        for table in qdata['tables']:
            category = table['category']
            districts = table['districts']
            headers = table['headers']

            table_data = {
                "fields": headers,
                "num_districts": len(districts),
                "districts": {}
            }

            for district, values in districts.items():
                if isinstance(values, dict):
                    row = {h: values.get(h, '') for h in headers}
                else:
                    row = {}
                    for i, h in enumerate(headers):
                        row[h] = values[i] if i < len(values) else ''
                table_data["districts"][district] = row

            quarter_obj["tables"][category] = table_data

        complete["quarters"][qk] = quarter_obj

    json_path = os.path.join(output_dir, 'chhattisgarh_complete.json')
    with open(json_path, 'w') as f:
        json.dump(complete, f, indent=2)

    print(f"\nSaved chhattisgarh_complete.json ({len(complete['quarters'])} quarters)")
    return complete


def build_timeseries(complete, output_dir):
    """Build and save chhattisgarh_fi_timeseries.json and .csv."""
    periods = []
    all_fields = set()

    for qk in sorted(complete['quarters'].keys()):
        qdata = complete['quarters'][qk]
        period_name = qdata['period']
        as_on_date = qdata['as_on_date']
        fy = qdata['fy']

        district_records = {}

        for category, table_data in qdata['tables'].items():
            fields = table_data['fields']
            for district, row in table_data['districts'].items():
                if district not in district_records:
                    district_records[district] = {
                        'district': district,
                        'period': period_name,
                    }
                for field_name, value in row.items():
                    key = f"{category}__{field_name}"
                    district_records[district][key] = value
                    all_fields.add(key)

        districts_list = []
        for d in CHHATTISGARH_DISTRICTS:
            if d in district_records:
                districts_list.append(district_records[d])

        if districts_list:
            periods.append({
                'period': period_name,
                'as_on_date': as_on_date,
                'fy': fy,
                'districts': districts_list,
            })

    timeseries = {
        "source": "SLBC Chhattisgarh",
        "state": "Chhattisgarh",
        "convenor": "State Bank of India",
        "description": "Chhattisgarh district-wise financial inclusion timeseries data",
        "num_periods": len(periods),
        "total_records": sum(len(p['districts']) for p in periods),
        "periods": periods,
    }

    # Save JSON
    json_path = os.path.join(output_dir, 'chhattisgarh_fi_timeseries.json')
    with open(json_path, 'w') as f:
        json.dump(timeseries, f, indent=2)
    print(f"Saved chhattisgarh_fi_timeseries.json ({len(periods)} periods, {timeseries['total_records']} records)")

    # Save CSV (wide format)
    sorted_fields = sorted(all_fields)
    csv_path = os.path.join(output_dir, 'chhattisgarh_fi_timeseries.csv')
    with open(csv_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['district', 'period', 'as_on_date', 'fy'] + sorted_fields)

        for period in periods:
            for rec in period['districts']:
                row = [
                    rec.get('district', ''),
                    rec.get('period', ''),
                    period.get('as_on_date', ''),
                    period.get('fy', ''),
                ]
                for field in sorted_fields:
                    row.append(rec.get(field, ''))
                writer.writerow(row)

    print(f"Saved chhattisgarh_fi_timeseries.csv")
    return timeseries


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python extract_chhattisgarh.py --all  (process all meetings in data-tables/)")
        print("  python extract_chhattisgarh.py 90 91 92 ...  (process specific meetings)")
        print("  python extract_chhattisgarh.py --output-dir DIR --all")
        sys.exit(1)

    # Parse args
    meetings = []
    output_dir = '../../public/slbc-data/chhattisgarh'
    data_base = 'data-tables'
    process_all = False

    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == '--output-dir' and i + 1 < len(sys.argv):
            output_dir = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--data-dir' and i + 1 < len(sys.argv):
            data_base = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--all':
            process_all = True
            i += 1
        else:
            try:
                meetings.append(int(sys.argv[i]))
            except ValueError:
                print(f"Warning: ignoring invalid meeting number '{sys.argv[i]}'")
            i += 1

    if process_all:
        # Find all meeting directories
        for d in sorted(glob.glob(os.path.join(data_base, '*'))):
            if os.path.isdir(d):
                try:
                    num = int(os.path.basename(d))
                    meetings.append(num)
                except ValueError:
                    pass

    if not meetings:
        print("No meetings to process!")
        sys.exit(1)

    meetings = sorted(set(meetings))
    print(f"Processing {len(meetings)} meeting(s): {meetings}")
    print(f"Output directory: {output_dir}")
    print(f"Data directory: {data_base}")

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, 'quarterly'), exist_ok=True)

    # Process each meeting
    all_quarters = {}

    for meeting_num in meetings:
        data_dir = os.path.join(data_base, str(meeting_num))
        if not os.path.isdir(data_dir):
            print(f"Warning: {data_dir} not found, skipping meeting {meeting_num}")
            continue

        try:
            result = process_meeting(meeting_num, data_dir)
        except Exception as e:
            print(f"  ERROR processing meeting {meeting_num}: {e}")
            import traceback
            traceback.print_exc()
            result = None

        if result:
            qk = result['quarter_key']

            # Save quarterly CSVs
            save_quarterly_csvs(result['tables'], qk, output_dir)

            if qk in all_quarters:
                # Merge: add new tables, replace existing ones
                existing_cats = {t['category'] for t in all_quarters[qk]['tables']}
                for t in result['tables']:
                    if t['category'] not in existing_cats:
                        all_quarters[qk]['tables'].append(t)
                    else:
                        all_quarters[qk]['tables'] = [
                            x for x in all_quarters[qk]['tables']
                            if x['category'] != t['category']
                        ]
                        all_quarters[qk]['tables'].append(t)
            else:
                all_quarters[qk] = {
                    'period': result['period'],
                    'as_on_date': result['as_on_date'],
                    'tables': result['tables'],
                }

    if not all_quarters:
        print("\nNo data extracted from any meeting!")
        sys.exit(1)

    # Build combined output files
    print(f"\n{'='*60}")
    print("Building combined output files")
    print(f"{'='*60}")

    complete = build_complete_json(all_quarters, output_dir)
    timeseries = build_timeseries(complete, output_dir)

    # Summary
    print(f"\n{'='*60}")
    print("EXTRACTION SUMMARY")
    print(f"{'='*60}")
    print(f"  Meetings processed: {len(meetings)}")
    print(f"  Quarters: {len(all_quarters)}")
    for qk in sorted(all_quarters.keys()):
        qd = all_quarters[qk]
        cats = [t['category'] for t in qd['tables']]
        print(f"    {qk} ({qd['period']}): {len(cats)} tables - {', '.join(cats)}")
    print(f"  Districts: {len(CHHATTISGARH_DISTRICTS)}")
    all_cats = set()
    for qd in all_quarters.values():
        for t in qd['tables']:
            all_cats.add(t['category'])
    print(f"  Total categories: {len(all_cats)}")
    print(f"  Categories: {', '.join(sorted(all_cats))}")
    print(f"\nOutput files:")
    print(f"  {output_dir}/chhattisgarh_complete.json")
    print(f"  {output_dir}/chhattisgarh_fi_timeseries.json")
    print(f"  {output_dir}/chhattisgarh_fi_timeseries.csv")
    print(f"  {output_dir}/quarterly/*/  (per-quarter CSVs)")


if __name__ == '__main__':
    main()
