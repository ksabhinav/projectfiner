#!/usr/bin/env python3
"""Extract district-wise SLBC data for Jammu & Kashmir (UT).

Source: J&K UTLBC convenor (J&K Bank Ltd) — https://www.jkslbc.com/
        annexure xlsx files in /pdf/LeadBankScheme/<year>/<quarter>/

The "ANNEX-N" sheet in each `CD RATIO DISTWISE` XLSX has identical structure
across all quarters we've sampled (Sep 2022 → Dec 2025):

  Row 1/2: title  "DISTRICT-WISE DEPOSITS, ADVANCES, CD RATIO, BRANCHES & GROSS NPA <DD.MM.YYYY>"
  Row 2/3: "AMOUNT IN CRORE"
  Row 3/4: header  # | DISTRICT | NO. OF BRANCHES | DEPOSITS | ADVANCES | CD RATIO | GROSS NPA
  Rows of district data, with two region-subtotal rows ("KASHMIR REGION", "JAMMU REGION") + final "TOTAL".

J&K reports monetary fields in Crores — we convert to Lakhs (×100) before emit,
to match FINER's canonical unit (1 Lakh = ₹100,000; 1 Crore = ₹1,00,00,000 = 100 Lakhs).

Output: jammu-kashmir_complete.json, jammu-kashmir_fi_timeseries.json,
        jammu-kashmir_fi_timeseries.csv, jammu-kashmir_fi_slim.json
        plus a meetings_audit.txt summary.
"""
import csv
import json
import os
import re
import sys
from glob import glob

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

HERE = os.path.abspath(os.path.dirname(__file__))
RAW = os.path.join(HERE, 'raw')
OUT_DIR = HERE  # outputs land here; copied to public/slbc-data/jammu-kashmir/ at the end


# 20 J&K districts (post-2019 reorg). State LGD code = 1.
DISTRICTS_CANONICAL = {
    'SRINAGAR': 'Srinagar', 'GANDERBAL': 'Ganderbal', 'BARAMULLA': 'Baramulla',
    'BANDIPORA': 'Bandipora', 'ANANTNAG': 'Anantnag', 'KULGAM': 'Kulgam',
    'PULWAMA': 'Pulwama', 'SHOPIAN': 'Shopian', 'BUDGAM': 'Budgam',
    'KUPWARA': 'Kupwara',
    'POONCH': 'Poonch', 'RAJOURI': 'Rajouri', 'JAMMU': 'Jammu',
    'SAMBA': 'Samba', 'UDHAMPUR': 'Udhampur', 'REASI': 'Reasi',
    'KATHUA': 'Kathua', 'DODA': 'Doda', 'RAMBAN': 'Ramban',
    'KISHTWAR': 'Kishtwar',
}

DATE_RE = re.compile(r'(\d{2})\.(\d{2})\.(\d{4})')
SOURCE_FILES = {
    # quarter_code: (raw subdir, xlsx filename)
    '2022-09': ('WAYBACK_2022', '7. CD RATIO DISTRICTWISE.xlsx'),
    '2023-09': ('WAYBACK_2023_NEW', '5. CD RATIO DISTRICTWISE_new.xlsx'),
    '2023-12': ('WAYBACK_2023', '5. CD RATIO DISTRICTWISE.xlsx'),
    '2024-03': ('WAYBACK_2024', '7. CD RATIO DISTRICTWISE.xlsx'),
    '2025-03': ('WAYBACK_2025_MAR', 'ATTACHMENT-4 CD RATIO DIST.xlsx'),
    '2025-12': ('DEC2025', '07. CD RATIO DISTWISE DEC25.xlsx'),
}
QUARTER_LABEL = {
    '2022-09': 'September 2022',
    '2023-09': 'September 2023',
    '2023-12': 'December 2023',
    '2024-03': 'March 2024',
    '2025-03': 'March 2025',
    '2025-12': 'December 2025',
}


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '')
    if not s or s.upper() in ('NA', 'N/A', '-', '#REF!', '#DIV/0!'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def find_header_row(ws):
    """Locate the row containing 'DISTRICT' & 'DEPOSITS' headers; returns 1-based row."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        cells = [str(c).upper().strip() if c is not None else '' for c in row]
        if 'DISTRICT' in cells and 'DEPOSITS' in cells and 'ADVANCES' in cells:
            return i, cells
    return None, None


def parse_xlsx(path: str) -> dict:
    """Parse a CD RATIO DISTWISE XLSX → {district -> {fields}} (monetary in Lakhs)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr_row, hdr_cells = find_header_row(ws)
    if hdr_row is None:
        raise ValueError(f"No data header found in {path}")
    col = {h: hdr_cells.index(h) for h in
           ('DISTRICT', 'NO. OF BRANCHES', 'DEPOSITS', 'ADVANCES', 'CD RATIO', 'GROSS NPA')
           if h in hdr_cells}
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= hdr_row:
            continue
        name_raw = row[col['DISTRICT']] if 'DISTRICT' in col else None
        if not name_raw:
            continue
        name_up = str(name_raw).strip().upper()
        # Skip subtotal rows
        if name_up in ('KASHMIR REGION', 'JAMMU REGION', 'TOTAL', 'GRAND TOTAL'):
            continue
        canonical = DISTRICTS_CANONICAL.get(name_up)
        if not canonical:
            continue  # unknown row
        deposits_cr = num(row[col['DEPOSITS']])
        advances_cr = num(row[col['ADVANCES']])
        cdr = num(row[col['CD RATIO']])
        branches = num(row[col['NO. OF BRANCHES']])
        gross_npa_cr = num(row[col['GROSS NPA']]) if 'GROSS NPA' in col and col['GROSS NPA'] < len(row) else None
        # Convert Crores → Lakhs (×100). cd_ratio is a percent (unitless).
        rec = {}
        if deposits_cr is not None:
            rec['credit_deposit_ratio__total_deposit'] = round(deposits_cr * 100, 2)
        if advances_cr is not None:
            rec['credit_deposit_ratio__total_advance'] = round(advances_cr * 100, 2)
        if cdr is not None:
            rec['credit_deposit_ratio__cd_ratio'] = round(cdr, 2)
        if branches is not None:
            rec['branch_network__total_branch'] = int(branches)
        if gross_npa_cr is not None:
            rec['credit_deposit_ratio__gross_npa'] = round(gross_npa_cr * 100, 2)
        out[canonical] = rec
    return out


def main():
    audit_lines = ['# J&K UTLBC — meetings/quarter audit\n']
    audit_lines.append(f'Source portal: https://www.jkslbc.com/  (UTLBC convenor: J&K Bank Ltd)\n')
    audit_lines.append(f'Live page lists only the current quarter; historical CD-ratio annexures retrieved via Wayback Machine.\n')

    complete = {'state': 'jammu-kashmir', 'state_lgd_code': 1, 'quarters': {}}
    timeseries = {'state': 'jammu-kashmir', 'periods': []}

    for qcode in sorted(SOURCE_FILES.keys()):
        subdir, fname = SOURCE_FILES[qcode]
        path = os.path.join(RAW, subdir, fname)
        if not os.path.exists(path):
            audit_lines.append(f'  [MISSING] {qcode} → {path}')
            continue
        try:
            district_rows = parse_xlsx(path)
        except Exception as e:
            audit_lines.append(f'  [FAIL] {qcode}: {e}')
            continue
        audit_lines.append(f'  {qcode}  ({QUARTER_LABEL[qcode]}) — {len(district_rows)} districts — {subdir}/{fname}')

        qkey = QUARTER_LABEL[qcode].lower().replace(' ', '_')
        # Build complete-style tables
        tables = {'credit_deposit_ratio': {'fields': [], 'districts': {}},
                  'branch_network':       {'fields': [], 'districts': {}}}
        for dname, rec in district_rows.items():
            cdr_rec = {k.split('__', 1)[1]: v for k, v in rec.items() if k.startswith('credit_deposit_ratio__')}
            bn_rec  = {k.split('__', 1)[1]: v for k, v in rec.items() if k.startswith('branch_network__')}
            if cdr_rec:
                tables['credit_deposit_ratio']['districts'][dname] = cdr_rec
                for f in cdr_rec:
                    if f not in tables['credit_deposit_ratio']['fields']:
                        tables['credit_deposit_ratio']['fields'].append(f)
            if bn_rec:
                tables['branch_network']['districts'][dname] = bn_rec
                for f in bn_rec:
                    if f not in tables['branch_network']['fields']:
                        tables['branch_network']['fields'].append(f)
        complete['quarters'][qkey] = {'period': QUARTER_LABEL[qcode], 'tables': tables}

        # Build timeseries (nested: periods → districts list with flat field keys)
        period_obj = {'period': QUARTER_LABEL[qcode], 'districts': []}
        for dname, rec in sorted(district_rows.items()):
            entry = {'district': dname, 'period': QUARTER_LABEL[qcode]}
            entry.update(rec)
            period_obj['districts'].append(entry)
        timeseries['periods'].append(period_obj)

    # Sort periods newest-first
    timeseries['periods'].sort(key=lambda p: p['period'], reverse=False)

    # Write outputs
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, 'jammu-kashmir_complete.json'), 'w') as f:
        json.dump(complete, f, indent=2)
    with open(os.path.join(OUT_DIR, 'jammu-kashmir_fi_timeseries.json'), 'w') as f:
        json.dump(timeseries, f, indent=2)
    # Slim is the same shape as the timeseries for J&K (we only carry the 2 indicator
    # categories that exist for J&K, both already in the 7-category whitelist).
    with open(os.path.join(OUT_DIR, 'jammu-kashmir_fi_slim.json'), 'w') as f:
        json.dump(timeseries, f, indent=2)

    # CSV: wide format
    all_field_keys = set()
    for period in timeseries['periods']:
        for d in period['districts']:
            for k in d:
                if k not in ('district', 'period'):
                    all_field_keys.add(k)
    field_order = sorted(all_field_keys)
    with open(os.path.join(OUT_DIR, 'jammu-kashmir_fi_timeseries.csv'), 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['district', 'period'] + field_order)
        for period in timeseries['periods']:
            for d in period['districts']:
                w.writerow([d['district'], d['period']] + [d.get(k, '') for k in field_order])

    with open(os.path.join(OUT_DIR, 'meetings_audit.txt'), 'w') as f:
        f.write('\n'.join(audit_lines) + '\n')
        f.write('\nNote: J&K SLBC reports amounts in Crores; extractor converts to Lakhs (×100)\n')
        f.write('to match FINER canonical unit. cd_ratio is left as a percent (unitless).\n')
        f.write('Categories captured: credit_deposit_ratio + branch_network (CD-ratio annexure only).\n')
        f.write('PMJDY/KCC/SHG/digital coverage tables exist in J&K UTLBC agendas but only as\n')
        f.write('bank-wise summaries (not district-wise) so they are not present in this extract.\n')

    print(f'Wrote outputs to {OUT_DIR}/')
    print('\n'.join(audit_lines))


if __name__ == '__main__':
    main()
