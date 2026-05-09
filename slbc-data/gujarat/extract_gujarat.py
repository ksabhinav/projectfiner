"""
Extract ALL district-level FI indicator data from Gujarat SLBC annexure Excel files.
Produces gujarat_complete.json and gujarat_fi_timeseries.json.

Source: https://slbcgujarat.in/slbc_meettings_held/
Meetings 179 (Sep 2023) through 188 (Dec 2025).

Indicators extracted:
  - CD Ratio (Annex1)
  - District-wise Banking Parameter (Annex3/21) — deposits, advances, NRI, agri, MSME, PSA, weaker sections
  - PMJDY (Annex A/1) — district-wise only available in meeting 188
  - Aadhaar/CASA seeding (Annex B/2)
  - Social Security / PMJJBY+PMSBY (Annex E/C/3) — district-wise enrolment
  - KCC (Annex20/49) — Kisan Credit Card district-wise
  - PMEGP (Annex8/31) — district-wise
  - ATM (Annex22/5) — district-wise
  - Branch (Annex23/6) — district-wise
  - SACP (Annex5/16) — Annual Credit Plan district-wise (multiple sub-categories)
  - PMMY/Mudra (Annex J/K/22) — district-wise (only meeting 188)
  - RSETI (Annex I/K/14) — district-wise
  - SARFAESI (Annex7B/52)
"""

import json
import os
import re
import warnings
import openpyxl

warnings.filterwarnings("ignore", category=UserWarning)

BASE = os.path.dirname(os.path.abspath(__file__))

# Gujarat district canonical names
GUJARAT_DISTRICTS = {
    "AHMEDABAD", "AHMADABAD", "AMRELI", "ANAND", "ARAVALLI", "ARVALLI", "ARAVALI",
    "BANASKANTHA", "BANAS KANTHA", "BHARUCH", "BHAVNAGAR", "BOTAD",
    "CHHOTAUDEPUR", "CHHOTA UDEPUR", "DAHOD", "DANG", "DANGS", "THE DANGS",
    "DEVBHUMI DWARKA", "DEVBHOOMI DWARKA", "GANDHINAGAR", "GIR SOMNATH",
    "JAMNAGAR", "JUNAGADH", "KHEDA", "KUTCH", "KACHCHH",
    "MAHISAGAR", "MAHESANA", "MEHSANA", "MORBI",
    "NARMADA", "NAVSARI", "PANCHMAHAL", "PANCHMAHALS", "PANCH MAHAL",
    "PATAN", "PORBANDAR", "RAJKOT", "SABARKANTHA", "SABAR KANTHA",
    "SURAT", "SURENDRANAGAR", "TAPI",
    "VADODARA", "VALSAD",
}

# Normalize district name mapping
DISTRICT_ALIASES = {
    "AHMADABAD": "Ahmedabad",
    "AHMEDABAD": "Ahmedabad",
    "ARVALLI": "Aravalli",
    "ARAVALI": "Aravalli",
    "ARAVALLI": "Aravalli",
    "BANAS KANTHA": "Banaskantha",
    "BANASKANTHA": "Banaskantha",
    "BANASAKNTHA": "Banaskantha",
    "CHHOTA UDEPUR": "Chhotaudepur",
    "CHHOTAUDEPUR": "Chhotaudepur",
    "CHHOTA  UDEPUR": "Chhotaudepur",
    "CHOTAUDEPUR": "Chhotaudepur",
    "DANG": "Dang",
    "DANGS": "Dang",
    "THE DANGS": "Dang",
    "DEVBHUMI DWARKA": "Devbhumi Dwarka",
    "DEVBHOOMI DWARKA": "Devbhumi Dwarka",
    "GIR SOMNATH": "Gir Somnath",
    "GIRSOMNATH": "Gir Somnath",
    "KUTCH": "Kutch",
    "KACHCHH": "Kutch",
    "MAHESANA": "Mehsana",
    "MEHSANA": "Mehsana",
    "PANCHMAHALS": "Panchmahal",
    "PANCH MAHAL": "Panchmahal",
    "PANCHMAHAL": "Panchmahal",
    "PANCH  MAHAL": "Panchmahal",
    "SABAR KANTHA": "Sabarkantha",
    "SABARKANTHA": "Sabarkantha",
    "SURENDRANAGAR": "Surendranagar",
    "SURENDRA NAGAR": "Surendranagar",
}


def normalize_district(name):
    """Normalize district name to canonical form."""
    if not name or not isinstance(name, str):
        return None
    name = name.strip().replace("\xa0", " ")
    # Remove leading serial numbers like "1", "2.", etc.
    name = re.sub(r"^\d+[\.\)\s]+", "", name).strip()
    if not name or len(name) < 3:
        return None

    upper = name.upper().strip()

    # Skip non-district entries
    skip_words = [
        "TOTAL", "GRAND", "STATE", "GUJARAT", "SOURCE", "DATA", "NO.", "SR.",
        "BANK", "DISTRICT", "AMOUNT", "NOTE", "PAGE", "ANNEXURE", "RS.",
        "LAKH", "CRORE", "NATIONAL", "PRIVATE", "RRB", "SFB", "COOPERATIVE",
        "ALL INDIA", "PARTICULARS"
    ]
    for sw in skip_words:
        if upper.startswith(sw):
            return None
    if "TOTAL" in upper:
        return None

    # Check alias mapping
    if upper in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[upper]

    # Check if it's a known Gujarat district
    for d in GUJARAT_DISTRICTS:
        if upper == d:
            return name.title()

    # Try fuzzy match - if it looks like a district
    # Only return if it roughly matches a known district
    for d in GUJARAT_DISTRICTS:
        if d in upper or upper in d:
            return DISTRICT_ALIASES.get(d, d.title())

    return None


def parse_number(val):
    """Convert Excel cell value to clean number string."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val != val:  # NaN check
            return None
        return str(round(val, 2))
    s = str(val).strip().replace(",", "").replace("\xa0", "")
    if s.startswith("=") or s == "#N/A" or s == "#VALUE!" or s == "#REF!" or s == "-" or s == "":
        return None
    try:
        return str(round(float(s), 2))
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """Safe string conversion."""
    if val is None:
        return ""
    return str(val).strip()


# ============ QUARTER METADATA ============

MEETING_QUARTERS = {
    179: {"quarter_key": "2023-09", "period": "September 2023", "as_on_date": "30-09-2023", "fy": "2023-24"},
    180: {"quarter_key": "2023-12", "period": "December 2023", "as_on_date": "31-12-2023", "fy": "2023-24"},
    181: {"quarter_key": "2024-03", "period": "March 2024", "as_on_date": "31-03-2024", "fy": "2023-24"},
    182: {"quarter_key": "2024-06", "period": "June 2024", "as_on_date": "30-06-2024", "fy": "2024-25"},
    183: {"quarter_key": "2024-09", "period": "September 2024", "as_on_date": "30-09-2024", "fy": "2024-25"},
    184: {"quarter_key": "2024-12", "period": "December 2024", "as_on_date": "31-12-2024", "fy": "2024-25"},
    185: {"quarter_key": "2025-03", "period": "March 2025", "as_on_date": "31-03-2025", "fy": "2024-25"},
    186: {"quarter_key": "2025-06", "period": "June 2025", "as_on_date": "30-06-2025", "fy": "2025-26"},
    187: {"quarter_key": "2025-09", "period": "September 2025", "as_on_date": "30-09-2025", "fy": "2025-26"},
    188: {"quarter_key": "2025-12", "period": "December 2025", "as_on_date": "31-12-2025", "fy": "2025-26"},
}


def find_annex_dir(meeting_num):
    """Find the annexure subdirectory for a meeting."""
    dirs = [d for d in os.listdir(BASE) if d.startswith(f"meeting{meeting_num}_") and os.path.isdir(os.path.join(BASE, d))]
    if not dirs:
        return None
    mdir = os.path.join(BASE, dirs[0])
    for sd in sorted(os.listdir(mdir)):
        sdp = os.path.join(mdir, sd)
        if os.path.isdir(sdp):
            return sdp
    return None


def find_file(annex_dir, patterns):
    """Find a file matching any of the given patterns (case-insensitive partial match)."""
    if not annex_dir:
        return None
    files = os.listdir(annex_dir)
    for pat in patterns:
        pat_lower = pat.lower()
        for f in files:
            if pat_lower in f.lower() and (f.endswith(".xlsx") or f.endswith(".xls")):
                return os.path.join(annex_dir, f)
    return None


def load_wb(filepath):
    """Load workbook safely."""
    if not filepath or not os.path.exists(filepath):
        return None
    try:
        return openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ERROR loading {filepath}: {e}")
        return None


def find_sheet(wb, patterns):
    """Find a sheet matching any of the given patterns."""
    if not wb:
        return None
    for pat in patterns:
        pat_lower = pat.lower()
        for sname in wb.sheetnames:
            if pat_lower in sname.lower():
                return wb[sname]
    return None


def get_all_rows(ws, min_row=1, max_col=None):
    """Get all rows from a worksheet."""
    rows = []
    for row in ws.iter_rows(min_row=min_row, max_col=max_col, values_only=True):
        rows.append(list(row))
    return rows


# ============ EXTRACTORS ============

def extract_cd_ratio(annex_dir, meeting_num):
    """Extract CD Ratio from Annex1/19 CD Ratio Districtwise."""
    fp = find_file(annex_dir, ["cd ratio districtwise"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = wb[wb.sheetnames[0]]
    districts = {}

    # Detect format
    row3_vals = []
    for row in ws.iter_rows(min_row=3, max_row=3, max_col=12, values_only=True):
        row3_vals = list(row)

    is_old = False
    if row3_vals and len(row3_vals) > 2:
        cell3 = safe_str(row3_vals[2])
        if "branch" in cell3.lower():
            is_old = True

    if is_old:
        # Old format: Sr, District, Branches, Deposits, Advances, CD Ratio
        for row in ws.iter_rows(min_row=6, max_col=6, values_only=True):
            sr, district, branches, deposits, advances, cd_ratio = row
            district = normalize_district(safe_str(district))
            if not district:
                continue
            dep = parse_number(deposits)
            adv = parse_number(advances)
            br = parse_number(branches)
            if dep and adv and float(dep) > 0:
                cdr = str(round(float(adv) / float(dep) * 100, 2))
            else:
                cdr = parse_number(cd_ratio)
            if dep or adv:
                d = {"total_deposit": dep or "0", "total_advances": adv or "0", "overall_cd_ratio": cdr or "0"}
                if br:
                    d["total_branch"] = br
                districts[district] = d
    else:
        # New format: find latest deposit/advances columns
        row3 = []
        row4 = []
        for row in ws.iter_rows(min_row=3, max_row=3, max_col=20, values_only=True):
            row3 = list(row)
        for row in ws.iter_rows(min_row=4, max_row=4, max_col=20, values_only=True):
            row4 = list(row)

        # Find the LAST (rightmost) "Total Deposit" header
        dep_col = None
        for i, cell in enumerate(row4):
            if cell and isinstance(cell, str) and "total deposit" in cell.lower():
                dep_col = i
        if dep_col is None:
            for i, cell in enumerate(row3):
                if cell and isinstance(cell, str) and "total deposit" in cell.lower():
                    dep_col = i
        if dep_col is None:
            dep_col = 4  # fallback

        adv_col = dep_col + 1
        cdr_col = dep_col + 2

        for row in ws.iter_rows(min_row=6, max_col=max(cdr_col + 1, 12), values_only=True):
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            dep = parse_number(row[dep_col]) if len(row) > dep_col else None
            adv = parse_number(row[adv_col]) if len(row) > adv_col else None
            if dep and adv and float(dep) > 0:
                cdr = str(round(float(adv) / float(dep) * 100, 2))
            else:
                cdr = parse_number(row[cdr_col]) if len(row) > cdr_col else None
            if dep or adv:
                districts[district] = {
                    "total_deposit": dep or "0",
                    "total_advances": adv or "0",
                    "overall_cd_ratio": cdr or "0",
                }

    wb.close()
    return districts


def extract_banking_parameter(annex_dir, meeting_num):
    """Extract Annex3/21 Districtwise Banking Parameter (Page 1 + Page 2)."""
    fp = find_file(annex_dir, ["districtwise banking parameter"])
    wb = load_wb(fp)
    if not wb:
        return {}

    districts = {}

    # Page 1: NRI Deposit, Agri Adv (A/c, Amt), %Agri, MSME (A/c, Amt), Other PS (A/c, Amt),
    #          Total Priority (A/c, Amt), Total Adv, %PSA
    ws1 = find_sheet(wb, ["page 1"])
    if ws1:
        # Detect if multi-period (meetings 187-188 have multiple date columns)
        rows = get_all_rows(ws1, min_row=1, max_col=32)
        header_row = None
        data_start = None

        # Find header row with "District" and data start
        for i, row in enumerate(rows):
            for cell in row:
                s = safe_str(cell).lower()
                if s == "district" or s == "district ":
                    header_row = i
                    break
            if header_row is not None:
                break

        if header_row is not None:
            # Data starts 2-3 rows after header
            data_start = header_row + 2

            # Check if it's multi-period format (meeting 187, 188)
            # In multi-period, we need the LATEST period columns
            # For meetings 187-188: look for "DECEMBER 2025" or last set of columns
            is_multi = meeting_num >= 187

            if is_multi:
                # Multi-period: find column offsets
                # Row structure: No, District, Total_Adv_prev, current_period_data...
                # Current period data: Agri(A/c,Amt), MSME(A/c,Amt), OtherPS(A/c,Amt), TotalPriority(A/c,Amt), %Agri..., %PSA...
                for row in ws1.iter_rows(min_row=data_start + 1, max_col=32, values_only=True):
                    district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
                    if not district:
                        continue
                    d = districts.setdefault(district, {})
                    # Col 2 = total adv prev period, cols 3-10 are current period
                    d["agri_adv_ac"] = parse_number(row[3]) if len(row) > 3 else None
                    d["agri_adv_amt"] = parse_number(row[4]) if len(row) > 4 else None
                    d["msme_ac"] = parse_number(row[5]) if len(row) > 5 else None
                    d["msme_amt"] = parse_number(row[6]) if len(row) > 6 else None
                    d["other_ps_ac"] = parse_number(row[7]) if len(row) > 7 else None
                    d["other_ps_amt"] = parse_number(row[8]) if len(row) > 8 else None
                    d["total_priority_ac"] = parse_number(row[9]) if len(row) > 9 else None
                    d["total_priority_amt"] = parse_number(row[10]) if len(row) > 10 else None
            else:
                # Old format: No, District, NRI Deposit, Agri(A/c, Amt), %Agri, MSME(A/c, Amt), OtherPS(A/c, Amt), TotalPriority(A/c, Amt), TotalAdv, %PSA
                for row in ws1.iter_rows(min_row=data_start + 1, max_col=25, values_only=True):
                    district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
                    if not district:
                        continue
                    d = districts.setdefault(district, {})
                    d["nri_deposit"] = parse_number(row[2]) if len(row) > 2 else None
                    d["agri_adv_ac"] = parse_number(row[3]) if len(row) > 3 else None
                    d["agri_adv_amt"] = parse_number(row[4]) if len(row) > 4 else None
                    d["pct_agri_adv"] = parse_number(row[5]) if len(row) > 5 else None
                    d["msme_ac"] = parse_number(row[6]) if len(row) > 6 else None
                    d["msme_amt"] = parse_number(row[7]) if len(row) > 7 else None
                    d["other_ps_ac"] = parse_number(row[8]) if len(row) > 8 else None
                    d["other_ps_amt"] = parse_number(row[9]) if len(row) > 9 else None
                    d["total_priority_ac"] = parse_number(row[10]) if len(row) > 10 else None
                    d["total_priority_amt"] = parse_number(row[11]) if len(row) > 11 else None
                    d["total_advances"] = parse_number(row[12]) if len(row) > 12 else None
                    d["pct_psa"] = parse_number(row[13]) if len(row) > 13 else None

    # Page 2: Weaker sections - Small/Marginal Farmers, SC, ST, DRI, SHG, Women, Minority
    ws2 = find_sheet(wb, ["page 2"])
    if ws2:
        # Find data start
        data_start = 5
        for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=10, max_col=3, values_only=True), 1):
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if district:
                data_start = i
                break

        for row in ws2.iter_rows(min_row=data_start, max_col=25, values_only=True):
            # In meetings 187-188, col 2 may be "Total Adv prev period"
            # Try col 1 first
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            d = districts.setdefault(district, {})

            # Columns vary slightly but general pattern:
            # Small/Marginal Farmers (A/c, Amt), SC (A/c, Amt), ST (A/c, Amt), DRI (A/c, Amt), SHG (A/c, Amt), Women (A/c, Amt), Minority (A/c, Amt)
            offset = 2
            if meeting_num >= 187:
                offset = 3  # extra "total adv" column
            idx = offset
            pairs = [
                ("small_marginal_farmers_ac", "small_marginal_farmers_amt"),
                ("sc_ac", "sc_amt"),
                ("st_ac", "st_amt"),
                ("dri_ac", "dri_amt"),
                ("shg_ac", "shg_amt"),
                ("women_ac", "women_amt"),
                ("minority_ac", "minority_amt"),
            ]
            for ac_key, amt_key in pairs:
                if len(row) > idx:
                    d[ac_key] = parse_number(row[idx])
                if len(row) > idx + 1:
                    d[amt_key] = parse_number(row[idx + 1])
                idx += 2

    wb.close()
    # Remove None values
    for dist in districts:
        districts[dist] = {k: v for k, v in districts[dist].items() if v is not None}
    return districts


def extract_pmjdy_dw(annex_dir, meeting_num):
    """Extract PMJDY district-wise data. Only meeting 188 has DW sheet."""
    # Meeting 188 has "Annex 1 PMJDY.xlsx" with DW sheet
    if meeting_num != 188:
        return {}

    fp = find_file(annex_dir, ["pmjdy"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["dw", "district"])
    if not ws:
        wb.close()
        return {}

    districts = {}
    # Headers in row 3: Sr, District, Total(prev periods...), Total(current), Progress, Rural, Urban,
    #                    Total Amount, Aadhaar Seeded, %Aadhaar, Zero Balance, %Zero, RuPay, %RuPay
    for row in ws.iter_rows(min_row=4, max_col=16, values_only=True):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if not district:
            continue
        # For meeting 188: cols are 0=Sr, 1=District, 2-4=prev totals, 5=current total, 6=progress,
        # 7=rural, 8=urban, 9=amount, 10=aadhaar, 11=%aadhaar, 12=zero_bal, 13=%zero, 14=rupay, 15=%rupay
        d = {}
        # Use the latest total (col 5 for 188)
        d["total_pmjdy_ac"] = parse_number(row[5]) if len(row) > 5 else None
        d["pmjdy_rural"] = parse_number(row[7]) if len(row) > 7 else None
        d["pmjdy_urban"] = parse_number(row[8]) if len(row) > 8 else None
        d["pmjdy_amount"] = parse_number(row[9]) if len(row) > 9 else None
        d["pmjdy_aadhaar_seeded"] = parse_number(row[10]) if len(row) > 10 else None
        d["pmjdy_pct_aadhaar"] = parse_number(row[11]) if len(row) > 11 else None
        d["pmjdy_zero_balance"] = parse_number(row[12]) if len(row) > 12 else None
        d["pmjdy_pct_zero_balance"] = parse_number(row[13]) if len(row) > 13 else None
        d["pmjdy_rupay_issued"] = parse_number(row[14]) if len(row) > 14 else None
        d["pmjdy_pct_rupay"] = parse_number(row[15]) if len(row) > 15 else None

        districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_aadhaar_casa(annex_dir, meeting_num):
    """Extract Aadhaar/CASA seeding district-wise data."""
    fp = find_file(annex_dir, ["aadhaar seeding", "casa aadhaar"])
    wb = load_wb(fp)
    if not wb:
        return {}

    districts = {}

    # Meeting 186 has special sheets: "District Wise Master CASA" and "District Wise Master Mobile See"
    ws_casa = find_sheet(wb, ["district wise master casa"])
    ws_mobile = find_sheet(wb, ["district wise master mobile"])

    if ws_casa:
        # Format: State, District, CASA_actual, CASA_lakhs, Aadhaar_actual, Aadhaar_lakhs, %Aadhaar, Auth, %Auth
        for row in ws_casa.iter_rows(min_row=3, max_col=12, values_only=True):
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            d = districts.setdefault(district, {})
            d["operative_casa"] = parse_number(row[2]) if len(row) > 2 else None  # actual number
            d["aadhaar_seeded_casa"] = parse_number(row[4]) if len(row) > 4 else None
            d["pct_aadhaar_seeding"] = parse_number(row[6]) if len(row) > 6 else None
            d["authenticated_casa"] = parse_number(row[7]) if len(row) > 7 else None
            d["pct_authentication"] = parse_number(row[8]) if len(row) > 8 else None

        if ws_mobile:
            # Format: State, District, SB_acs, Mobile_seeded, Not_seeded, %Mobile
            for row in ws_mobile.iter_rows(min_row=3, max_col=8, values_only=True):
                district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
                if not district:
                    continue
                d = districts.setdefault(district, {})
                d["operative_sb"] = parse_number(row[2]) if len(row) > 2 else None
                d["mobile_seeded_sb"] = parse_number(row[3]) if len(row) > 3 else None
                d["pct_mobile_seeding"] = parse_number(row[5]) if len(row) > 5 else None

        # Remove None values
        for dist in districts:
            districts[dist] = {k: v for k, v in districts[dist].items() if v is not None}
        wb.close()
        return districts

    # Look for DW sheet (meeting 188)
    ws = find_sheet(wb, ["dw", "district"])
    if not ws:
        wb.close()
        return {}

    rows = get_all_rows(ws, min_row=1, max_col=17)
    # Find data start (first row with a district)
    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    for row in rows[data_start:]:
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if not district:
            continue

        d = {}
        if meeting_num == 188:
            # Cols: Sr, District, CASA_prev, CASA_current, Aadhaar, %Aadhaar_prev, %Aadhaar_current,
            #        Auth, %Auth_prev, %Auth_current, SB_prev, SB_current, Mobile, %Mobile_prev, %Mobile_current
            d["operative_casa"] = parse_number(row[3]) if len(row) > 3 else None
            d["aadhaar_seeded_casa"] = parse_number(row[4]) if len(row) > 4 else None
            d["pct_aadhaar_seeding"] = parse_number(row[6]) if len(row) > 6 else None
            d["authenticated_casa"] = parse_number(row[7]) if len(row) > 7 else None
            d["pct_authentication"] = parse_number(row[9]) if len(row) > 9 else None
            d["operative_sb"] = parse_number(row[11]) if len(row) > 11 else None
            d["mobile_seeded_sb"] = parse_number(row[12]) if len(row) > 12 else None
            d["pct_mobile_seeding"] = parse_number(row[14]) if len(row) > 14 else None
        else:
            # Generic: Sr, District, CASA, Aadhaar, %Aadhaar, Auth, %Auth, SB, Mobile, %Mobile
            d["operative_casa"] = parse_number(row[2]) if len(row) > 2 else None
            d["aadhaar_seeded_casa"] = parse_number(row[3]) if len(row) > 3 else None
            d["pct_aadhaar_seeding"] = parse_number(row[4]) if len(row) > 4 else None
            d["authenticated_casa"] = parse_number(row[5]) if len(row) > 5 else None
            d["pct_authentication"] = parse_number(row[6]) if len(row) > 6 else None
            d["operative_sb"] = parse_number(row[7]) if len(row) > 7 else None
            d["mobile_seeded_sb"] = parse_number(row[8]) if len(row) > 8 else None
            d["pct_mobile_seeding"] = parse_number(row[9]) if len(row) > 9 else None

        districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_pmjjby_pmsby(annex_dir, meeting_num):
    """Extract PMJJBY/PMSBY district-wise enrolment."""
    # Different file names across meetings
    if meeting_num == 188:
        fp = find_file(annex_dir, ["social security scheme"])
    elif meeting_num in [182, 185, 186]:
        # These may not have PMJJBY DW (no Annex E)
        fp = find_file(annex_dir, ["pmjjby pmsby", "pmjjby"])
        if not fp:
            fp = find_file(annex_dir, ["social security scheme"])
    else:
        fp = find_file(annex_dir, ["pmjjby pmsby", "pmjjby"])

    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["dw", "district"])
    if not ws:
        wb.close()
        return {}

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=15)

    # Find data start
    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    for row in rows[data_start:]:
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if not district:
            continue

        d = {}
        if meeting_num == 188:
            # Meeting 188 Social Security DW:
            # Sr, District, PMJJBY(prev periods + current + progress), PMSBY(prev + current + progress)
            # Cols: 0=Sr, 1=Dist, 2-5=PMJJBY(dates), 6=progress, 7-10=PMSBY(dates), 11=progress
            d["pmjjby_enrolment"] = parse_number(row[5]) if len(row) > 5 else None
            d["pmsby_enrolment"] = parse_number(row[10]) if len(row) > 10 else None
        elif meeting_num >= 183:
            # Older DW format: Sr, District, PMJJBY_target, PMJJBY_achievement, %PMJJBY, PMSBY_target, PMSBY_achievement, %PMSBY
            d["pmjjby_target"] = parse_number(row[2]) if len(row) > 2 else None
            d["pmjjby_achievement"] = parse_number(row[3]) if len(row) > 3 else None
            d["pmjjby_pct"] = parse_number(row[4]) if len(row) > 4 else None
            d["pmsby_target"] = parse_number(row[5]) if len(row) > 5 else None
            d["pmsby_achievement"] = parse_number(row[6]) if len(row) > 6 else None
            d["pmsby_pct"] = parse_number(row[7]) if len(row) > 7 else None
        else:
            d["pmjjby_target"] = parse_number(row[2]) if len(row) > 2 else None
            d["pmjjby_achievement"] = parse_number(row[3]) if len(row) > 3 else None
            d["pmjjby_pct"] = parse_number(row[4]) if len(row) > 4 else None
            d["pmsby_target"] = parse_number(row[5]) if len(row) > 5 else None
            d["pmsby_achievement"] = parse_number(row[6]) if len(row) > 6 else None
            d["pmsby_pct"] = parse_number(row[7]) if len(row) > 7 else None

        districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_kcc(annex_dir, meeting_num):
    """Extract KCC district-wise data."""
    fp = find_file(annex_dir, ["kcc"])
    # Meeting 180 also has separate "KCC DW.xlsx"
    if meeting_num == 180:
        fp2 = find_file(annex_dir, ["kcc dw"])
        if fp2:
            fp = fp2

    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["district", "kisan_credit"])
    if not ws:
        wb.close()
        return {}

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=21)

    # Find data start
    data_start = None
    for i, row in enumerate(rows):
        for ci in [1, 0]:
            district = normalize_district(safe_str(row[ci]) if len(row) > ci else "")
            if district:
                data_start = i
                break
        if data_start is not None:
            break

    if data_start is None:
        wb.close()
        return {}

    # Detect format
    # Meeting 180 "KCC DW.xlsx" has different format:
    #   SR, District, Issued_Qtr_Ac, Issued_Qtr_Amt, Total_KCC_Ac, OS_Amt, RuPay, Activated, Issued_FY_Ac, Issued_FY_Amt
    # Meetings 181+ have multi-category (AH, Fisheries, Total):
    #   No, District, AH(prev_ac, prev_amt, issued_ac, issued_amt, current_ac, current_amt),
    #                  Fish(same), Total(same)

    is_kcc_dw_format = False
    for row in rows[:10]:
        for cell in row:
            s = safe_str(cell).lower()
            if "kcc issued during" in s or "no. of kcc issued" in s:
                is_kcc_dw_format = True
                break

    for row in rows[data_start:]:
        district_col = 1
        district = normalize_district(safe_str(row[district_col]) if len(row) > district_col else "")
        if not district:
            continue

        d = {}
        if is_kcc_dw_format:
            # Simple KCC DW format (meeting 180)
            d["kcc_issued_qtr_ac"] = parse_number(row[2]) if len(row) > 2 else None
            d["kcc_issued_qtr_amt"] = parse_number(row[3]) if len(row) > 3 else None
            d["total_kcc_ac"] = parse_number(row[4]) if len(row) > 4 else None
            d["kcc_os_amt"] = parse_number(row[5]) if len(row) > 5 else None
            d["kcc_rupay"] = parse_number(row[6]) if len(row) > 6 else None
            d["kcc_activated"] = parse_number(row[7]) if len(row) > 7 else None
        else:
            # Multi-category format
            # Total KCC section starts at col 14 (0-indexed)
            # Total: prev_ac(14), prev_amt(15), issued_ac(16), issued_amt(17), current_ac(18), current_amt(19)
            d["kcc_ah_os_ac"] = parse_number(row[6]) if len(row) > 6 else None
            d["kcc_ah_os_amt"] = parse_number(row[7]) if len(row) > 7 else None
            d["kcc_fish_os_ac"] = parse_number(row[12]) if len(row) > 12 else None
            d["kcc_fish_os_amt"] = parse_number(row[13]) if len(row) > 13 else None
            d["total_kcc_ac"] = parse_number(row[18]) if len(row) > 18 else None
            d["total_kcc_amt"] = parse_number(row[19]) if len(row) > 19 else None
            d["kcc_issued_qtr_ac"] = parse_number(row[16]) if len(row) > 16 else None
            d["kcc_issued_qtr_amt"] = parse_number(row[17]) if len(row) > 17 else None

        districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_atm(annex_dir, meeting_num):
    """Extract ATM district-wise data."""
    fp = find_file(annex_dir, ["atm districtwise", "atm district"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["atm"])
    if not ws:
        ws = wb[wb.sheetnames[0]]

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=16)

    # Find data start
    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    # Meetings 187-188 have multi-period format
    if meeting_num >= 187:
        # Cols: No, District, Rural(prev periods), SemiUrban(prev periods), Urban(prev periods), Total(prev periods), NewATMs
        # For 188: 3 date cols each: March2025, Sept2025, Dec2025
        # Use the LAST date col for each category
        for row in rows[data_start:]:
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            d = {}
            # Dec 2025 is cols 4, 7, 10, 13 for meeting 188; cols 3, 5, 7, 9 for meeting 187
            if meeting_num == 188:
                d["atm_rural"] = parse_number(row[4]) if len(row) > 4 else None
                d["atm_semi_urban"] = parse_number(row[7]) if len(row) > 7 else None
                d["atm_urban"] = parse_number(row[10]) if len(row) > 10 else None
                d["atm_total"] = parse_number(row[13]) if len(row) > 13 else None
                d["atm_new_fy"] = parse_number(row[14]) if len(row) > 14 else None
            else:
                d["atm_rural"] = parse_number(row[3]) if len(row) > 3 else None
                d["atm_semi_urban"] = parse_number(row[5]) if len(row) > 5 else None
                d["atm_urban"] = parse_number(row[7]) if len(row) > 7 else None
                d["atm_total"] = parse_number(row[9]) if len(row) > 9 else None
                d["atm_new_fy"] = parse_number(row[10]) if len(row) > 10 else None
            districts[district] = {k: v for k, v in d.items() if v is not None}
    else:
        # Old format: No, District, Rural, Semi Urban, Urban, Total
        for row in rows[data_start:]:
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            d = {}
            d["atm_rural"] = parse_number(row[2]) if len(row) > 2 else None
            d["atm_semi_urban"] = parse_number(row[3]) if len(row) > 3 else None
            d["atm_urban"] = parse_number(row[4]) if len(row) > 4 else None
            d["atm_total"] = parse_number(row[5]) if len(row) > 5 else None
            districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_branch(annex_dir, meeting_num):
    """Extract Branch district-wise data."""
    fp = find_file(annex_dir, ["branch"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["district"])
    if not ws:
        wb.close()
        return {}

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=16)

    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    if meeting_num >= 187:
        # Multi-period layout: 3 date cols per category (Mar, Jun/Sep, Sep/Dec)
        # Rural(cols 2-4), SemiUrban(cols 5-7), Urban(cols 8-10), Total(cols 11-13), NewBranches(col 14)
        # We want the LATEST date column (rightmost in each group): col 4, 7, 10, 13
        for row in rows[data_start:]:
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            d = {}
            d["branch_rural"] = parse_number(row[4]) if len(row) > 4 else None
            d["branch_semi_urban"] = parse_number(row[7]) if len(row) > 7 else None
            d["branch_urban"] = parse_number(row[10]) if len(row) > 10 else None
            d["total_branch"] = parse_number(row[13]) if len(row) > 13 else None
            d["branch_new_fy"] = parse_number(row[14]) if len(row) > 14 else None
            districts[district] = {k: v for k, v in d.items() if v is not None}
    else:
        # Old format: No, District, Rural, Semi Urban, Urban, Total
        for row in rows[data_start:]:
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            d = {}
            d["branch_rural"] = parse_number(row[2]) if len(row) > 2 else None
            d["branch_semi_urban"] = parse_number(row[3]) if len(row) > 3 else None
            d["branch_urban"] = parse_number(row[4]) if len(row) > 4 else None
            d["total_branch"] = parse_number(row[5]) if len(row) > 5 else None
            districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_pmegp(annex_dir, meeting_num):
    """Extract PMEGP district-wise data."""
    fp = find_file(annex_dir, ["pmegp"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["dist", "pmegp dist"])
    if not ws:
        wb.close()
        return {}

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=12)

    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    # Format: Sr, District, Target_No, Target_MM, Forwarded_No, Forwarded_MM, Sanctioned_No, Sanctioned_MM, Rejected_No, Rejected_MM, Pending_No, Pending_MM
    for row in rows[data_start:]:
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if not district:
            continue
        d = {}
        d["pmegp_target_no"] = parse_number(row[2]) if len(row) > 2 else None
        d["pmegp_target_mm"] = parse_number(row[3]) if len(row) > 3 else None
        d["pmegp_forwarded_no"] = parse_number(row[4]) if len(row) > 4 else None
        d["pmegp_forwarded_mm"] = parse_number(row[5]) if len(row) > 5 else None
        d["pmegp_sanctioned_no"] = parse_number(row[6]) if len(row) > 6 else None
        d["pmegp_sanctioned_mm"] = parse_number(row[7]) if len(row) > 7 else None
        d["pmegp_rejected_no"] = parse_number(row[8]) if len(row) > 8 else None
        d["pmegp_rejected_mm"] = parse_number(row[9]) if len(row) > 9 else None
        d["pmegp_pending_no"] = parse_number(row[10]) if len(row) > 10 else None
        d["pmegp_pending_mm"] = parse_number(row[11]) if len(row) > 11 else None
        districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_sacp(annex_dir, meeting_num):
    """Extract SACP (Annual Credit Plan) district-wise data - multiple sub-categories."""
    fp = find_file(annex_dir, ["sacp districtwise", "sacp district"])
    wb = load_wb(fp)
    if not wb:
        return {}

    districts = {}

    # SACP has multiple sheets: ACP (total PS), CROP, TERM, Agri_Infra, Total Agri, Total MSME, Edu_PS, Housing_PS, T Other PS
    sheet_mapping = {
        "ACP": "sacp_total_ps",
        "CROP": "sacp_crop_loan",
        "TERM": "sacp_term_loan",
        "Agri_Infra": "sacp_agri_infra",
        "Total Agri": "sacp_total_agri",
        "Total MSME": "sacp_total_msme",
        "Edu_PS": "sacp_education",
        "Housing_PS": "sacp_housing",
        "T Other PS": "sacp_other_ps",
    }

    for sheet_key, prefix in sheet_mapping.items():
        ws = find_sheet(wb, [sheet_key])
        if not ws:
            continue

        rows = get_all_rows(ws, min_row=1, max_col=10)
        data_start = None
        for i, row in enumerate(rows):
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if district:
                data_start = i
                break

        if data_start is None:
            continue

        # Format: No, District, Target_Ac, Target_Amt, Disb_Ac, Disb_Amt, %Ach_Ac, %Ach_Amt, OS_Ac, OS_Amt
        for row in rows[data_start:]:
            district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not district:
                continue
            d = districts.setdefault(district, {})
            d[f"{prefix}_target_ac"] = parse_number(row[2]) if len(row) > 2 else None
            d[f"{prefix}_target_amt"] = parse_number(row[3]) if len(row) > 3 else None
            d[f"{prefix}_disb_ac"] = parse_number(row[4]) if len(row) > 4 else None
            d[f"{prefix}_disb_amt"] = parse_number(row[5]) if len(row) > 5 else None
            d[f"{prefix}_pct_ach_ac"] = parse_number(row[6]) if len(row) > 6 else None
            d[f"{prefix}_pct_ach_amt"] = parse_number(row[7]) if len(row) > 7 else None
            d[f"{prefix}_os_ac"] = parse_number(row[8]) if len(row) > 8 else None
            d[f"{prefix}_os_amt"] = parse_number(row[9]) if len(row) > 9 else None

    wb.close()
    for dist in districts:
        districts[dist] = {k: v for k, v in districts[dist].items() if v is not None}
    return districts


def extract_rseti(annex_dir, meeting_num):
    """Extract RSETI district-wise data."""
    fp = find_file(annex_dir, ["rseti"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["annx", "rseti", "sheet1"])
    if not ws:
        ws = wb[wb.sheetnames[0]]

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=20)

    # Find data start - RSETI has district in col 2 (0-indexed), with lead bank in col 0
    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[2]) if len(row) > 2 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    # Format: LeadBank, Sr, District, Target_Programs, Target_Candidates,
    #         Programs_Qtr, Cumulative_Programs, Beneficiaries_Qtr, Cumulative_Beneficiaries,
    #         Settled(Bank, Own, Wage, Total), %Settlement, SC, ...
    for row in rows[data_start:]:
        district = normalize_district(safe_str(row[2]) if len(row) > 2 else "")
        if not district:
            continue
        d = {}
        d["rseti_target_programs"] = parse_number(row[3]) if len(row) > 3 else None
        d["rseti_target_candidates"] = parse_number(row[4]) if len(row) > 4 else None
        d["rseti_programs_qtr"] = parse_number(row[5]) if len(row) > 5 else None
        d["rseti_cumulative_programs"] = parse_number(row[6]) if len(row) > 6 else None
        d["rseti_beneficiaries_qtr"] = parse_number(row[7]) if len(row) > 7 else None
        d["rseti_cumulative_beneficiaries"] = parse_number(row[8]) if len(row) > 8 else None
        d["rseti_settled_bank"] = parse_number(row[9]) if len(row) > 9 else None
        d["rseti_settled_own"] = parse_number(row[10]) if len(row) > 10 else None
        d["rseti_settled_wage"] = parse_number(row[11]) if len(row) > 11 else None
        d["rseti_settled_total"] = parse_number(row[12]) if len(row) > 12 else None
        d["rseti_settlement_pct"] = parse_number(row[13]) if len(row) > 13 else None
        districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


def extract_sarfaesi(annex_dir, meeting_num):
    """Extract SARFAESI district-wise data."""
    fp = find_file(annex_dir, ["sarfaesi"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["districtwise", "district", "cmm"])
    if not ws:
        wb.close()
        return {}

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=15)

    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    for row in rows[data_start:]:
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if not district:
            continue
        d = {}
        # Extract available numeric cols
        for ci in range(2, min(len(row), 12)):
            val = parse_number(row[ci])
            if val:
                d[f"sarfaesi_col{ci}"] = val
        if d:
            districts[district] = d

    wb.close()
    return districts


def extract_pmmy_dw(annex_dir, meeting_num):
    """Extract PMMY (Mudra) district-wise data. Only meeting 188 has DW sheet."""
    fp = find_file(annex_dir, ["pmmy"])
    wb = load_wb(fp)
    if not wb:
        return {}

    ws = find_sheet(wb, ["dw", "district"])
    if not ws:
        wb.close()
        return {}

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=20)

    data_start = None
    for i, row in enumerate(rows):
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if district:
            data_start = i
            break

    if data_start is None:
        wb.close()
        return {}

    # Format (meeting 188): Sr, District, Target, PrevTotal(Ac,Amt), CurrentTotal(Ac,Amt),
    #   Shishu(Ac,Amt), Kishore(Ac,Amt), Tarun(Ac,Amt), TarunPlus(Ac,Amt), Total(Ac,Amt)
    for row in rows[data_start:]:
        district = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if not district:
            continue
        d = {}
        d["pmmy_target"] = parse_number(row[2]) if len(row) > 2 else None
        # Current total
        d["pmmy_total_ac"] = parse_number(row[5]) if len(row) > 5 else None
        d["pmmy_total_amt"] = parse_number(row[6]) if len(row) > 6 else None
        d["pmmy_shishu_ac"] = parse_number(row[7]) if len(row) > 7 else None
        d["pmmy_shishu_amt"] = parse_number(row[8]) if len(row) > 8 else None
        d["pmmy_kishore_ac"] = parse_number(row[9]) if len(row) > 9 else None
        d["pmmy_kishore_amt"] = parse_number(row[10]) if len(row) > 10 else None
        d["pmmy_tarun_ac"] = parse_number(row[11]) if len(row) > 11 else None
        d["pmmy_tarun_amt"] = parse_number(row[12]) if len(row) > 12 else None
        d["pmmy_tarun_plus_ac"] = parse_number(row[13]) if len(row) > 13 else None
        d["pmmy_tarun_plus_amt"] = parse_number(row[14]) if len(row) > 14 else None
        districts[district] = {k: v for k, v in d.items() if v is not None}

    wb.close()
    return districts


# ============ MAIN EXTRACTION ============

def extract_meeting(meeting_num):
    """Extract all district-wise data from a single meeting."""
    annex_dir = find_annex_dir(meeting_num)
    if not annex_dir:
        print(f"  Meeting {meeting_num}: directory not found")
        return {}

    print(f"\n=== Meeting {meeting_num} ({MEETING_QUARTERS[meeting_num]['period']}) ===")
    print(f"  Annexure dir: {annex_dir}")

    all_data = {}  # district -> category -> fields

    # 1. CD Ratio
    cd = extract_cd_ratio(annex_dir, meeting_num)
    print(f"  CD Ratio: {len(cd)} districts")
    for dist, data in cd.items():
        all_data.setdefault(dist, {})["credit_deposit_ratio"] = data

    # 2. Banking Parameter (Annex3/21)
    bp = extract_banking_parameter(annex_dir, meeting_num)
    print(f"  Banking Parameter: {len(bp)} districts")
    for dist, data in bp.items():
        all_data.setdefault(dist, {})["banking_parameter"] = data

    # 3. PMJDY (DW only in meeting 188)
    pmjdy = extract_pmjdy_dw(annex_dir, meeting_num)
    if pmjdy:
        print(f"  PMJDY: {len(pmjdy)} districts")
        for dist, data in pmjdy.items():
            all_data.setdefault(dist, {})["pmjdy"] = data

    # 4. Aadhaar/CASA
    aadhaar = extract_aadhaar_casa(annex_dir, meeting_num)
    print(f"  Aadhaar/CASA: {len(aadhaar)} districts")
    for dist, data in aadhaar.items():
        all_data.setdefault(dist, {})["aadhaar_authentication"] = data

    # 5. PMJJBY/PMSBY
    pmjjby = extract_pmjjby_pmsby(annex_dir, meeting_num)
    print(f"  PMJJBY/PMSBY: {len(pmjjby)} districts")
    for dist, data in pmjjby.items():
        all_data.setdefault(dist, {})["social_security_schemes"] = data

    # 6. KCC
    kcc = extract_kcc(annex_dir, meeting_num)
    print(f"  KCC: {len(kcc)} districts")
    for dist, data in kcc.items():
        all_data.setdefault(dist, {})["kcc"] = data

    # 7. ATM
    atm = extract_atm(annex_dir, meeting_num)
    print(f"  ATM: {len(atm)} districts")
    for dist, data in atm.items():
        all_data.setdefault(dist, {})["atm"] = data

    # 8. Branch
    branch = extract_branch(annex_dir, meeting_num)
    print(f"  Branch: {len(branch)} districts")
    for dist, data in branch.items():
        all_data.setdefault(dist, {})["branch_network"] = data

    # 9. PMEGP
    pmegp = extract_pmegp(annex_dir, meeting_num)
    print(f"  PMEGP: {len(pmegp)} districts")
    for dist, data in pmegp.items():
        all_data.setdefault(dist, {})["pmegp"] = data

    # 10. SACP
    sacp = extract_sacp(annex_dir, meeting_num)
    print(f"  SACP: {len(sacp)} districts")
    for dist, data in sacp.items():
        all_data.setdefault(dist, {})["sacp"] = data

    # 11. RSETI
    rseti = extract_rseti(annex_dir, meeting_num)
    print(f"  RSETI: {len(rseti)} districts")
    for dist, data in rseti.items():
        all_data.setdefault(dist, {})["rseti"] = data

    # 12. SARFAESI
    sarfaesi = extract_sarfaesi(annex_dir, meeting_num)
    print(f"  SARFAESI: {len(sarfaesi)} districts")
    for dist, data in sarfaesi.items():
        all_data.setdefault(dist, {})["sarfaesi"] = data

    # 13. PMMY (DW only in meeting 188)
    pmmy = extract_pmmy_dw(annex_dir, meeting_num)
    if pmmy:
        print(f"  PMMY: {len(pmmy)} districts")
        for dist, data in pmmy.items():
            all_data.setdefault(dist, {})["pmmy"] = data

    print(f"  TOTAL: {len(all_data)} districts, {sum(len(cats) for cats in all_data.values())} category entries")
    return all_data


def build_complete_json(all_quarters):
    """Build gujarat_complete.json."""
    result = {
        "source": "SLBC Gujarat (State Level Bankers' Committee, Gujarat)",
        "state": "Gujarat",
        "amount_unit": "Rs. Lakhs",
        "quarters": {},
    }

    for q in all_quarters:
        tables = {}
        for dist, categories in q["data"].items():
            for cat_name, cat_data in categories.items():
                if cat_name not in tables:
                    tables[cat_name] = {"fields": set(), "districts": {}}
                tables[cat_name]["districts"][dist] = cat_data
                tables[cat_name]["fields"].update(cat_data.keys())

        # Convert field sets to sorted lists
        for cat_name in tables:
            tables[cat_name]["fields"] = sorted(tables[cat_name]["fields"])

        result["quarters"][q["quarter_key"]] = {
            "period": q["period"],
            "as_on_date": q["as_on_date"],
            "fy": q["fy"],
            "tables": tables,
        }

    return result


def build_timeseries_json(all_quarters):
    """Build gujarat_fi_timeseries.json."""
    periods = []

    for q in all_quarters:
        district_rows = []
        for dist, categories in q["data"].items():
            row = {
                "district": dist,
                "period": q["period"],
            }
            for cat_name, cat_data in categories.items():
                for field, value in cat_data.items():
                    row[f"{cat_name}__{field}"] = value
            district_rows.append(row)

        periods.append({
            "period": q["period"],
            "districts": district_rows,
        })

    return {"periods": periods}


def main():
    os.chdir(BASE)

    all_quarters = []

    for meeting_num in sorted(MEETING_QUARTERS.keys()):
        qinfo = MEETING_QUARTERS[meeting_num]
        data = extract_meeting(meeting_num)

        if data:
            all_quarters.append({
                "quarter_key": qinfo["quarter_key"],
                "period": qinfo["period"],
                "as_on_date": qinfo["as_on_date"],
                "fy": qinfo["fy"],
                "data": data,
            })

    all_quarters.sort(key=lambda q: q["quarter_key"])

    # Build outputs
    complete = build_complete_json(all_quarters)
    timeseries = build_timeseries_json(all_quarters)

    with open("gujarat_complete.json", "w") as f:
        json.dump(complete, f, indent=2, ensure_ascii=False)
    print(f"\nWrote gujarat_complete.json ({len(all_quarters)} quarters)")

    with open("gujarat_fi_timeseries.json", "w") as f:
        json.dump(timeseries, f, indent=2, ensure_ascii=False)
    print(f"Wrote gujarat_fi_timeseries.json")

    # Summary
    print("\n=== SUMMARY ===")
    for q in all_quarters:
        cats = set()
        for dist, categories in q["data"].items():
            cats.update(categories.keys())
        print(f"  {q['period']}: {len(q['data'])} districts, categories: {sorted(cats)}")


if __name__ == "__main__":
    main()
