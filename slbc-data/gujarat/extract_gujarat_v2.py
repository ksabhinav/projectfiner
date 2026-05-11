"""
Gujarat SLBC Data Extractor v2
================================

Extends extract_gujarat.py to cover older meetings 153-178 (March 2017 -> June 2023)
in addition to the existing 179-188 (Sept 2023 -> Dec 2025).

Source: https://slbcgujarat.in/slbc_meettings_held/

For 179-188 the existing v1 extractor logic is delegated to.
For 153-178 we parse new annexure naming patterns from the data-tables/ unzipped tree:
  - Annex1DW / Annex1 CD Ratio Districtwise / "Annex 1 Districtwise CD Ratio"  -> credit_deposit_ratio
  - Annex3 Page1_ 2 Districtwise Banking Parameter                              -> banking_parameter
  - Annex5 SACP Districtwise (multi-sheet)                                      -> sacp
  - Annex23/24/26/27/31 Branch / Annex 31 Branch / Annex 30 DistrictWise AtM    -> branch_network, atm

Older meetings (128-149, .xls BIFF format) are not extracted in this pass — they
have district-wise data scattered across non-standard sheets and the format
varies meeting-to-meeting; downloading them is in place so a future pass can
target them with xlrd.
"""

import json
import os
import re
import sys
import warnings
from copy import deepcopy

import openpyxl

import extract_gujarat as v1
from extract_gujarat import (
    BASE,
    MEETING_QUARTERS as V1_MEETING_QUARTERS,
    normalize_district,
    parse_number,
    safe_str,
    extract_cd_ratio as v1_extract_cd_ratio,
    extract_banking_parameter as v1_extract_banking_parameter,
    extract_pmjdy_dw as v1_extract_pmjdy_dw,
    extract_aadhaar_casa as v1_extract_aadhaar_casa,
    extract_pmjjby_pmsby as v1_extract_pmjjby_pmsby,
    extract_kcc as v1_extract_kcc,
    extract_atm as v1_extract_atm,
    extract_branch as v1_extract_branch,
    extract_pmegp as v1_extract_pmegp,
    extract_sacp as v1_extract_sacp,
    extract_rseti as v1_extract_rseti,
    extract_sarfaesi as v1_extract_sarfaesi,
    extract_pmmy_dw as v1_extract_pmmy_dw,
    find_annex_dir as v1_find_annex_dir,
    find_file,
    load_wb,
    find_sheet,
    get_all_rows,
    build_complete_json,
    build_timeseries_json,
)

warnings.filterwarnings("ignore", category=UserWarning)


# Older meetings: each ZIP unzips into data-tables/<meeting>/<sub>/
# We need a different annex-dir resolver because the parent dir is meeting num
# not "meetingN_period".
DATA_TABLES = os.path.join(BASE, "data-tables")


# Older meeting -> quarter mapping (extension of v1)
MEETING_QUARTERS_OLD = {
    153: {"quarter_key": "2017-03", "period": "March 2017",      "as_on_date": "31-03-2017", "fy": "2016-17"},
    154: {"quarter_key": "2017-06", "period": "June 2017",       "as_on_date": "30-06-2017", "fy": "2017-18"},
    155: {"quarter_key": "2017-09", "period": "September 2017",  "as_on_date": "30-09-2017", "fy": "2017-18"},
    156: {"quarter_key": "2017-12", "period": "December 2017",   "as_on_date": "31-12-2017", "fy": "2017-18"},
    157: {"quarter_key": "2018-03", "period": "March 2018",      "as_on_date": "31-03-2018", "fy": "2017-18"},
    158: {"quarter_key": "2018-06", "period": "June 2018",       "as_on_date": "30-06-2018", "fy": "2018-19"},
    159: {"quarter_key": "2018-09", "period": "September 2018",  "as_on_date": "30-09-2018", "fy": "2018-19"},
    160: {"quarter_key": "2018-12", "period": "December 2018",   "as_on_date": "31-12-2018", "fy": "2018-19"},
    161: {"quarter_key": "2019-03", "period": "March 2019",      "as_on_date": "31-03-2019", "fy": "2018-19"},
    162: {"quarter_key": "2019-06", "period": "June 2019",       "as_on_date": "30-06-2019", "fy": "2019-20"},
    163: {"quarter_key": "2019-09", "period": "September 2019",  "as_on_date": "30-09-2019", "fy": "2019-20"},
    164: {"quarter_key": "2019-12", "period": "December 2019",   "as_on_date": "31-12-2019", "fy": "2019-20"},
    165: {"quarter_key": "2020-03", "period": "March 2020",      "as_on_date": "31-03-2020", "fy": "2019-20"},
    166: {"quarter_key": "2020-06", "period": "June 2020",       "as_on_date": "30-06-2020", "fy": "2020-21"},
    167: {"quarter_key": "2020-09", "period": "September 2020",  "as_on_date": "30-09-2020", "fy": "2020-21"},
    168: {"quarter_key": "2020-12", "period": "December 2020",   "as_on_date": "31-12-2020", "fy": "2020-21"},
    169: {"quarter_key": "2021-03", "period": "March 2021",      "as_on_date": "31-03-2021", "fy": "2020-21"},
    170: {"quarter_key": "2021-06", "period": "June 2021",       "as_on_date": "30-06-2021", "fy": "2021-22"},
    171: {"quarter_key": "2021-09", "period": "September 2021",  "as_on_date": "30-09-2021", "fy": "2021-22"},
    172: {"quarter_key": "2021-12", "period": "December 2021",   "as_on_date": "31-12-2021", "fy": "2021-22"},
    173: {"quarter_key": "2022-03", "period": "March 2022",      "as_on_date": "31-03-2022", "fy": "2021-22"},
    174: {"quarter_key": "2022-06", "period": "June 2022",       "as_on_date": "30-06-2022", "fy": "2022-23"},
    175: {"quarter_key": "2022-09", "period": "September 2022",  "as_on_date": "30-09-2022", "fy": "2022-23"},
    176: {"quarter_key": "2022-12", "period": "December 2022",   "as_on_date": "31-12-2022", "fy": "2022-23"},
    177: {"quarter_key": "2023-03", "period": "March 2023",      "as_on_date": "31-03-2023", "fy": "2022-23"},
    178: {"quarter_key": "2023-06", "period": "June 2023",       "as_on_date": "30-06-2023", "fy": "2023-24"},
}

ALL_MEETING_QUARTERS = {**MEETING_QUARTERS_OLD, **V1_MEETING_QUARTERS}


# ----------------------------------------------------------------------
# Older-meeting helpers
# ----------------------------------------------------------------------

def find_old_annex_dir(meeting_num):
    """Resolve data-tables/<meeting>/...  walking nested folders to find the
    directory that actually holds the .xlsx/.xls annexure files.

    Returns the deepest directory containing at least one .xlsx file (preferring
    the dir that holds an 'Annex1' / 'CD Ratio' file). Falls back to the
    largest subdir or the meeting root.
    """
    mdir = os.path.join(DATA_TABLES, str(meeting_num))
    if not os.path.isdir(mdir):
        return None

    # Walk and score directories by # of district-wise candidate files
    candidates = []
    for root, dirs, files in os.walk(mdir):
        n_xlsx = sum(1 for f in files if f.lower().endswith(('.xlsx', '.xls')))
        if n_xlsx == 0:
            continue
        score = n_xlsx
        # Boost if there's an Annex1 or CD Ratio file here
        for f in files:
            fl = f.lower()
            if 'annex1' in fl or 'annex 1' in fl or 'cd ratio' in fl or 'districtwise' in fl:
                score += 20
                break
        candidates.append((root, score, n_xlsx))

    if not candidates:
        return mdir

    candidates.sort(key=lambda x: -x[1])
    return candidates[0][0]


# ----------------------------------------------------------------------
# Extractors for older meeting layouts (153-178)
# ----------------------------------------------------------------------

def _find_file_any(annex_dir, patterns_list):
    """Try each pattern tuple in order; return first match."""
    if not annex_dir or not os.path.isdir(annex_dir):
        return None
    files = os.listdir(annex_dir)
    for substrs in patterns_list:
        substrs_lower = [s.lower() for s in substrs]
        for f in files:
            fl = f.lower()
            if not (fl.endswith('.xlsx') or fl.endswith('.xls')):
                continue
            if all(s in fl for s in substrs_lower):
                return os.path.join(annex_dir, f)
    return None


def extract_cd_ratio_old(annex_dir, meeting_num):
    """Annex1DW / Annex1 CD Ratio Districtwise / Annex 1 Districtwise CD Ratio.

    Header at row 3: No. | District | No. of Branches | Total Deposits |
    Total Advances | CD Ratio. Data starts at row 6.
    """
    fp = _find_file_any(annex_dir, [
        ('annex1dw',),
        ('annex 1dw',),
        ('annex 1 districtwise cd ratio',),
        ('annex1 cd ratio districtwise',),
        ('annex 1 cd ratio districtwise',),
        ('cd ratio districtwise',),
        ('districtwise cd ratio',),
        ('cd ratio district wise',),     # M162, M163
        ('cd ratio dist wise',),          # M161, M168
        ('cd ratio distwise',),           # M169
        ('distwise cd ratio',),           # M166: "Annex 1 Distwise CD ratio"
        ('annex 1 dist',),
    ])
    if not fp:
        return {}

    if fp.lower().endswith('.xls'):
        return {}  # skip BIFF format for now

    wb = load_wb(fp)
    if not wb:
        return {}
    ws = wb[wb.sheetnames[0]]
    districts = {}
    for row in ws.iter_rows(min_row=6, max_col=6, values_only=True):
        if not row or not row[1]:
            continue
        district = normalize_district(safe_str(row[1]))
        if not district:
            continue
        br = parse_number(row[2]) if len(row) > 2 else None
        dep = parse_number(row[3]) if len(row) > 3 else None
        adv = parse_number(row[4]) if len(row) > 4 else None
        cd_raw = parse_number(row[5]) if len(row) > 5 else None
        if dep and adv and float(dep) > 0:
            cdr = str(round(float(adv) / float(dep) * 100, 2))
        else:
            cdr = cd_raw
        if dep or adv:
            d = {"total_deposit": dep or "0", "total_advances": adv or "0", "overall_cd_ratio": cdr or "0"}
            if br:
                d["total_branch"] = br
            districts[district] = d
    wb.close()
    return districts


def extract_banking_parameter_old(annex_dir, meeting_num):
    """Annex3 Page1_ 2 Districtwise Banking Parameter (xlsx).

    Page 1: NRI Deposit | Agri (A/c, Amt) | %Agri | MSME (A/c, Amt) | Other PS
    (A/c, Amt) | Total Priority (A/c, Amt) | Total Adv | %PSA
    Page 2: Weaker sections (small/marginal farmers, SC, ST, DRI, SHG, Women,
    Minority — each as A/c + Amt pair).

    Layout matches v1's "Old format" branch; the data rows can start at row 7-8
    depending on header decoration.
    """
    fp = _find_file_any(annex_dir, [
        ('districtwise banking parameter',),
        ('district wise banking parameter',),
        ('annex3 page1', 'districtwise'),
        ('annex 3 page', 'districtwise'),
        ('annex3_districtwise',),               # M159
        ('districtwise banking figure',),       # M163
        ('annex 3 p1_p2 districtwise',),
        ('annex3_p1_p2', 'banking'),            # M161
        ('annex 3_p1_p2 banking',),             # M164
    ])
    if not fp or fp.lower().endswith('.xls'):
        return {}

    wb = load_wb(fp)
    if not wb:
        return {}
    districts = {}

    # Page 1 — find data start by scanning for a row whose col-1 is a known district
    ws1 = find_sheet(wb, ["page 1", "page1"]) or wb[wb.sheetnames[0]]
    if ws1:
        rows = get_all_rows(ws1, min_row=1, max_col=20)
        data_start = None
        for i, row in enumerate(rows):
            d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if d:
                data_start = i
                break
        if data_start is not None:
            for row in rows[data_start:]:
                d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
                if not d:
                    continue
                rec = districts.setdefault(d, {})
                rec["nri_deposit"] = parse_number(row[2]) if len(row) > 2 else None
                rec["agri_adv_ac"] = parse_number(row[3]) if len(row) > 3 else None
                rec["agri_adv_amt"] = parse_number(row[4]) if len(row) > 4 else None
                rec["pct_agri_adv"] = parse_number(row[5]) if len(row) > 5 else None
                rec["msme_ac"] = parse_number(row[6]) if len(row) > 6 else None
                rec["msme_amt"] = parse_number(row[7]) if len(row) > 7 else None
                rec["other_ps_ac"] = parse_number(row[8]) if len(row) > 8 else None
                rec["other_ps_amt"] = parse_number(row[9]) if len(row) > 9 else None
                rec["total_priority_ac"] = parse_number(row[10]) if len(row) > 10 else None
                rec["total_priority_amt"] = parse_number(row[11]) if len(row) > 11 else None
                rec["total_advances"] = parse_number(row[12]) if len(row) > 12 else None
                rec["pct_psa"] = parse_number(row[13]) if len(row) > 13 else None

    # Page 2 — weaker sections
    ws2 = find_sheet(wb, ["page 2", "page2"])
    if ws2:
        rows = get_all_rows(ws2, min_row=1, max_col=20)
        data_start = None
        for i, row in enumerate(rows):
            d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if d:
                data_start = i
                break
        if data_start is not None:
            for row in rows[data_start:]:
                d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
                if not d:
                    continue
                rec = districts.setdefault(d, {})
                pairs = [
                    ("small_marginal_farmers_ac", "small_marginal_farmers_amt"),
                    ("sc_ac", "sc_amt"),
                    ("st_ac", "st_amt"),
                    ("dri_ac", "dri_amt"),
                    ("shg_ac", "shg_amt"),
                    ("women_ac", "women_amt"),
                    ("minority_ac", "minority_amt"),
                ]
                idx = 2
                for ac_key, amt_key in pairs:
                    if len(row) > idx:
                        rec[ac_key] = parse_number(row[idx])
                    if len(row) > idx + 1:
                        rec[amt_key] = parse_number(row[idx + 1])
                    idx += 2

    wb.close()
    # Strip Nones
    for d in districts:
        districts[d] = {k: v for k, v in districts[d].items() if v is not None}
    return districts


def extract_branch_old(annex_dir, meeting_num):
    """Annex23/24/27/31 Branch with 'District' sub-sheet.

    Header at row 3: No. | DISTRICT | Rural | Semi Urban | Urban | Total.
    """
    fp = _find_file_any(annex_dir, [
        ('annex23 branch',),
        ('annex24 branch',),
        ('annex 31 branch',),
        ('annex31 branch',),
        ('annex27 branch',),
        ('annex 27 branch',),
        ('annex 23 branch',),
        ('annex 24 branch',),
        ('annex28 branch',),                # M167
        ('annex 28 branch',),
        ('annex31-branches',),              # M159
        ('annex31 branch summary',),        # M161
    ])
    if not fp or fp.lower().endswith('.xls'):
        return {}
    wb = load_wb(fp)
    if not wb:
        return {}
    # Prefer "District" sheet, falling back to first sheet whose data has districts
    ws = None
    for sname in wb.sheetnames:
        if 'district' in sname.lower():
            ws = wb[sname]
            break
    if not ws:
        wb.close()
        return {}

    districts = {}
    rows = get_all_rows(ws, min_row=1, max_col=10)
    data_start = None
    for i, row in enumerate(rows):
        d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if d:
            data_start = i
            break
    if data_start is None:
        wb.close()
        return {}
    for row in rows[data_start:]:
        d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if not d:
            continue
        rec = {
            "branch_rural": parse_number(row[2]) if len(row) > 2 else None,
            "branch_semi_urban": parse_number(row[3]) if len(row) > 3 else None,
            "branch_urban": parse_number(row[4]) if len(row) > 4 else None,
            "total_branch": parse_number(row[5]) if len(row) > 5 else None,
        }
        rec = {k: v for k, v in rec.items() if v is not None}
        if rec:
            districts[d] = rec
    wb.close()
    return districts


def extract_atm_old(annex_dir, meeting_num):
    """Annex23/26/30 ATM Districtwise (or 'BWATm' sheet in older 165 layout)."""
    fp = _find_file_any(annex_dir, [
        ('atm districtwise',),
        ('districtwise atm',),
        ('atm district',),
        ('atm distwise',),                       # M159 'ATM DISTWISE'
        ('atm dist wise',),                      # M161
        ('annex 30 districtwise',),
        ('annex 30 district',),
        ('annex 30  district wise atm',),         # M164
        ('annex23 atm',),
        ('annex 23 atm',),
        ('annex22 atm',),
        ('annex 22 atm',),
        ('annex26', 'atm'),
        ('annex 26', 'atm'),
        ('annex27 bankstatement30atmdw',),       # M167
        ('annex30-atm distwise',),               # M159
        ('annex 30  district wise',),
    ])
    if not fp or fp.lower().endswith('.xls'):
        return {}
    wb = load_wb(fp)
    if not wb:
        return {}

    districts = {}
    # Two possible layouts:
    # (a) "atm"/"District" sheet with simple 6-col layout: No, District, Rural, SemiUrban, Urban, Total
    # (b) 165's 'BWATm' sheet with wide multi-area layout (Metro on/off/total, Urban,
    #     Semi Urban, Rural ...). We only try to pull a 'Total' for (b) where
    #     possible.
    ws = None
    for sname in wb.sheetnames:
        if 'district' in sname.lower() or sname.lower() in ('atm', 'bwatm'):
            ws = wb[sname]
            break
    if not ws:
        ws = wb[wb.sheetnames[0]]

    rows = get_all_rows(ws, min_row=1, max_col=16)
    data_start = None
    for i, row in enumerate(rows):
        d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
        if d:
            data_start = i
            break
    if data_start is None:
        wb.close()
        return {}

    # Layout (a): simple
    first_row = rows[data_start]
    cells = [safe_str(c) for c in first_row[:8]]
    # If exactly 4 numeric cells after col 1, treat as simple
    simple = True
    try:
        nums = [parse_number(c) for c in first_row[2:6]]
        if sum(1 for n in nums if n is not None) >= 3:
            simple = True
        else:
            simple = False
    except Exception:
        simple = True

    if simple:
        for row in rows[data_start:]:
            d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not d:
                continue
            rec = {
                "atm_rural": parse_number(row[2]) if len(row) > 2 else None,
                "atm_semi_urban": parse_number(row[3]) if len(row) > 3 else None,
                "atm_urban": parse_number(row[4]) if len(row) > 4 else None,
                "atm_total": parse_number(row[5]) if len(row) > 5 else None,
            }
            rec = {k: v for k, v in rec.items() if v is not None}
            if rec:
                districts[d] = rec

    wb.close()
    return districts


def extract_sacp_old(annex_dir, meeting_num):
    """Annex5 SACP Districtwise (multi-sheet)."""
    fp = _find_file_any(annex_dir, [
        ('sacp districtwise',),
        ('sacp district',),
        ('ps_sacp_districtwise',),
        ('districtwise sacp',),                  # M163
        ('annex5 bank_ps_sacp_districtwise',),    # M159, M167
        ('annex 5 bank_ps_sacp_districtwise',),
    ])
    if not fp or fp.lower().endswith('.xls'):
        return {}
    wb = load_wb(fp)
    if not wb:
        return {}

    districts = {}
    sheet_mapping = {
        "ACP": "sacp_total_ps",
        "CROP": "sacp_crop_loan",
        "TERM": "sacp_term_loan",
        "Agri_Infra": "sacp_agri_infra",
        "Agri. Infra": "sacp_agri_infra",
        "Total Agri": "sacp_total_agri",
        "Total MSME": "sacp_total_msme",
        "Edu_PS": "sacp_education",
        "Edu (PS)": "sacp_education",
        "Housing_PS": "sacp_housing",
        "Housing (PS)": "sacp_housing",
        "T Other PS": "sacp_other_ps",
    }

    for sheet_key, prefix in sheet_mapping.items():
        ws = find_sheet(wb, [sheet_key])
        if not ws:
            continue
        rows = get_all_rows(ws, min_row=1, max_col=12)
        data_start = None
        for i, row in enumerate(rows):
            d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if d:
                data_start = i
                break
        if data_start is None:
            continue
        for row in rows[data_start:]:
            d = normalize_district(safe_str(row[1]) if len(row) > 1 else "")
            if not d:
                continue
            rec = districts.setdefault(d, {})
            rec[f"{prefix}_target_ac"] = parse_number(row[2]) if len(row) > 2 else None
            rec[f"{prefix}_target_amt"] = parse_number(row[3]) if len(row) > 3 else None
            rec[f"{prefix}_disb_ac"] = parse_number(row[4]) if len(row) > 4 else None
            rec[f"{prefix}_disb_amt"] = parse_number(row[5]) if len(row) > 5 else None
            rec[f"{prefix}_pct_ach_ac"] = parse_number(row[6]) if len(row) > 6 else None
            rec[f"{prefix}_pct_ach_amt"] = parse_number(row[7]) if len(row) > 7 else None
            rec[f"{prefix}_os_ac"] = parse_number(row[8]) if len(row) > 8 else None
            rec[f"{prefix}_os_amt"] = parse_number(row[9]) if len(row) > 9 else None
    wb.close()
    for d in districts:
        districts[d] = {k: v for k, v in districts[d].items() if v is not None}
    return districts


# ----------------------------------------------------------------------
# Main per-meeting orchestrator
# ----------------------------------------------------------------------

def extract_meeting_new(meeting_num):
    """Extract a single 179-188 meeting by delegating to v1 extractors."""
    annex_dir = v1_find_annex_dir(meeting_num)
    if not annex_dir:
        return {}

    all_data = {}

    cd = v1_extract_cd_ratio(annex_dir, meeting_num)
    for d, data in cd.items():
        all_data.setdefault(d, {})["credit_deposit_ratio"] = data

    bp = v1_extract_banking_parameter(annex_dir, meeting_num)
    for d, data in bp.items():
        all_data.setdefault(d, {})["banking_parameter"] = data

    for tag, fn, cat in [
        ("PMJDY", v1_extract_pmjdy_dw, "pmjdy"),
        ("Aadhaar", v1_extract_aadhaar_casa, "aadhaar_authentication"),
        ("PMJJBY", v1_extract_pmjjby_pmsby, "social_security_schemes"),
        ("KCC", v1_extract_kcc, "kcc"),
        ("ATM", v1_extract_atm, "atm"),
        ("Branch", v1_extract_branch, "branch_network"),
        ("PMEGP", v1_extract_pmegp, "pmegp"),
        ("SACP", v1_extract_sacp, "sacp"),
        ("RSETI", v1_extract_rseti, "rseti"),
        ("SARFAESI", v1_extract_sarfaesi, "sarfaesi"),
        ("PMMY", v1_extract_pmmy_dw, "pmmy"),
    ]:
        try:
            data = fn(annex_dir, meeting_num)
        except Exception as e:
            print(f"  [WARN] M{meeting_num} {tag}: {e}")
            data = {}
        for d, dd in (data or {}).items():
            all_data.setdefault(d, {})[cat] = dd

    return all_data


def extract_meeting_old(meeting_num):
    """Extract a single 153-178 meeting from the data-tables/<n>/ tree."""
    annex_dir = find_old_annex_dir(meeting_num)
    if not annex_dir or not os.path.isdir(annex_dir):
        return {}

    all_data = {}

    cd = extract_cd_ratio_old(annex_dir, meeting_num)
    for d, data in cd.items():
        all_data.setdefault(d, {})["credit_deposit_ratio"] = data

    bp = extract_banking_parameter_old(annex_dir, meeting_num)
    for d, data in bp.items():
        all_data.setdefault(d, {})["banking_parameter"] = data

    br = extract_branch_old(annex_dir, meeting_num)
    for d, data in br.items():
        # If CD ratio already provided total_branch, banking_network row stays informative
        all_data.setdefault(d, {})["branch_network"] = data

    atm = extract_atm_old(annex_dir, meeting_num)
    for d, data in atm.items():
        all_data.setdefault(d, {})["atm"] = data

    sacp = extract_sacp_old(annex_dir, meeting_num)
    for d, data in sacp.items():
        all_data.setdefault(d, {})["sacp"] = data

    return all_data


def main():
    os.chdir(BASE)

    targets = sorted(ALL_MEETING_QUARTERS.keys())
    if len(sys.argv) > 1:
        if sys.argv[1] != '--all':
            targets = [int(a) for a in sys.argv[1:] if a.lstrip('-').isdigit()]

    print(f"Target meetings: {targets[0]}..{targets[-1]} ({len(targets)} total)")

    all_quarters = []
    for n in targets:
        qinfo = ALL_MEETING_QUARTERS[n]
        if n in V1_MEETING_QUARTERS:
            data = extract_meeting_new(n)
        else:
            data = extract_meeting_old(n)

        cats = set()
        for dist, categories in data.items():
            cats.update(categories.keys())
        print(f"  M{n} ({qinfo['period']:>17}): {len(data)} districts, "
              f"{len(cats)} categories: {sorted(cats)}")

        if data:
            all_quarters.append({
                "quarter_key": qinfo["quarter_key"],
                "period": qinfo["period"],
                "as_on_date": qinfo["as_on_date"],
                "fy": qinfo["fy"],
                "data": data,
            })

    all_quarters.sort(key=lambda q: q["quarter_key"])

    complete = build_complete_json(all_quarters)
    timeseries = build_timeseries_json(all_quarters)

    out_complete = os.path.join(BASE, "gujarat_complete.json")
    out_timeseries = os.path.join(BASE, "gujarat_fi_timeseries.json")
    with open(out_complete, "w") as f:
        json.dump(complete, f, indent=2, ensure_ascii=False)
    with open(out_timeseries, "w") as f:
        json.dump(timeseries, f, indent=2, ensure_ascii=False)

    print(f"\nWrote {out_complete} and {out_timeseries}")
    print(f"  Quarters: {len(all_quarters)}")
    cats_all = set()
    for q in all_quarters:
        for dist, categories in q["data"].items():
            cats_all.update(categories.keys())
    print(f"  Categories: {sorted(cats_all)}")


if __name__ == "__main__":
    main()
