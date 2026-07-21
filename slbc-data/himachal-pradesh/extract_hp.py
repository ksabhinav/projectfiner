#!/usr/bin/env python3
"""
Himachal Pradesh SLBC CQR (Comprehensive Quarterly Return) extractor.

Source: https://slbchp.com/Slbcmeetingspage/agenda-minutes.aspx
        -> 179th SLBC Agenda - Quarter Dec 25 ZIP (4 XLSX annexure bundles).

The XLSX bundle ships 74 annexures; ~30 are *district-wise* (12 rows = the
12 HP districts: Bilaspur, Chamba, Hamirpur, Kangra, Kinnaur, Kullu,
Lahul Spiti, Mandi, Shimla, Sirmaur, Solan, Una).

Units: "(Rs. In lacs)" per the table headers — NO ×100 Crore→Lakh
conversion needed (FINER's canonical unit is already Rs Lakhs).

Outputs (under public/slbc-data/himachal-pradesh/):
  - himachal-pradesh_fi_timeseries.json   (periods → districts)
  - himachal-pradesh_complete.json
  - himachal-pradesh_fi_slim.json
  - himachal-pradesh_fi_timeseries.csv

Quarter scope: Dec 2025 (the only quarter HP publishes bundled XLSX
annexures for). Older quarters are PDF-only; a future pdfplumber
pipeline can backfill, mirroring slbc-data/uttarakhand/.
"""

import argparse
import csv
import glob
import json
import os
import re
import sys
import warnings

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(os.path.dirname(HERE))
RAW_DIR = os.path.join(HERE, "raw", "agenda179")
PUBLIC_OUT = os.path.join(PROJECT, "public", "slbc-data", "himachal-pradesh")
LOCAL_OUT = HERE

CANONICAL_DISTRICTS = [
    "Bilaspur", "Chamba", "Hamirpur", "Kangra", "Kinnaur", "Kullu",
    "Lahul and Spiti",  # matches districts.name in finer.db (lgd 21)
    "Mandi", "Shimla", "Sirmaur", "Solan", "Una",
]
DISTRICT_LOOKUP = {d.upper(): d for d in CANONICAL_DISTRICTS}
# Common variants observed in PNB/SLBC HP source text.
DISTRICT_ALIASES = {
    "LAHUL & SPITI": "Lahul and Spiti",
    "LAHUL AND SPITI": "Lahul and Spiti",
    "LAHUL-SPITI": "Lahul and Spiti",
    "L&S": "Lahul and Spiti",
    "L & S": "Lahul and Spiti",
    "LAHAUL SPITI": "Lahul and Spiti",
    "LAHAUL & SPITI": "Lahul and Spiti",
    "LAHAUL AND SPITI": "Lahul and Spiti",
    "LAHAUL-SPITI": "Lahul and Spiti",
    "SIRMOUR": "Sirmaur",
    "SHIMLA URBAN": "Shimla",
    "TOTAL": None,
    "GRAND TOTAL": None,
}

# Set per-run from --quarter (default = the original Dec 2025 bundle).
QUARTER_LABEL = "December 2025"   # what FINER stores
QUARTER_CODE = "2025-12"
SOURCE_FILE_TAG = "himachal-pradesh"

MONTH_NAMES = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May",
               6: "June", 7: "July", 8: "August", 9: "September",
               10: "October", 11: "November", 12: "December"}
MONTH_NUM = {v: k for k, v in MONTH_NAMES.items()}

# Per-quarter column overrides.
#
# The CQR annexures are not column-stable across quarters. Mar 2026's ANNEX-52
# inserted a "Target" No./Amt pair ahead of every "Disbursements" block, so the
# Dec 2025 indices land on the (all-zero) target columns and shift every
# disbursement field one block left — which silently reported total NRLM
# disbursements as the women-SHG figure (a 60-760x jump the validator caught).
# Mapping verified by value continuity against Dec 2025: (5, 6.86)->(5, 6.97),
# (2, 12)->(2, 12), 2->2, and the NULM tail 6/0.97/6/6/6.56/26 -> 8/1.96/8/
# 7/6.68/27.
#
# Field NAMES are kept as-is for series continuity even though Dec 2025's
# "nrlm_women_shg"/"nrlm_individual" keys actually hold NULM-SEP(I)/(G) blocks
# — renaming them is a separate correction, not this quarter's job.
ANNEXURE_OVERRIDES = {
    "2026-03": {
        "ANNEX-52": {
            "category": "nrlm",
            "data_start_row": 6,
            "cols": {
                5: "nrlm_shg_disbursement_no",
                6: "nrlm_shg_disbursement_amt",
                9: "nrlm_women_shg_disbursement_no",
                10: "nrlm_women_shg_disbursement_amt",
                13: "nrlm_individual_disbursement_no",
                14: "nrlm_individual_disbursement_amt",
                15: "nrlm_beneficiaries_no",
                16: "nulm_shg_disbursement_no",
                17: "nulm_shg_disbursement_amt",
                18: "nulm_beneficiaries_no",
                19: "nulm_women_shg_disbursement_no",
                20: "nulm_women_shg_disbursement_amt",
                21: "nulm_women_beneficiaries_no",
            },
        },
    },
}

F_TS = "himachal-pradesh_fi_timeseries.json"
F_COMPLETE = "himachal-pradesh_complete.json"
F_SLIM = "himachal-pradesh_fi_slim.json"
F_CSV = "himachal-pradesh_fi_timeseries.csv"


def load_json(path, default):
    """Existing output, or `default` on a first run."""
    if not os.path.exists(path):
        return json.loads(json.dumps(default))
    with open(path) as fh:
        return json.load(fh)


def period_sort_key(p):
    """'March 2026' -> ('2026', 3). Sorting the label lexically would put
    September ahead of December."""
    m = re.match(r"([A-Za-z]+)\s+(\d{4})", p.get("period", ""))
    if not m:
        return ("0000", 0)
    return (m.group(2), MONTH_NUM.get(m.group(1), 0))


def quarter_sort_key(item):
    """complete.json keys are YYYY-MM, except one legacy 'dec_2025' written by
    an earlier run of this script. Sort that alongside its YYYY-MM peers."""
    key = item[0]
    if re.fullmatch(r"\d{4}-\d{2}", key):
        return key
    m = re.match(r"([a-z]+)_(\d{4})", key)
    if m:
        for num, name in MONTH_NAMES.items():
            if name.lower().startswith(m.group(1)):
                return f"{m.group(2)}-{num:02d}"
    return key


# ----------------------------------------------------------------------------
# Annexure → (category, field-header) mapping
# Each entry: list of (column_index_1based, field_key)
# field_key is "<category>__<field_name>" (matches FINER's "__" convention)
# Column indices are 1-based to match openpyxl ws.cell().
#
# Categories used (aligned with FINER canonical category prefixes):
#   branch_network, credit_deposit_ratio, priority_sector,
#   non_priority_sector, annual_credit_plan, digital_transactions,
#   kcc, education_loan, shg, pmegp, nrlm, minority_loans, sc_st,
#   women_finance, pmjdy, mudra, standup_india, aadhaar_authentication,
#   investment_credit, npa
# ----------------------------------------------------------------------------

ANNEXURES = {
    # (filename_glob, sheet_name): {data_start_row, columns}
    "ANNEX-02": {  # Branch network by district
        "category": "branch_network",
        "data_start_row": 6,
        "cols": {
            3: "branch_rural",
            4: "branch_semi_urban",
            5: "branch_urban",
            6: "total_branch",
            7: "atm_rural",
            8: "atm_semi_urban",
            9: "atm_urban",
            10: "total_atm",
        },
    },
    "ANNEX-04": {  # Banking outlets / BC outlets
        "category": "branch_network",
        "data_start_row": 5,
        "cols": {
            3: "banking_outlets_no",
            4: "fixed_bc_outlets_no",
            5: "other_bc_outlets_no",
        },
    },
    "ANNEX-06": {  # Deposits, advances, CD ratio
        # NOTE: Header column "Amt" is the canonical "total_deposit" /
        # "total_advance" (₹ Lakhs); column "No" is account count.
        "category": "credit_deposit_ratio",
        "data_start_row": 6,
        "cols": {
            3: "previous_qtr_deposit_no",
            4: "previous_qtr_deposit_amt",
            5: "total_deposit_no",
            6: "total_deposit",          # canonical (₹ Lakhs)
            7: "previous_qtr_advance_no",
            8: "previous_qtr_advance_amt",
            9: "total_advance_no",
            10: "total_advance",         # canonical (₹ Lakhs)
            11: "previous_qtr_cd_ratio_no",
            12: "previous_qtr_cd_ratio",
            13: "current_cd_ratio_no",
            14: "overall_cd_ratio",      # canonical % (matches other states)
        },
    },
    "ANNEX-12": {  # Agriculture loans
        "category": "priority_sector",
        "data_start_row": 6,
        "cols": {
            3: "farm_credit_total_no",
            4: "farm_credit_total_amt",
            5: "crop_loan_no",
            6: "crop_loan_amt",
            7: "investment_credit_no",
            8: "investment_credit_amt",
            12: "agri_infra_no",
            13: "agri_infra_amt",
            14: "ancillary_no",
            15: "ancillary_amt",
            16: "total_agri_no",
            17: "total_agri_amt",
        },
    },
    "ANNEX-14": {  # MSME loans
        "category": "priority_sector",
        "data_start_row": 6,
        "cols": {
            3: "msme_micro_no",
            4: "msme_micro_amt",
            5: "msme_small_no",
            6: "msme_small_amt",
            7: "msme_medium_no",
            8: "msme_medium_amt",
            9: "other_msme_no",
            10: "other_msme_amt",
            11: "msme_total_no",
            12: "msme_total_amt",
        },
    },
    "ANNEX-16": {  # Other priority sector
        "category": "priority_sector",
        "data_start_row": 6,
        "cols": {
            3: "export_no",
            4: "export_amt",
            5: "education_no",
            6: "education_amt",
            7: "housing_no",
            8: "housing_amt",
            9: "social_infra_no",
            10: "social_infra_amt",
            11: "renewable_no",
            12: "renewable_amt",
            13: "others_no",
            14: "others_amt",
            15: "total_priority_no",
            16: "total_priority_amt",
        },
    },
    "ANNEX-18": {  # Weaker sections
        "category": "priority_sector",
        "data_start_row": 6,
        "cols": {
            3: "weaker_small_marginal_no",
            4: "weaker_small_marginal_amt",
            5: "weaker_sc_st_no",
            6: "weaker_sc_st_amt",
            7: "weaker_shg_no",
            8: "weaker_shg_amt",
            9: "weaker_minority_no",
            10: "weaker_minority_amt",
            11: "od_pmjdy_no",
            12: "od_pmjdy_amt",
            19: "total_weaker_no",
            20: "total_weaker_amt",
        },
    },
    "ANNEX-20": {  # Non-priority sector
        "category": "non_priority_sector",
        "data_start_row": 7,
        "cols": {
            3: "nps_agri_no",
            4: "nps_agri_amt",
            5: "nps_education_no",
            6: "nps_education_amt",
            7: "nps_housing_no",
            8: "nps_housing_amt",
            9: "nps_personal_no",
            10: "nps_personal_amt",
            11: "nps_others_no",
            12: "nps_others_amt",
            13: "nps_total_no",
            14: "nps_total_amt",
        },
    },
    "ANNEX-22": {  # ACP agri
        "category": "annual_credit_plan",
        "data_start_row": 6,
        "cols": {
            3: "acp_farm_credit_target_no",
            4: "acp_farm_credit_target_amt",
            5: "acp_farm_credit_achievement_no",
            6: "acp_farm_credit_achievement_amt",
            7: "acp_farm_credit_achievement_pct",
            8: "acp_crop_loan_target_no",
            9: "acp_crop_loan_target_amt",
            10: "acp_crop_loan_achievement_no",
            11: "acp_crop_loan_achievement_amt",
            12: "acp_crop_loan_achievement_pct",
        },
    },
    "ANNEX-26": {  # ACP MSME
        "category": "annual_credit_plan",
        "data_start_row": 6,
        "cols": {
            3: "acp_msme_target_no",
            4: "acp_msme_target_amt",
            5: "acp_msme_achievement_no",
            6: "acp_msme_achievement_amt",
            7: "acp_msme_achievement_pct",
        },
    },
    "ANNEX-30": {  # ACP other priority + total priority
        "category": "annual_credit_plan",
        "data_start_row": 6,
        "cols": {
            18: "acp_total_priority_target_no",
            19: "acp_total_priority_target_amt",
            20: "acp_total_priority_achievement_no",
            21: "acp_total_priority_achievement_amt",
            22: "acp_total_priority_achievement_pct",
        },
    },
    "ANNEX-32A": {  # ACP non-priority total
        "category": "annual_credit_plan",
        "data_start_row": 6,
        "cols": {
            13: "acp_nps_total_target_no",
            14: "acp_nps_total_target_amt",
            15: "acp_nps_total_achievement_no",
            16: "acp_nps_total_achievement_amt",
            17: "acp_nps_total_achievement_pct",
        },
    },
    "ANNEX-44": {  # KCC crop loans
        # Canonical KCC count = total_no_of_kcc (matches Meghalaya pattern)
        "category": "kcc",
        "data_start_row": 5,
        "cols": {
            3: "kcc_issued_during_quarter_no",
            4: "kcc_issued_during_quarter_amt",
            5: "total_no_of_kcc",
            6: "kcc_outstanding_amt",
            7: "kcc_active_no",
            8: "kcc_card_activated_no",
        },
    },
    "ANNEX-45": {  # KCC animal husbandry
        "category": "kcc",
        "data_start_row": 4,
        "cols": {
            3: "kcc_ah_issued_during_quarter_no",
            4: "kcc_ah_issued_during_quarter_amt",
            5: "kcc_ah_total_no",
            6: "kcc_ah_outstanding_amt",
            7: "kcc_ah_active_no",
            8: "kcc_ah_card_activated_no",
        },
    },
    "ANNEX-46": {  # KCC fisheries
        "category": "kcc",
        "data_start_row": 4,
        "cols": {
            3: "kcc_fish_issued_during_quarter_no",
            4: "kcc_fish_issued_during_quarter_amt",
            5: "kcc_fish_total_no",
            6: "kcc_fish_outstanding_amt",
            7: "kcc_fish_active_no",
            8: "kcc_fish_card_activated_no",
        },
    },
    "ANNEX-50": {  # SHG linkage & JLGs
        "category": "shg",
        "data_start_row": 6,
        "cols": {
            3: "savings_linked_no",
            4: "savings_linked_amt",
            5: "credit_linked_no",
            6: "credit_linked_amt",
            7: "current_fy_savings_linked_no",
            8: "current_fy_savings_linked_amt",
            9: "current_fy_credit_linked_no",
            10: "current_fy_credit_linked_amt",
            11: "shg_loan_outstanding_no",
            12: "shg_loan_outstanding_amt",
            13: "shg_npa_no",
            14: "shg_npa_amt",
            15: "jlg_disbursement_no",
            16: "jlg_disbursement_amt",
            17: "jlg_outstanding_no",
            18: "jlg_outstanding_amt",
        },
    },
    "ANNEX-50A": {  # PMEGP
        "category": "pmegp",
        "data_start_row": 5,
        "cols": {
            3: "pmegp_sanction_no",
            4: "pmegp_sanction_amt",
            5: "pmegp_disbursed_no",
            6: "pmegp_disbursed_amt",
            7: "pmegp_outstanding_no",
            8: "pmegp_outstanding_amt",
        },
    },
    "ANNEX-52": {  # NRLM & NULM disbursements
        "category": "nrlm",
        "data_start_row": 6,
        "cols": {
            3: "nrlm_shg_disbursement_no",
            4: "nrlm_shg_disbursement_amt",
            5: "nrlm_women_shg_disbursement_no",
            6: "nrlm_women_shg_disbursement_amt",
            7: "nrlm_individual_disbursement_no",
            8: "nrlm_individual_disbursement_amt",
            9: "nrlm_beneficiaries_no",
            10: "nulm_shg_disbursement_no",
            11: "nulm_shg_disbursement_amt",
            12: "nulm_beneficiaries_no",
            13: "nulm_women_shg_disbursement_no",
            14: "nulm_women_shg_disbursement_amt",
            15: "nulm_women_beneficiaries_no",
        },
    },
    "ANNEX-58": {  # SC/ST loans
        "category": "sc_st",
        "data_start_row": 6,
        "cols": {
            3: "sc_outstanding_no",
            4: "sc_outstanding_amt",
            5: "st_outstanding_no",
            6: "st_outstanding_amt",
            7: "sc_disbursement_no",
            8: "sc_disbursement_amt",
            9: "st_disbursement_no",
            10: "st_disbursement_amt",
        },
    },
    "ANNEX-60": {  # Women loans
        "category": "women_finance",
        "data_start_row": 5,
        "cols": {
            3: "women_outstanding_no",
            4: "women_outstanding_amt",
            5: "individual_women_beneficiaries_no",
            6: "individual_women_beneficiaries_amt",
            7: "women_disbursed_no",
            8: "women_disbursed_amt",
        },
    },
    "ANNEX-62": {  # PMJDY
        # Canonical field names (matches Meghalaya/Tripura/etc. pattern):
        # total_pmjdy_no, male_no, female_no, no_of_aadhaar_seeded
        "category": "pmjdy",
        "data_start_row": 6,
        "cols": {
            3: "rural_no",
            4: "urban_no",
            5: "total_pmjdy_no",
            6: "male_no",
            7: "female_no",
            8: "total_no",
            9: "no_of_zero_balance_a_c",
            10: "zero_balance_pct",
            11: "deposits_held_in_a_c_amt",
            12: "overdraft_no",
            13: "overdraft_amt",
            14: "no_of_rupay_card_issued",
            15: "rupay_card_activated",
            16: "no_of_aadhaar_seeded",
            17: "aadhaar_seeding_pct",
        },
    },
    "ANNEX-64": {  # MUDRA
        "category": "mudra",
        "data_start_row": 6,
        "cols": {
            3: "shishu_outstanding_no",
            4: "shishu_outstanding_amt",
            5: "shishu_disbursement_no",
            6: "shishu_disbursement_amt",
            7: "shishu_npa_no",
            8: "shishu_npa_amt",
            9: "kishore_outstanding_no",
            10: "kishore_outstanding_amt",
            11: "kishore_disbursement_no",
            12: "kishore_disbursement_amt",
            13: "kishore_npa_no",
            14: "kishore_npa_amt",
            15: "tarun_outstanding_no",
            16: "tarun_outstanding_amt",
            17: "tarun_disbursement_no",
            18: "tarun_disbursement_amt",
            19: "tarun_npa_no",
            20: "tarun_npa_amt",
            21: "mudra_total_outstanding_no",
            22: "mudra_total_outstanding_amt",
            23: "mudra_total_disbursement_no",
            24: "mudra_total_disbursement_amt",
            25: "mudra_total_npa_no",
            26: "mudra_total_npa_amt",
        },
    },
    "ANNEX-66": {  # StandUp India
        "category": "standup_india",
        "data_start_row": 6,
        "cols": {
            3: "target_no",
            4: "female_sanctioned_no",
            5: "female_sanctioned_amt",
            6: "sc_sanctioned_no",
            7: "sc_sanctioned_amt",
            8: "st_sanctioned_no",
            9: "st_sanctioned_amt",
            10: "total_sanctioned_no",
            11: "total_sanctioned_amt",
        },
    },
    "ANNEX-37": {  # Priority sector NPA - agri
        "category": "npa",
        "data_start_row": 7,
        "cols": {
            3: "npa_ps_farm_credit_no",
            4: "npa_ps_farm_credit_amt",
            5: "npa_ps_crop_loan_no",
            6: "npa_ps_crop_loan_amt",
            7: "npa_ps_investment_credit_no",
            8: "npa_ps_investment_credit_amt",
            15: "npa_ps_total_agri_no",
            16: "npa_ps_total_agri_amt",
        },
    },
    "ANNEX-37A": {  # NPA - MSME / education
        "category": "npa",
        "data_start_row": 7,
        "cols": {
            3: "npa_msme_micro_no",
            4: "npa_msme_micro_amt",
            11: "npa_msme_total_no",
            12: "npa_msme_total_amt",
            15: "npa_education_no",
            16: "npa_education_amt",
        },
    },
    "ANNEX-37B": {  # NPA - PS totals & %
        "category": "npa",
        "data_start_row": 7,
        "cols": {
            11: "npa_ps_total_no",
            12: "npa_ps_total_amt",
            13: "npa_ps_pct",
        },
    },
    "ANNEX-38": {  # NPA - NPS, total NPA, total advances
        "category": "npa",
        "data_start_row": 6,
        "cols": {
            13: "npa_nps_total_no",
            14: "npa_nps_total_amt",
            15: "npa_total_no",
            16: "npa_total_amt",
            17: "total_advances_no",
            18: "total_advances_amt",
            19: "npa_pct",
        },
    },
    "ANNEX-56": {  # Minority community loan disbursements
        "category": "minority_loans",
        "data_start_row": 5,
        "cols": {
            3: "christians_no",
            4: "christians_amt",
            5: "muslims_no",
            6: "muslims_amt",
            7: "buddhists_no",
            8: "buddhists_amt",
            9: "sikhs_no",
            10: "sikhs_amt",
            11: "zoroastrians_no",
            12: "zoroastrians_amt",
            13: "jains_no",
            14: "jains_amt",
            15: "minority_total_no",
            16: "minority_total_amt",
        },
    },
}

# Map each ANNEX key to (file, sheet_name_in_workbook).
# In the Dec 2025 agenda bundle the sheets are reshuffled relative to their
# numeric order (e.g. ANNEX-58, ANNEX-60 ship inside the 1-20 file).
ANNEX_SHEETS = {
    "ANNEX-02"  : "ANNEX-02",
    "ANNEX-04"  : "ANNEX-04",
    "ANNEX-06"  : "ANNEX-06",
    "ANNEX-12"  : "ANNEX-12",
    "ANNEX-14"  : "ANNEX-14",
    "ANNEX-16"  : "ANNEX-16",
    "ANNEX-18"  : "ANNEX-18",
    "ANNEX-20"  : "ANNEX-20",
    "ANNEX-58"  : "ANNEX-58",
    "ANNEX-60"  : "ANNEX-60",
    "ANNEX-22"  : "ANNEX-22",
    "ANNEX-26"  : "ANNEX-26",
    "ANNEX-30"  : "ANNEX-30",
    "ANNEX-32A" : "ANNEX-32A",
    "ANNEX-37"  : "ANNEX-37",
    "ANNEX-37A" : "ANNEX-37A",
    "ANNEX-37B" : "ANNEX-37B",
    "ANNEX-38"  : "ANNEX-38",
    "ANNEX-50"  : "ANNEX-50",
    "ANNEX-44"  : "ANNEX-44",
    "ANNEX-45"  : "ANNEX-45",
    "ANNEX-46"  : "ANNEX-46",
    "ANNEX-50A" : "ANNEX-49A",
    "ANNEX-52"  : "ANNEX-52",
    "ANNEX-56"  : "ANNEX-56",
    "ANNEX-62"  : "ANNEX-62",
    "ANNEX-64"  : "ANNEX-64",
    "ANNEX-66"  : "ANNEX-66",
}


def canonical_district(raw_name):
    if raw_name is None:
        return None
    s = str(raw_name).strip().upper()
    s = re.sub(r"\s+", " ", s)
    if s in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[s]
    if s in DISTRICT_LOOKUP:
        return DISTRICT_LOOKUP[s]
    # Fuzzy strip of trailing "TOTAL"
    if "TOTAL" in s:
        return None
    return None


def to_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if v != v:  # NaN
            return None
        return float(v)
    s = str(v).strip()
    if not s or s in {"-", "NA", "N/A", "Nil", "nil", "NIL", "#DIV/0!", "None"}:
        return None
    try:
        return float(s.replace(",", "").replace("%", ""))
    except ValueError:
        return None


def extract_annex(wb, sheet_name, spec):
    """Returns {district: {field_key: value}} for one annexure."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    start = spec["data_start_row"]
    out = {}
    for r in range(start, ws.max_row + 1):
        raw = ws.cell(r, 2).value
        dist = canonical_district(raw)
        if not dist:
            continue
        rec = out.setdefault(dist, {})
        for col, fname in spec["cols"].items():
            if col > ws.max_column:
                continue
            val = to_number(ws.cell(r, col).value)
            if val is None:
                continue
            # Build "<category>__<field>" key
            key = f"{spec['category']}__{fname}"
            rec[key] = val
    return out


def build_sheet_index(raw_dir):
    """sheet name -> file path, scanning every XLSX in the bundle.

    The CQR bundles do NOT keep a stable annexure->file assignment between
    quarters: ANNEX-58/60 shipped in the "1To20" workbook for Dec 2025 but in
    "51To74" for Mar 2026, and ANNEX-49A moved from "51To74" to "33TO50".
    Resolving by sheet name instead of filename survives that drift (and the
    per-quarter " Dec 25" / " (2)" filename suffixes).
    """
    index = {}
    for path in sorted(glob.glob(os.path.join(raw_dir, "*.xlsx"))):
        if os.path.basename(path).startswith("~$"):
            continue          # Excel lock file
        wb = openpyxl.load_workbook(path, read_only=True)
        for sheet in wb.sheetnames:
            # First file wins; only ANNEX-67/68 are duplicated and neither is
            # mapped in ANNEXURES, so this is unambiguous for what we read.
            index.setdefault(sheet, path)
        wb.close()
    return index


def main():
    global QUARTER_LABEL, QUARTER_CODE

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--raw-dir", default=RAW_DIR,
                    help="directory holding the quarter's 4 CQR XLSX files")
    ap.add_argument("--quarter", default=QUARTER_CODE,
                    help="quarter key, YYYY-MM (e.g. 2026-03)")
    args = ap.parse_args()

    raw_dir = args.raw_dir
    QUARTER_CODE = args.quarter
    if not re.fullmatch(r"\d{4}-\d{2}", QUARTER_CODE):
        print(f"ERROR: --quarter must be YYYY-MM, got {QUARTER_CODE!r}",
              file=sys.stderr)
        sys.exit(1)
    QUARTER_LABEL = f"{MONTH_NAMES[int(QUARTER_CODE[5:])]} {QUARTER_CODE[:4]}"

    if not os.path.exists(raw_dir):
        print(f"ERROR: raw dir not found: {raw_dir}", file=sys.stderr)
        sys.exit(1)

    sheet_index = build_sheet_index(raw_dir)
    print(f"{raw_dir}\n  -> {QUARTER_LABEL} ({QUARTER_CODE}), "
          f"{len(sheet_index)} sheets across "
          f"{len(set(sheet_index.values()))} workbooks")

    # Open each XLSX once
    wb_cache = {}
    def get_wb(path):
        if path not in wb_cache:
            wb_cache[path] = openpyxl.load_workbook(path, data_only=True)
        return wb_cache[path]

    # district -> {field_key: value}
    district_recs = {d: {"district": d, "period": QUARTER_LABEL}
                     for d in CANONICAL_DISTRICTS}

    # complete.json style: quarters -> tables -> category -> {fields, districts}
    tables_by_cat = {}

    overrides = ANNEXURE_OVERRIDES.get(QUARTER_CODE, {})
    if overrides:
        print(f"  applying {len(overrides)} column override(s) for "
              f"{QUARTER_CODE}: {', '.join(sorted(overrides))}")

    annex_summary = []
    for annex_key, spec in ANNEXURES.items():
        spec = overrides.get(annex_key, spec)
        sheet_name = ANNEX_SHEETS.get(annex_key)
        if not sheet_name:
            print(f"  skip {annex_key}: no sheet mapping")
            continue
        path = sheet_index.get(sheet_name)
        if not path:
            print(f"  skip {annex_key}: sheet {sheet_name} not in bundle")
            continue
        wb = get_wb(path)
        data = extract_annex(wb, sheet_name, spec)
        if not data:
            print(f"  WARN {annex_key}: no rows extracted (sheet missing?)")
            continue
        # Merge into district records
        cat = spec["category"]
        cat_table = tables_by_cat.setdefault(
            cat, {"fields": [], "districts": {}}
        )
        for dist, kv in data.items():
            district_recs[dist].update(kv)
            ctab_d = cat_table["districts"].setdefault(dist, {})
            for fk, val in kv.items():
                _, fname_no_cat = fk.split("__", 1)
                ctab_d[fname_no_cat] = val
                if fname_no_cat not in cat_table["fields"]:
                    cat_table["fields"].append(fname_no_cat)
        annex_summary.append((annex_key, cat, len(data),
                              len(spec["cols"])))

    # ---- derive CD ratio when the source leaves it blank -------------------
    # HP's ANNEX-06 prints deposits, advances AND the C.D RATIO columns, but
    # the Mar 2026 bundle ships those four columns empty. Deposits/advances
    # are present and continuous, so compute the ratio rather than shipping a
    # quarter that is invisible on the CD-ratio choropleth. Same convention as
    # db/normalize_wayback_madhya_pradesh.py's map_cd, incl. the sanity bound
    # (see CLAUDE.md gotcha #51 — outside 5-400% it's a unit mismatch, not a
    # real ratio, and no data beats nonsense).
    CD = "credit_deposit_ratio__"
    derived = []
    for d, rec in district_recs.items():
        if rec.get(CD + "overall_cd_ratio") is not None:
            continue
        try:
            dep = float(rec[CD + "total_deposit"])
            adv = float(rec[CD + "total_advance"])
        except (KeyError, TypeError, ValueError):
            continue
        if dep <= 0:
            continue
        val = adv / dep * 100.0
        if not (5 <= val <= 400):
            print(f"  WARN {d}: derived CD {val:.2f}% out of bounds, skipped")
            continue
        rec[CD + "overall_cd_ratio"] = round(val, 2)
        tbl = tables_by_cat.get("credit_deposit_ratio")
        if tbl is not None:
            tbl["districts"].setdefault(d, {})["overall_cd_ratio"] = round(val, 2)
            if "overall_cd_ratio" not in tbl["fields"]:
                tbl["fields"].append("overall_cd_ratio")
        derived.append(d)
    if derived:
        print(f"  derived overall_cd_ratio for {len(derived)}/"
              f"{len(CANONICAL_DISTRICTS)} districts (source columns blank)")

    # ---- merge this quarter into the existing outputs ---------------------
    # HP's timeseries/complete carry 11 further quarters recovered by
    # db/normalize_wayback_hp.py. This script used to write single-quarter
    # files, so re-running it wiped that history — always upsert, never
    # replace wholesale.
    timeseries = load_json(os.path.join(PUBLIC_OUT, F_TS), {"periods": []})
    periods = [p for p in timeseries.get("periods", [])
               if p.get("period") != QUARTER_LABEL]
    periods.append({
        "period": QUARTER_LABEL,
        "districts": [district_recs[d] for d in CANONICAL_DISTRICTS],
    })
    periods.sort(key=period_sort_key)
    timeseries["periods"] = periods

    complete = load_json(os.path.join(PUBLIC_OUT, F_COMPLETE), {"quarters": {}})
    quarters = complete.setdefault("quarters", {})
    # Drop any legacy alias of this quarter first — an earlier version of this
    # script keyed Dec 2025 as "dec_2025", so a plain assignment to "2025-12"
    # would leave the same quarter in the file twice.
    for stale in [k for k in quarters
                  if k != QUARTER_CODE
                  and quarter_sort_key((k, None)) == QUARTER_CODE]:
        print(f"  dropping legacy duplicate quarter key {stale!r}")
        del quarters[stale]
    quarters[QUARTER_CODE] = {"period": QUARTER_LABEL, "tables": tables_by_cat}
    complete["quarters"] = dict(sorted(quarters.items(), key=quarter_sort_key))

    # Slim = the merged timeseries restricted to the 7 homepage categories.
    # Rebuilt from the full series (not just this quarter) because the
    # committed slim had drifted to a single period while the timeseries held
    # twelve.
    SLIM_PREFIXES = (
        "credit_deposit_ratio__", "pmjdy__", "branch_network__",
        "kcc__", "shg__", "digital_transactions__",
        "aadhaar_authentication__",
    )
    slim_periods = []
    for p in timeseries["periods"]:
        recs = []
        for d in p["districts"]:
            rec = {"district": d["district"], "period": d["period"]}
            rec.update({k: v for k, v in d.items()
                        if k not in ("district", "period")
                        and k.startswith(SLIM_PREFIXES)})
            recs.append(rec)
        slim_periods.append({"period": p["period"], "districts": recs})
    slim = {"periods": slim_periods}

    # Wide CSV, rebuilt across every period in the merged timeseries (it is
    # the Downloads-page artefact, so it must not shrink to one quarter).
    seen = set()
    for p in timeseries["periods"]:
        for d in p["districts"]:
            seen.update(k for k in d if k not in ("district", "period"))
    all_field_keys = sorted(seen)
    csv_headers = ["district", "period"] + all_field_keys

    # Write outputs
    os.makedirs(LOCAL_OUT, exist_ok=True)
    os.makedirs(PUBLIC_OUT, exist_ok=True)

    for outdir in (LOCAL_OUT, PUBLIC_OUT):
        with open(os.path.join(outdir, F_TS), "w") as f:
            json.dump(timeseries, f, indent=2, ensure_ascii=False)
        with open(os.path.join(outdir, F_COMPLETE), "w") as f:
            json.dump(complete, f, indent=2, ensure_ascii=False)
        with open(os.path.join(outdir, F_SLIM), "w") as f:
            json.dump(slim, f, indent=2, ensure_ascii=False)
        with open(os.path.join(outdir, F_CSV), "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(csv_headers)
            for p in timeseries["periods"]:
                for d in p["districts"]:
                    w.writerow([d["district"], d["period"]]
                               + [d.get(k, "") for k in all_field_keys])

    # Summary
    total_cells = sum(
        len(v) - 2  # exclude district, period
        for v in district_recs.values()
    )
    print(f"\nExtracted {QUARTER_LABEL} for {len(CANONICAL_DISTRICTS)} HP districts.")
    print(f"Total numeric cells: {total_cells:,}")
    print(f"Distinct fields: {len(all_field_keys)}")
    print(f"Categories: {sorted(tables_by_cat.keys())}\n")
    print("Per-annexure:")
    for ak, cat, ndist, ncols in annex_summary:
        print(f"  {ak:10s} -> {cat:28s} {ndist} districts x {ncols} cols")
    print(f"\nOutputs written to:\n  {LOCAL_OUT}\n  {PUBLIC_OUT}")


if __name__ == "__main__":
    main()
