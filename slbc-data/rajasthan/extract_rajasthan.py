"""
Comprehensive extraction of district-level FI indicator data from Rajasthan SLBC
annexure Excel files.

Source: https://slbcrajasthan.in/Annexture.aspx
Two sets of files:
  - annex_1-12_2025.xlsx, annex_13-20_2025.xlsx, annex_21-33_2025.xlsx → Sep 2025
  - annex_1-12_2026.xlsx, annex_13-20_2026.xlsx, annex_21-33_2026.xlsx → Dec 2025

District-level sheets extracted:
  Annex 3  — Deposits, Advances, CD Ratio
  Annex 12 — ACP (Annual Credit Plan) — target vs achievement %
  Annex 12A — ACP Accounts — agriculture, MSME, priority sector
  Annex 12B — ACP Agriculture breakdown — Farm credit, KCC, infra
  Annex 12C — ACP MSME breakdown — Micro, Small, Medium
  Annex 12D — ACP Other Priority Sector — Export, Education, Housing, Renewable, Social Infra
  Annex 12E — Non-Priority Sector — Agriculture, Education, Housing, Personal, Others
  Annex 13A/13B (PMJDY district-wise) — accounts, deposits, RuPay, Aadhaar seeded
  Annex 13D/14A (Jan Suraksha district-wise) — PMSBY, PMJJBY enrollments
  Annex 14A/14C (APY district-wise) — enrollments, targets, cumulative
  Annex 15A — NWR Pledge Financing — disbursement and outstanding
  Annex 17A — NRLM (SHG) — targets, sanctioned, disbursed, pending
  Annex 18A — PMEGP — forwarded, sanctioned, MM claimed, disbursed, pending
  Annex 19A — Dr Ambedkar SC/ST scheme district-wise
  Annex 20A — PM-AJAY district-wise
  Annex 23A — MNSUPY district-wise
  Annex 25A — Vishwakarma district-wise (2026 only)
  Annex 28 — RSETI district-wise training/settlement
"""

import json
import os
import re
import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_number(val):
    """Convert Excel cell value to clean number string."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return str(round(val, 2))
    s = str(val).strip().replace(",", "")
    if s.startswith("="):
        return None
    try:
        return str(round(float(s), 2))
    except (ValueError, TypeError):
        return None


def clean_district_name(name):
    """Normalize district name to title case."""
    if not name or not isinstance(name, str):
        return None
    name = name.strip()
    # Strip leading serial numbers like "1)", "2)" etc
    name = re.sub(r"^\d+\)\s*", "", name)
    # Skip header/total/bank rows
    skip_exact = {
        "sr. no", "sr.no", "sr. no.", "sno", "s.no", "s. no",
        "grand total", "total", "district", "district name",
        "nationalized", "sub total", "name of bank", "banks",
        "nationalized banks", "name", "bank name", "others",
    }
    if name.lower().strip() in skip_exact:
        return None
    if name.lower().startswith("grand total"):
        return None
    if name.lower().startswith("total"):
        return None
    if name.lower().startswith("sub total"):
        return None
    if name.lower().startswith("state total"):
        return None
    if "rajasthan" in name.lower():
        return None
    return name.strip().title()


# District name normalization map
DISTRICT_ALIASES = {
    "Balotra": "Balotra",
    "Balotara": "Balotra",
    "Beawar": "Beawar",
    "Deeg": "Deeg",
    "Didwana Kuchaman": "Didwana-Kuchaman",
    "Didwana-Kuchaman": "Didwana-Kuchaman",
    "Khairthal-Tijara": "Khairthal-Tijara",
    "Khairthal Tijara": "Khairthal-Tijara",
    "Kotputli-Behror": "Kotputli-Behror",
    "Kotputli Behror": "Kotputli-Behror",
    "Phalodi": "Phalodi",
    "Salumbar": "Salumbar",
    "Sri Ganganagar": "Ganganagar",
    "Ganganagar": "Ganganagar",
    "Rajasamand": "Rajsamand",
    "Rajsamand": "Rajsamand",
    "Chittaurgarh": "Chittorgarh",
    "Chittorgarh": "Chittorgarh",
    "Chhitorgarh": "Chittorgarh",
    "Chittaurgarh": "Chittorgarh",
    "Budni": "Bundi",
    "Bundi": "Bundi",
    "Balekhan, Jaipur": None,  # RSETI sub-centre, skip
    "Jaipur Gramin": "Jaipur Gramin",
    "Jodhpur Gramin": "Jodhpur Gramin",
    "Dudu": "Dudu",
    "Gangapur City": "Gangapur City",
    "Kekri": "Kekri",
    "Neem Ka Thana": "Neem Ka Thana",
    "Shahpura": "Shahpura",
    "Sanchore": "Sanchore",
    "Anupgarh": "Anupgarh",
    "Bhiwadi": "Kotputli-Behror",  # Bhiwadi is in Kotputli-Behror district
    # Misspellings found in various sheets
    "Dungrpur": "Dungarpur",
    "Dungarpur": "Dungarpur",
    "Kotpulti-Behror": "Kotputli-Behror",
    "Kotputli Bahror": "Kotputli-Behror",
    "Kotputli Behroad": "Kotputli-Behror",
    "Kotputli- Behror": "Kotputli-Behror",
    "Nagour": "Nagaur",
    "Nagaur": "Nagaur",
    "Pratapgrah": "Pratapgarh",
    "Pratapgarh": "Pratapgarh",
    "Sawaimadhopur": "Sawai Madhopur",
    "Sawai Madhopur": "Sawai Madhopur",
    "Salumber": "Salumbar",
    "Slumber": "Salumbar",
    "Jaipur Rural": "Jaipur",  # PMEGP/Vishwakarma split Jaipur urban/rural, map to Jaipur
    "Jaipur Urban": "Jaipur",
    "Jalor": "Jalore",
    "Jalore": "Jalore",
    "Jhunjhunun": "Jhunjhunu",
    "Jhunjhunu": "Jhunjhunu",
    "Jodhpur (Rural)": "Jodhpur",
    "Karoli": "Karauli",
    "Karauli": "Karauli",
    "Raj Samand": "Rajsamand",
    "Chittaurgarh": "Chittorgarh",
    "Chittaurgarh ": "Chittorgarh",
    "Rajasthan Total": None,  # skip state total row
}


def normalize_district(name):
    """Apply district alias mapping after title-casing."""
    if not name:
        return None
    name = name.strip().title()
    if name in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[name]
    return name


def open_sheet(filepath, sheetname):
    """Open workbook and return worksheet. Returns None if sheet doesn't exist."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheetname not in wb.sheetnames:
            wb.close()
            return None, None
        return wb, wb[sheetname]
    except Exception as e:
        print(f"  ERROR opening {filepath}/{sheetname}: {e}")
        return None, None


def read_all_rows(ws, min_row=1, max_col=25):
    """Read all rows from a worksheet."""
    rows = []
    for row in ws.iter_rows(min_row=min_row, max_col=max_col, values_only=True):
        rows.append(list(row))
    return rows


def find_district_col(rows, max_check=10):
    """Find which column contains district names by checking first rows."""
    known_districts = {"ajmer", "alwar", "jaipur", "jodhpur", "udaipur", "kota",
                       "bikaner", "bhilwara", "barmer", "bundi", "sikar", "tonk",
                       "churu", "dausa", "nagaur", "pali", "baran"}
    for row in rows[:max_check]:
        for ci, cell in enumerate(row):
            if cell and isinstance(cell, str) and cell.strip().lower() in known_districts:
                return ci
    return 1  # default: column B


# ---------------------------------------------------------------------------
# Per-sheet extractors
# ---------------------------------------------------------------------------

def extract_annex3(filepath):
    """Annex 3 — Deposits, Advances, CD Ratio."""
    wb, ws = open_sheet(filepath, "Annex 3")
    if not ws:
        return {}
    districts = {}
    for row in ws.iter_rows(min_row=7, max_col=5, values_only=True):
        sr, district, deposit, advance, cd_ratio = row
        district = clean_district_name(str(district) if district else "")
        district = normalize_district(district)
        if not district:
            continue
        dep = parse_number(deposit)
        adv = parse_number(advance)
        if dep and adv and float(dep) > 0:
            cdr = str(round(float(adv) / float(dep) * 100, 2))
        else:
            cdr = parse_number(cd_ratio)
        if dep or adv:
            districts[district] = {
                "total_deposit": dep or "0",
                "total_advances": adv or "0",
                "overall_cd_ratio": cdr or "0",
            }
    wb.close()
    print(f"  Annex 3: {len(districts)} districts")
    return districts


def extract_annex12(filepath):
    """Annex 12 — ACP target vs achievement (amounts)."""
    wb, ws = open_sheet(filepath, "Annex 12")
    if not ws:
        return {}
    districts = {}
    # Columns: Sr, District, Agri Target, Agri Ach, Agri%, MSE Target, MSE Ach, MSE%,
    #   ME Target, ME Ach, ME%, Total MSME Target, Total MSME Ach, Total MSME%,
    #   OPS Target, OPS Ach, OPS%, Total PS Target, Total PS Ach, Total PS%
    for row in ws.iter_rows(min_row=10, max_col=20, values_only=True):
        cells = list(row)
        district = clean_district_name(str(cells[1]) if cells[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "acp_agri_target"), (3, "acp_agri_achievement"), (4, "acp_agri_pct"),
            (5, "acp_mse_target"), (6, "acp_mse_achievement"), (7, "acp_mse_pct"),
            (8, "acp_me_target"), (9, "acp_me_achievement"), (10, "acp_me_pct"),
            (11, "acp_msme_target"), (12, "acp_msme_achievement"), (13, "acp_msme_pct"),
            (14, "acp_ops_target"), (15, "acp_ops_achievement"), (16, "acp_ops_pct"),
            (17, "acp_total_ps_target"), (18, "acp_total_ps_achievement"), (19, "acp_total_ps_pct"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(cells[ci]) if ci < len(cells) else None
            if v is not None:
                # Convert fraction percentages to real percentages
                if fname.endswith("_pct") and v and float(v) <= 10:
                    v = str(round(float(v) * 100, 2))
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 12: {len(districts)} districts")
    return districts


def extract_annex12a(filepath):
    """Annex 12A — ACP accounts breakup: Agri, MSME, OPS, Total PS, Weaker Section."""
    wb, ws = open_sheet(filepath, "Annex 12A")
    if not ws:
        return {}
    districts = {}
    # Cols: Sr, Dist, Agri A/C, Agri AMT, MSE A/C, MSE AMT, ME A/C, ME AMT,
    #   Total MSME A/C, Total MSME AMT, OPS A/C, OPS AMT, Total PS A/C, Total PS AMT, Weaker A/C
    for row in ws.iter_rows(min_row=9, max_col=15, values_only=True):
        cells = list(row)
        district = clean_district_name(str(cells[1]) if cells[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "acp_agri_ac"), (3, "acp_agri_amt"),
            (4, "acp_mse_ac"), (5, "acp_mse_amt"),
            (6, "acp_me_ac"), (7, "acp_me_amt"),
            (8, "acp_msme_ac"), (9, "acp_msme_amt"),
            (10, "acp_ops_ac"), (11, "acp_ops_amt"),
            (12, "acp_total_ps_ac"), (13, "acp_total_ps_amt"),
            (14, "acp_weaker_section_ac"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(cells[ci]) if ci < len(cells) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 12A: {len(districts)} districts")
    return districts


def extract_annex12b(filepath):
    """Annex 12B — ACP Agriculture breakdown: Farm Credit (ST, TL, KCC-WC), Agri Infra, Ancillary."""
    wb, ws = open_sheet(filepath, "Annex 12B")
    if not ws:
        return {}
    districts = {}
    # Cols: Sr, Dist, ST/Crop A/C, ST/Crop AMT, TL A/C, TL AMT, KCC-WC A/C, KCC-WC AMT,
    #   Total Farm A/C, Total Farm AMT, Agri Infra A/C, Agri Infra AMT, Ancillary A/C, Ancillary AMT, Total Agri A/C
    for row in ws.iter_rows(min_row=10, max_col=16, values_only=True):
        cells = list(row)
        district = clean_district_name(str(cells[1]) if cells[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "acp_crop_ac"), (3, "acp_crop_amt"),
            (4, "acp_tl_ac"), (5, "acp_tl_amt"),
            (6, "acp_kcc_wc_ac"), (7, "acp_kcc_wc_amt"),
            (8, "acp_farm_credit_ac"), (9, "acp_farm_credit_amt"),
            (10, "acp_agri_infra_ac"), (11, "acp_agri_infra_amt"),
            (12, "acp_ancillary_ac"), (13, "acp_ancillary_amt"),
            (14, "acp_total_agri_ac"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(cells[ci]) if ci < len(cells) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 12B: {len(districts)} districts")
    return districts


def extract_annex12c(filepath):
    """Annex 12C — ACP MSME breakdown: Micro (Khadi/Village), Small, Others, Medium."""
    wb, ws = open_sheet(filepath, "Annex 12C")
    if not ws:
        return {}
    districts = {}
    for row in ws.iter_rows(min_row=9, max_col=12, values_only=True):
        cells = list(row)
        district = clean_district_name(str(cells[1]) if cells[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "acp_micro_ac"), (3, "acp_micro_amt"),
            (4, "acp_small_ac"), (5, "acp_small_amt"),
            (6, "acp_msme_others_ac"), (7, "acp_msme_others_amt"),
            (8, "acp_medium_ac"), (9, "acp_medium_amt"),
            (10, "acp_total_msme_ac"), (11, "acp_total_msme_amt"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(cells[ci]) if ci < len(cells) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 12C: {len(districts)} districts")
    return districts


def extract_annex12d(filepath):
    """Annex 12D — ACP Other Priority Sector: Export, Education, Housing, Renewable, Others, Social Infra."""
    wb, ws = open_sheet(filepath, "Annex 12D")
    if not ws:
        return {}
    districts = {}
    for row in ws.iter_rows(min_row=9, max_col=15, values_only=True):
        cells = list(row)
        district = clean_district_name(str(cells[1]) if cells[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "acp_export_ac"), (3, "acp_export_amt"),
            (4, "acp_education_ac"), (5, "acp_education_amt"),
            (6, "acp_housing_ac"), (7, "acp_housing_amt"),
            (8, "acp_renewable_ac"), (9, "acp_renewable_amt"),
            (10, "acp_others_ac"), (11, "acp_others_amt"),
            (12, "acp_social_infra_ac"), (13, "acp_social_infra_amt"),
            (14, "acp_total_ops_ac"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(cells[ci]) if ci < len(cells) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 12D: {len(districts)} districts")
    return districts


def extract_annex12e(filepath):
    """Annex 12E — Non-Priority Sector: Agriculture, Education, Housing, Personal, Others."""
    wb, ws = open_sheet(filepath, "Annex 12E")
    if not ws:
        return {}
    districts = {}
    # Col B is "Banks" but actually district names
    for row in ws.iter_rows(min_row=8, max_col=14, values_only=True):
        cells = list(row)
        district = clean_district_name(str(cells[1]) if cells[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "nps_agri_ac"), (3, "nps_agri_amt"),
            (4, "nps_education_ac"), (5, "nps_education_amt"),
            (6, "nps_housing_ac"), (7, "nps_housing_amt"),
            (8, "nps_personal_ac"), (9, "nps_personal_amt"),
            (10, "nps_others_ac"), (11, "nps_others_amt"),
            (12, "nps_total_ac"), (13, "nps_total_amt"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(cells[ci]) if ci < len(cells) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 12E: {len(districts)} districts")
    return districts


def extract_pmjdy_district(filepath, sheetname):
    """PMJDY district-wise: 13B (2025 set) or 13A (2026 set).
    Cols: Sr, District, Rural A/C, Urban A/C, Total A/C, Total Deposit,
          Zero Balance, %Zero, RuPay Issued, %RuPay, Aadhaar Seeded, %Aadhaar
    """
    wb, ws = open_sheet(filepath, sheetname)
    if not ws:
        return {}
    districts = {}
    for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
        cells = list(row)
        district = clean_district_name(str(cells[1]) if cells[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "pmjdy_rural_ac"), (3, "pmjdy_urban_ac"), (4, "pmjdy_total_ac"),
            (5, "pmjdy_total_deposit"), (6, "pmjdy_zero_balance"),
            (7, "pmjdy_zero_balance_pct"), (8, "pmjdy_rupay_issued"),
            (9, "pmjdy_rupay_pct"), (10, "pmjdy_aadhaar_seeded"),
            (11, "pmjdy_aadhaar_pct"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(cells[ci]) if ci < len(cells) else None
            if v is not None:
                # Convert fraction percentages
                if fname.endswith("_pct") and v and float(v) <= 1.5:
                    v = str(round(float(v) * 100, 2))
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  {sheetname} (PMJDY): {len(districts)} districts")
    return districts


def extract_jan_suraksha_district(filepath, sheetname):
    """Jan Suraksha district-wise: 13D (2025 set) or 14A (2026 set).
    Cols: Sr, District, PMSBY, PMJJBY, Total
    """
    wb, ws = open_sheet(filepath, sheetname)
    if not ws:
        return {}
    districts = {}
    # Find where district data starts by looking for a known district
    rows = read_all_rows(ws, min_row=1, max_col=6)
    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().lower() in (
            "ajmer", "alwar", "banswara", "baran"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [(2, "jan_suraksha_pmsby"), (3, "jan_suraksha_pmjjby"), (4, "jan_suraksha_total")]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  {sheetname} (Jan Suraksha): {len(districts)} districts")
    return districts


def extract_apy_district(filepath, sheetname):
    """APY district-wise: 14A (2025 set) or 14C (2026 set).
    Cols: Sr, District, Lead Bank, Branches, Annual Target, Current FY Achievement,
          Target achievement %, Cumulative
    """
    wb, ws = open_sheet(filepath, sheetname)
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=9)
    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BANSWARA", "BARAN"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (3, "apy_branches"), (4, "apy_annual_target"),
            (5, "apy_current_fy_achievement"), (6, "apy_target_pct"),
            (7, "apy_cumulative"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                if fname == "apy_target_pct" and v and float(v) <= 10:
                    v = str(round(float(v) * 100, 2))
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  {sheetname} (APY): {len(districts)} districts")
    return districts


def extract_annex15a(filepath):
    """Annex 15A — NWR Pledge Financing.
    Cols: Sr, District, Disbursement No, Disbursement Amt, Outstanding No, Outstanding Amt
    """
    wb, ws = open_sheet(filepath, "Annex 15A")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=7)
    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BANSWARA"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "nwr_disbursement_no"), (3, "nwr_disbursement_amt"),
            (4, "nwr_outstanding_no"), (5, "nwr_outstanding_amt"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 15A: {len(districts)} districts")
    return districts


def extract_annex17a(filepath):
    """Annex 17A — NRLM (SHG) district-wise.
    2026: Sr, District, Targets, App Forwarded, Sanctioned, Disbursed, Total Pending, >14d, >30d, >90d
    2025: Sr, District, Targets, App Forwarded, Sanctioned, Disbursed, Total Pending, >14d, >30d, >90d
    """
    wb, ws = open_sheet(filepath, "Annex 17A")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=10)
    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BALOTRA ", "BANSWARA"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "nrlm_targets"), (3, "nrlm_app_forwarded"),
            (4, "nrlm_sanctioned"), (5, "nrlm_disbursed"),
            (6, "nrlm_total_pending"),
            (7, "nrlm_pending_gt14d"), (8, "nrlm_pending_gt30d"), (9, "nrlm_pending_gt90d"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 17A: {len(districts)} districts")
    return districts


def extract_annex18a(filepath):
    """Annex 18A — PMEGP district-wise.
    2026 cols: Sr, Name, Forwarded to Bank, Sanctioned No, Sanctioned MM,
              EDP Online, EDP Offline, MM Claimed No, MM Claimed Amt,
              MM Disbursed No, MM Disbursed Amt, Returned, Pending No, Pending Amt, Pending Rectification
    2025 cols: Sr, Name, Target MM, Forwarded No, Forwarded Amt, Sanctioned No, Sanctioned Amt,
              MM Claimed No, MM Claimed Amt, MM Disbursed No, MM Disbursed Amt,
              Returned No, Returned Amt, Pending No, Pending Amt
    """
    wb, ws = open_sheet(filepath, "Annex 18A")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=16)

    # Detect which format based on header row
    is_2026_format = False
    for row in rows[:5]:
        for cell in row:
            if cell and isinstance(cell, str) and "EDP" in str(cell):
                is_2026_format = True
                break

    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BANSWARA"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}

    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        if is_2026_format:
            fields = [
                (2, "pmegp_forwarded_to_bank"),
                (3, "pmegp_sanctioned_no"), (4, "pmegp_sanctioned_mm"),
                (7, "pmegp_mm_claimed_no"), (8, "pmegp_mm_claimed_amt"),
                (9, "pmegp_mm_disbursed_no"), (10, "pmegp_mm_disbursed_amt"),
                (11, "pmegp_returned"), (12, "pmegp_pending_no"), (13, "pmegp_pending_amt"),
            ]
        else:
            fields = [
                (3, "pmegp_forwarded_no"), (4, "pmegp_forwarded_amt"),
                (5, "pmegp_sanctioned_no"), (6, "pmegp_sanctioned_amt"),
                (7, "pmegp_mm_claimed_no"), (8, "pmegp_mm_claimed_amt"),
                (9, "pmegp_mm_disbursed_no"), (10, "pmegp_mm_disbursed_amt"),
                (11, "pmegp_returned_no"), (12, "pmegp_returned_amt"),
                (13, "pmegp_pending_no"), (14, "pmegp_pending_amt"),
            ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 18A: {len(districts)} districts")
    return districts


def extract_annex19a(filepath):
    """Annex 19A — Dr Ambedkar SC/ST scheme district-wise.
    Cols: Sr, District, Targets, App Forwarded, Sanctioned, Achieved% (or Disbursed),
          Disbursed (or Total Pending), Total Pending, >14d, >30d, >90d
    """
    wb, ws = open_sheet(filepath, "Annex 19A")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=11)

    # Detect format: 2026 has "Achieved In %age" at col 5, 2025 has "Disbursed" at col 5
    header_row = rows[1] if len(rows) > 1 else []
    has_pct_col = any(c and isinstance(c, str) and "%" in str(c) for c in header_row)

    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BALOTRA ", "BANSWARA"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        if has_pct_col:
            # 2026 format: Sr, Dist, Targets, Fwd, Sanctioned, Ach%, Disbursed, Pending, >14, >30, >90
            fields = [
                (2, "ambedkar_targets"), (3, "ambedkar_app_forwarded"),
                (4, "ambedkar_sanctioned"), (5, "ambedkar_ach_pct"),
                (6, "ambedkar_disbursed"), (7, "ambedkar_total_pending"),
                (8, "ambedkar_pending_gt14d"), (9, "ambedkar_pending_gt30d"),
                (10, "ambedkar_pending_gt90d"),
            ]
        else:
            # 2025 format: Sr, Dist, Targets, Fwd, Sanctioned, Disbursed, Pending, >14, >30, >90, Ach%
            fields = [
                (2, "ambedkar_targets"), (3, "ambedkar_app_forwarded"),
                (4, "ambedkar_sanctioned"), (5, "ambedkar_disbursed"),
                (6, "ambedkar_total_pending"),
                (7, "ambedkar_pending_gt14d"), (8, "ambedkar_pending_gt30d"),
                (9, "ambedkar_pending_gt90d"), (10, "ambedkar_ach_pct"),
            ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                if fname == "ambedkar_ach_pct" and v and float(v) <= 10:
                    v = str(round(float(v) * 100, 2))
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 19A: {len(districts)} districts")
    return districts


def extract_annex20a(filepath):
    """Annex 20A — PM-AJAY district-wise.
    Cols: Sr, District, Targets, App Forwarded, Sanctioned, Disbursed, Total Pending, >14d, >30d, >90d
    """
    wb, ws = open_sheet(filepath, "Annex 20A")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=10)
    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BALOTRA ", "BANSWARA"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "pmajay_targets"), (3, "pmajay_app_forwarded"),
            (4, "pmajay_sanctioned"), (5, "pmajay_disbursed"),
            (6, "pmajay_total_pending"),
            (7, "pmajay_pending_gt14d"), (8, "pmajay_pending_gt30d"),
            (9, "pmajay_pending_gt90d"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 20A: {len(districts)} districts")
    return districts


def extract_annex23a(filepath):
    """Annex 23A — MNSUPY district-wise.
    2026 cols: Sr, District, Targets, Forwarded(FI), Rejected(FI), Approved(FI), Ach%, Disbursed, Subsidy, Pending
    2025 cols: Sr, District, Targets, Fwd, Sanctioned, Disbursed, Pending, >14d, >30d, >90d, Ach%
    """
    wb, ws = open_sheet(filepath, "Annex 23A")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=11)

    # Detect format: 2026 has "Forwarded (To FI)" / "Rejected"
    is_2026 = False
    for row in rows[:4]:
        for cell in row:
            if cell and isinstance(cell, str) and ("Rejected" in str(cell) or "Approved" in str(cell) or "Subsidy" in str(cell)):
                is_2026 = True
                break

    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BALOTRA ", "BALOTARA", "BANSWARA"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        if is_2026:
            fields = [
                (2, "mnsupy_targets"), (3, "mnsupy_forwarded"),
                (4, "mnsupy_rejected"), (5, "mnsupy_approved"),
                (6, "mnsupy_ach_pct"), (7, "mnsupy_disbursed"),
                (8, "mnsupy_subsidy"), (9, "mnsupy_pending"),
            ]
        else:
            fields = [
                (2, "mnsupy_targets"), (3, "mnsupy_forwarded"),
                (4, "mnsupy_sanctioned"), (5, "mnsupy_disbursed"),
                (6, "mnsupy_total_pending"),
                (7, "mnsupy_pending_gt14d"), (8, "mnsupy_pending_gt30d"),
                (9, "mnsupy_pending_gt90d"), (10, "mnsupy_ach_pct"),
            ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                if fname == "mnsupy_ach_pct" and v and float(v) <= 10:
                    v = str(round(float(v) * 100, 2))
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 23A: {len(districts)} districts")
    return districts


def extract_annex25a(filepath):
    """Annex 25A — Vishwakarma (VYUPY) district-wise (2026 only).
    Cols: Sr, DIC Name, Financial Target, Recd in DIC, Forwarded FI, Avg Forward,
          Pending FI, Sanctioned FI, Disbursed FI
    """
    wb, ws = open_sheet(filepath, "Annex 25A")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=10)
    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().upper() in (
            "AJMER", "ALWAR", "BALOTRA", "BANSWARA"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (2, "vishwakarma_target_amt"), (3, "vishwakarma_recd_dic"),
            (4, "vishwakarma_forwarded_fi"), (5, "vishwakarma_avg_forward_pct"),
            (6, "vishwakarma_pending_fi"), (7, "vishwakarma_sanctioned"),
            (8, "vishwakarma_disbursed"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                if fname == "vishwakarma_avg_forward_pct" and v and float(v) <= 2:
                    v = str(round(float(v) * 100, 2))
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 25A: {len(districts)} districts")
    return districts


def extract_annex28(filepath):
    """Annex 28 — RSETI district-wise training/settlement.
    Cols: Sr, Name, Sponsor Bank, Target Progs, Target Candidates,
          Progs Conducted, Candidates Trained, Settled,
          Self Emp, Wage Emp, With Bank Finance, With Self Finance,
          % Settled, % Credit Linkage
    """
    wb, ws = open_sheet(filepath, "Annex 28")
    if not ws:
        return {}
    districts = {}
    rows = read_all_rows(ws, min_row=1, max_col=14)
    data_start = 0
    for i, row in enumerate(rows):
        if row[1] and isinstance(row[1], str) and row[1].strip().title() in (
            "Ajmer", "Alwar", "Banswara", "Bundi", "Chittorgarh", "Churu"):
            data_start = i
            break
    if data_start == 0:
        wb.close()
        return {}
    for row in rows[data_start:]:
        district = clean_district_name(str(row[1]) if row[1] else "")
        district = normalize_district(district)
        if not district:
            continue
        d = {}
        fields = [
            (3, "rseti_target_progs"), (4, "rseti_target_candidates"),
            (5, "rseti_progs_conducted"), (6, "rseti_candidates_trained"),
            (7, "rseti_settled"), (8, "rseti_self_employment"),
            (9, "rseti_wage_employment"), (10, "rseti_bank_finance"),
            (11, "rseti_self_finance"),
            (12, "rseti_settled_pct"), (13, "rseti_credit_linkage_pct"),
        ]
        has_data = False
        for ci, fname in fields:
            v = parse_number(row[ci]) if ci < len(row) else None
            if v is not None:
                if fname.endswith("_pct") and v and float(v) <= 2:
                    v = str(round(float(v) * 100, 2))
                d[fname] = v
                has_data = True
        if has_data:
            districts[district] = d
    wb.close()
    print(f"  Annex 28: {len(districts)} districts")
    return districts


# ---------------------------------------------------------------------------
# Quarter configurations
# ---------------------------------------------------------------------------

QUARTERS = [
    {
        "quarter_key": "2025-09",
        "period": "September 2025",
        "as_on_date": "30-09-2025",
        "fy": "2025-26",
        "files": {
            "1-12": "annex_1-12_2025.xlsx",
            "13-20": "annex_13-20_2025.xlsx",
            "21-33": "annex_21-33_2025.xlsx",
        },
        "pmjdy_sheet": "Annex 13B",  # district-wise PMJDY in 2025 set
        "jan_suraksha_sheet": "Annex 13D",
        "apy_sheet": "Annex 14A",
        "has_vishwakarma_25a": False,
    },
    {
        "quarter_key": "2025-12",
        "period": "December 2025",
        "as_on_date": "31-12-2025",
        "fy": "2025-26",
        "files": {
            "1-12": "annex_1-12_2026.xlsx",
            "13-20": "annex_13-20_2026.xlsx",
            "21-33": "annex_21-33_2026.xlsx",
        },
        "pmjdy_sheet": "Annex 13A",  # district-wise PMJDY in 2026 set
        "jan_suraksha_sheet": "Annex 14A",
        "apy_sheet": "Annex 14C",
        "has_vishwakarma_25a": True,
    },
]


# ---------------------------------------------------------------------------
# Category table definitions for output
# ---------------------------------------------------------------------------

CATEGORY_DEFS = {
    "credit_deposit_ratio": {
        "display_name": "Credit Deposit Ratio",
        "extractor": "annex3",
        "fields_order": ["total_deposit", "total_advances", "overall_cd_ratio"],
    },
    "acp_summary": {
        "display_name": "Annual Credit Plan — Summary",
        "extractor": "annex12",
        "fields_order": [
            "acp_agri_target", "acp_agri_achievement", "acp_agri_pct",
            "acp_mse_target", "acp_mse_achievement", "acp_mse_pct",
            "acp_me_target", "acp_me_achievement", "acp_me_pct",
            "acp_msme_target", "acp_msme_achievement", "acp_msme_pct",
            "acp_ops_target", "acp_ops_achievement", "acp_ops_pct",
            "acp_total_ps_target", "acp_total_ps_achievement", "acp_total_ps_pct",
        ],
    },
    "acp_accounts": {
        "display_name": "Annual Credit Plan — Accounts",
        "extractor": "annex12a",
        "fields_order": [
            "acp_agri_ac", "acp_agri_amt", "acp_mse_ac", "acp_mse_amt",
            "acp_me_ac", "acp_me_amt", "acp_msme_ac", "acp_msme_amt",
            "acp_ops_ac", "acp_ops_amt", "acp_total_ps_ac", "acp_total_ps_amt",
            "acp_weaker_section_ac",
        ],
    },
    "acp_agriculture": {
        "display_name": "Annual Credit Plan — Agriculture Breakdown",
        "extractor": "annex12b",
        "fields_order": [
            "acp_crop_ac", "acp_crop_amt", "acp_tl_ac", "acp_tl_amt",
            "acp_kcc_wc_ac", "acp_kcc_wc_amt", "acp_farm_credit_ac", "acp_farm_credit_amt",
            "acp_agri_infra_ac", "acp_agri_infra_amt", "acp_ancillary_ac", "acp_ancillary_amt",
            "acp_total_agri_ac",
        ],
    },
    "acp_msme": {
        "display_name": "Annual Credit Plan — MSME Breakdown",
        "extractor": "annex12c",
        "fields_order": [
            "acp_micro_ac", "acp_micro_amt", "acp_small_ac", "acp_small_amt",
            "acp_msme_others_ac", "acp_msme_others_amt", "acp_medium_ac", "acp_medium_amt",
            "acp_total_msme_ac", "acp_total_msme_amt",
        ],
    },
    "acp_other_priority": {
        "display_name": "Annual Credit Plan — Other Priority Sector",
        "extractor": "annex12d",
        "fields_order": [
            "acp_export_ac", "acp_export_amt", "acp_education_ac", "acp_education_amt",
            "acp_housing_ac", "acp_housing_amt", "acp_renewable_ac", "acp_renewable_amt",
            "acp_others_ac", "acp_others_amt", "acp_social_infra_ac", "acp_social_infra_amt",
            "acp_total_ops_ac",
        ],
    },
    "non_priority_sector": {
        "display_name": "Non-Priority Sector Lending",
        "extractor": "annex12e",
        "fields_order": [
            "nps_agri_ac", "nps_agri_amt", "nps_education_ac", "nps_education_amt",
            "nps_housing_ac", "nps_housing_amt", "nps_personal_ac", "nps_personal_amt",
            "nps_others_ac", "nps_others_amt", "nps_total_ac", "nps_total_amt",
        ],
    },
    "pmjdy": {
        "display_name": "PM Jan Dhan Yojana",
        "extractor": "pmjdy",
        "fields_order": [
            "pmjdy_rural_ac", "pmjdy_urban_ac", "pmjdy_total_ac",
            "pmjdy_total_deposit", "pmjdy_zero_balance", "pmjdy_zero_balance_pct",
            "pmjdy_rupay_issued", "pmjdy_rupay_pct",
            "pmjdy_aadhaar_seeded", "pmjdy_aadhaar_pct",
        ],
    },
    "jan_suraksha": {
        "display_name": "Jan Suraksha Schemes (PMSBY + PMJJBY)",
        "extractor": "jan_suraksha",
        "fields_order": ["jan_suraksha_pmsby", "jan_suraksha_pmjjby", "jan_suraksha_total"],
    },
    "apy": {
        "display_name": "Atal Pension Yojana (APY)",
        "extractor": "apy",
        "fields_order": [
            "apy_branches", "apy_annual_target", "apy_current_fy_achievement",
            "apy_target_pct", "apy_cumulative",
        ],
    },
    "nwr_pledge": {
        "display_name": "NWR Pledge Financing",
        "extractor": "annex15a",
        "fields_order": [
            "nwr_disbursement_no", "nwr_disbursement_amt",
            "nwr_outstanding_no", "nwr_outstanding_amt",
        ],
    },
    "nrlm": {
        "display_name": "NRLM (SHG Bank Linkage)",
        "extractor": "annex17a",
        "fields_order": [
            "nrlm_targets", "nrlm_app_forwarded", "nrlm_sanctioned", "nrlm_disbursed",
            "nrlm_total_pending", "nrlm_pending_gt14d", "nrlm_pending_gt30d", "nrlm_pending_gt90d",
        ],
    },
    "pmegp": {
        "display_name": "PMEGP (PM Employment Generation Programme)",
        "extractor": "annex18a",
        "fields_order": [
            "pmegp_forwarded_to_bank", "pmegp_forwarded_no", "pmegp_forwarded_amt",
            "pmegp_sanctioned_no", "pmegp_sanctioned_mm", "pmegp_sanctioned_amt",
            "pmegp_mm_claimed_no", "pmegp_mm_claimed_amt",
            "pmegp_mm_disbursed_no", "pmegp_mm_disbursed_amt",
            "pmegp_returned", "pmegp_returned_no", "pmegp_returned_amt",
            "pmegp_pending_no", "pmegp_pending_amt",
        ],
    },
    "ambedkar_scheme": {
        "display_name": "Dr Bhimrao Ambedkar SC/ST Scheme",
        "extractor": "annex19a",
        "fields_order": [
            "ambedkar_targets", "ambedkar_app_forwarded", "ambedkar_sanctioned",
            "ambedkar_ach_pct", "ambedkar_disbursed", "ambedkar_total_pending",
            "ambedkar_pending_gt14d", "ambedkar_pending_gt30d", "ambedkar_pending_gt90d",
        ],
    },
    "pmajay": {
        "display_name": "PM-AJAY",
        "extractor": "annex20a",
        "fields_order": [
            "pmajay_targets", "pmajay_app_forwarded", "pmajay_sanctioned",
            "pmajay_disbursed", "pmajay_total_pending",
            "pmajay_pending_gt14d", "pmajay_pending_gt30d", "pmajay_pending_gt90d",
        ],
    },
    "mnsupy": {
        "display_name": "MNSUPY (Mukhyamantri Nishulk Udyam Protsahan Yojana)",
        "extractor": "annex23a",
        "fields_order": [
            "mnsupy_targets", "mnsupy_forwarded", "mnsupy_rejected", "mnsupy_approved",
            "mnsupy_sanctioned", "mnsupy_ach_pct", "mnsupy_disbursed",
            "mnsupy_subsidy", "mnsupy_pending", "mnsupy_total_pending",
            "mnsupy_pending_gt14d", "mnsupy_pending_gt30d", "mnsupy_pending_gt90d",
        ],
    },
    "vishwakarma": {
        "display_name": "Vishwakarma Yuva Udyami Protsahan Yojana",
        "extractor": "annex25a",
        "fields_order": [
            "vishwakarma_target_amt", "vishwakarma_recd_dic", "vishwakarma_forwarded_fi",
            "vishwakarma_avg_forward_pct", "vishwakarma_pending_fi",
            "vishwakarma_sanctioned", "vishwakarma_disbursed",
        ],
    },
    "rseti": {
        "display_name": "RSETI (Rural Self Employment Training Institute)",
        "extractor": "annex28",
        "fields_order": [
            "rseti_target_progs", "rseti_target_candidates",
            "rseti_progs_conducted", "rseti_candidates_trained",
            "rseti_settled", "rseti_self_employment", "rseti_wage_employment",
            "rseti_bank_finance", "rseti_self_finance",
            "rseti_settled_pct", "rseti_credit_linkage_pct",
        ],
    },
}


# ---------------------------------------------------------------------------
# Main extraction pipeline
# ---------------------------------------------------------------------------

def extract_all_for_quarter(qinfo):
    """Extract all district-level data for one quarter."""
    f112 = qinfo["files"]["1-12"]
    f1320 = qinfo["files"]["13-20"]
    f2133 = qinfo["files"]["21-33"]

    print(f"\n{'='*60}")
    print(f"Quarter: {qinfo['period']} ({qinfo['quarter_key']})")
    print(f"  Files: {f112}, {f1320}, {f2133}")
    print(f"{'='*60}")

    data = {}

    # Annex 3 — CD Ratio
    data["credit_deposit_ratio"] = extract_annex3(f112)

    # Annex 12 series — ACP
    data["acp_summary"] = extract_annex12(f112)
    data["acp_accounts"] = extract_annex12a(f112)
    data["acp_agriculture"] = extract_annex12b(f112)
    data["acp_msme"] = extract_annex12c(f112)
    data["acp_other_priority"] = extract_annex12d(f112)
    data["non_priority_sector"] = extract_annex12e(f112)

    # PMJDY
    data["pmjdy"] = extract_pmjdy_district(f1320, qinfo["pmjdy_sheet"])

    # Jan Suraksha
    data["jan_suraksha"] = extract_jan_suraksha_district(f1320, qinfo["jan_suraksha_sheet"])

    # APY
    data["apy"] = extract_apy_district(f1320, qinfo["apy_sheet"])

    # NWR
    data["nwr_pledge"] = extract_annex15a(f1320)

    # NRLM
    data["nrlm"] = extract_annex17a(f1320)

    # PMEGP
    data["pmegp"] = extract_annex18a(f1320)

    # Ambedkar scheme
    data["ambedkar_scheme"] = extract_annex19a(f1320)

    # PM-AJAY
    data["pmajay"] = extract_annex20a(f1320)

    # MNSUPY
    data["mnsupy"] = extract_annex23a(f2133)

    # Vishwakarma (2026 only)
    if qinfo["has_vishwakarma_25a"]:
        data["vishwakarma"] = extract_annex25a(f2133)
    else:
        data["vishwakarma"] = {}

    # RSETI
    data["rseti"] = extract_annex28(f2133)

    return data


def build_complete_json(all_quarters):
    """Build rajasthan_complete.json."""
    result = {
        "source": "SLBC Rajasthan (State Level Bankers' Committee, Rajasthan)",
        "state": "Rajasthan",
        "amount_unit": "Rs. Lakhs",
        "quarters": {},
    }

    for q in all_quarters:
        tables = {}
        for cat_key, cat_def in CATEGORY_DEFS.items():
            cat_data = q["data"].get(cat_key, {})
            if not cat_data:
                continue
            # Collect all fields actually present
            all_fields = set()
            for dist_data in cat_data.values():
                all_fields.update(dist_data.keys())
            # Order fields: use defined order, then any extras
            ordered = [f for f in cat_def["fields_order"] if f in all_fields]
            extras = sorted(all_fields - set(ordered))
            fields = ordered + extras

            tables[cat_key] = {
                "fields": fields,
                "districts": cat_data,
            }

        result["quarters"][q["quarter_key"]] = {
            "period": q["period"],
            "as_on_date": q["as_on_date"],
            "fy": q["fy"],
            "tables": tables,
        }

    return result


def build_timeseries_json(all_quarters):
    """Build rajasthan_fi_timeseries.json."""
    periods = []

    for q in all_quarters:
        # Collect all districts across all categories
        all_districts = set()
        for cat_data in q["data"].values():
            all_districts.update(cat_data.keys())

        district_rows = []
        for dist_name in sorted(all_districts):
            row = {
                "district": dist_name,
                "period": q["period"],
            }
            # Flatten all categories into the row
            for cat_key, cat_data in q["data"].items():
                if dist_name in cat_data:
                    for field, value in cat_data[dist_name].items():
                        row[f"{cat_key}__{field}"] = value

            district_rows.append(row)

        periods.append({
            "period": q["period"],
            "districts": district_rows,
        })

    return {"periods": periods}


def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    all_quarters = []
    for qinfo in QUARTERS:
        # Check files exist
        missing = False
        for fkey, fname in qinfo["files"].items():
            if not os.path.exists(fname):
                print(f"MISSING: {fname}")
                missing = True
        if missing:
            continue

        data = extract_all_for_quarter(qinfo)
        all_quarters.append({
            "quarter_key": qinfo["quarter_key"],
            "period": qinfo["period"],
            "as_on_date": qinfo["as_on_date"],
            "fy": qinfo["fy"],
            "data": data,
        })

    # Sort chronologically
    all_quarters.sort(key=lambda q: q["quarter_key"])

    # Build outputs
    complete = build_complete_json(all_quarters)
    timeseries = build_timeseries_json(all_quarters)

    # Write
    with open("rajasthan_complete.json", "w") as f:
        json.dump(complete, f, indent=2, ensure_ascii=False)
    print(f"\nWrote rajasthan_complete.json")

    with open("rajasthan_fi_timeseries.json", "w") as f:
        json.dump(timeseries, f, indent=2, ensure_ascii=False)
    print(f"Wrote rajasthan_fi_timeseries.json")

    # Summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    for q in all_quarters:
        all_dists = set()
        for cat_data in q["data"].values():
            all_dists.update(cat_data.keys())
        print(f"\n{q['period']} ({q['quarter_key']}):")
        print(f"  Total unique districts: {len(all_dists)}")
        for cat_key in CATEGORY_DEFS:
            cat_data = q["data"].get(cat_key, {})
            if cat_data:
                n_fields = sum(len(v) for v in cat_data.values()) // max(len(cat_data), 1)
                print(f"  {cat_key}: {len(cat_data)} districts, ~{n_fields} fields each")


if __name__ == "__main__":
    main()
