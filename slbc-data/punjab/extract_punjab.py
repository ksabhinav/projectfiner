#!/usr/bin/env python3
"""
Punjab SLBC Data Extractor
==========================

Punjab's SLBC convenor is Punjab National Bank. Source:
  https://slbcpunjab.pnb.bank.in/slbc-agenda/

GRANULARITY FINDING (DO NOT EXPAND WITHOUT CHECKING SOURCE):
  - The "Basic Banking Data" Annexures (Annexure-3 / -16 / -20) are BANK-WISE
    only — they list ~30+ banks' Deposits / Advances / Branches / CD-ratio
    for the state aggregate. They do NOT break out by Punjab's 23 districts.
  - The only consistently district-wise quarterly data Punjab publishes is the
    "ACP District-wise" Annexure (Annual Credit Plan target-vs-achievement)
    which breaks out Priority Sector advances by district into 4 buckets:
      1. Agriculture & Allied
      2. MSME (Manufacturing + Services)
      3. Other Priority Sector
      4. Total Priority Sector (1+2+3)
    Each bucket carries: Target, Achievement, % Achievement.
  - Per-district CD-ratio mini-tables exist for ~6 Doaba districts (Jalandhar,
    Kapurthala, SBS Nagar, Hoshiarpur, Pathankot, Rupnagar) in the 2022/03
    annexures, but these are NOT a full Punjab CD-ratio table.

Punjab is therefore a PRIORITY-SECTOR-DISTRICT-WISE state in FINER (analogous
to Madhya Pradesh which is CD-ratio-only). The CD-ratio / deposits / advances
/ branches choropleths will not have Punjab district data — those metrics are
state-aggregate only in the source.

District count caveat:
  - Pre-May-2021 source files show 22 districts (no Malerkotla).
  - Post-May-2021 source files show 23 districts (Malerkotla carved from
    Sangrur).
  Both are valid for their respective periods.

Unit: Source ACP tables clearly state "Amount in cr" (Crores).
FINER canonical unit is Lakhs => *100 conversion applied to all monetary
fields (Target_amt, Achievement_amt). Percentage achievement is unitless
and not converted.

Output:
  punjab_complete.json
  punjab_fi_timeseries.json
"""

import json
import os
import re
import warnings

import openpyxl

warnings.filterwarnings('ignore')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOADS_DIR = os.path.join(BASE_DIR, 'downloads')
PUBLIC_DIR = os.path.abspath(
    os.path.join(BASE_DIR, '..', '..', 'public', 'slbc-data', 'punjab')
)


# ---------------------------------------------------------------------------
# Quarter mapping per ACP file (parsed from each file's title row)
# Keys: filename in downloads/. Value: (quarter_label, period_code, end_date)
# ---------------------------------------------------------------------------

ACP_FILES = {
    'acp_2022_03_Annexure-27-ACP-Districtwise.xlsx':
        ('September 2020', '2020-09', '30-09-2020'),
    'acp_2022_03_Annexure-22-ACP-Districtwise.xlsx':
        ('December 2020', '2020-12', '31-12-2020'),
    'acp_2022_03_Annexure-21-ACP-Districtwise.xlsx':
        ('March 2021', '2021-03', '31-03-2021'),
    'acp_2022_04_Annexure-21-ACP-Districtwise.xlsx':
        ('September 2021', '2021-09', '30-09-2021'),
    'acp_2022_04_Annexure-18-ACP-Districtwise.xlsx':
        ('December 2021', '2021-12', '31-12-2021'),
    'acp_2023_03_Annexure-31-ACP-Districtwise-1.xlsx':
        ('March 2022', '2022-03', '31-03-2022'),
    'acp_2023_03_Annexure-32-ACP-Districtwise-1.xlsx':
        ('June 2022', '2022-06', '30-06-2022'),
    'acp_2023_03_Annexure-29-ACP-District-wise.xlsx':
        ('September 2022', '2022-09', '30-09-2022'),
    'acp_2023_04_Annexure-20-ACP-District-wise.xlsx':
        ('December 2022', '2022-12', '31-12-2022'),
    'acp_2023_06_Annexure-20-ACP-District-wise.xlsx':
        ('March 2023', '2023-03', '31-03-2023'),
    'acp_2023_08_Annexure-2-ACP-District-wise.xlsx':
        ('June 2023', '2023-06', '30-06-2023'),
    'acp_2023_12_Annexure-2-ACP-District-wise.xlsx':
        ('September 2023', '2023-09', '30-09-2023'),
    'acp_2024_03.xlsx':
        ('December 2023', '2023-12', '31-12-2023'),
    'acp_2024_09.xlsx':
        ('June 2024', '2024-06', '30-06-2024'),
    'acp_2024_11.xlsx':
        ('September 2024', '2024-09', '30-09-2024'),
    'acp_2025_02.xlsx':
        ('December 2024', '2024-12', '31-12-2024'),
    'acp_2025_05.xlsx':
        ('March 2025', '2025-03', '31-03-2025'),
    'acp_2025_08.xlsx':
        ('June 2025', '2025-06', '30-06-2025'),
    'acp_2026_02.xlsx':
        ('December 2025', '2025-12', '31-12-2025'),
}


# ---------------------------------------------------------------------------
# Districts + aliases
# ---------------------------------------------------------------------------

# Canonical names match FINER SQLite `districts.name` for Punjab (state_lgd=3).
# Malerkotla carved from Sangrur in May 2021.
PUNJAB_DISTRICTS = [
    "Amritsar", "Barnala", "Bathinda", "Faridkot", "Fatehgarh Sahib",
    "Fazilka", "Ferozepur", "Gurdaspur", "Hoshiarpur", "Jalandhar",
    "Kapurthala", "Ludhiana", "Malerkotla", "Mansa", "Moga",
    "S.a.s Nagar", "Sri Muktsar Sahib", "Pathankot", "Patiala", "Rupnagar",
    "Sangrur", "Shahid Bhagat Singh Nagar", "Tarn Taran",
]

DISTRICT_ALIASES_UPPER = {
    # S.a.s Nagar (Mohali / SAS Nagar)
    'SAS NAGAR': 'S.a.s Nagar',
    'S.A.S. NAGAR': 'S.a.s Nagar',
    'S A S NAGAR': 'S.a.s Nagar',
    'SAHIBZADA AJIT SINGH NAGAR': 'S.a.s Nagar',
    'MOHALI': 'S.a.s Nagar',
    # Sri Muktsar Sahib
    'SRI MUKTSAR SAHIB': 'Sri Muktsar Sahib',
    'MUKTSAR SAHIB': 'Sri Muktsar Sahib',
    'MUKTSAR': 'Sri Muktsar Sahib',
    'SRI MUKATSAR SAHIB': 'Sri Muktsar Sahib',
    'MUKATSAR': 'Sri Muktsar Sahib',
    # Shahid Bhagat Singh Nagar (Nawanshahr / SBS Nagar)
    'SBS NAGAR': 'Shahid Bhagat Singh Nagar',
    'S.B.S. NAGAR': 'Shahid Bhagat Singh Nagar',
    'S B S NAGAR': 'Shahid Bhagat Singh Nagar',
    'SHAHEED BHAGAT SINGH NAGAR': 'Shahid Bhagat Singh Nagar',
    'SHAHID BHAGAT SINGH NAGAR': 'Shahid Bhagat Singh Nagar',
    'NAWANSHAHR': 'Shahid Bhagat Singh Nagar',
    # Rupnagar (Ropar)
    'RUPNAGAR': 'Rupnagar',
    'RUPANAGAR': 'Rupnagar',
    'ROPAR': 'Rupnagar',
    # Ferozepur (alt spellings)
    'FEROZEPUR': 'Ferozepur',
    'FEROZPUR': 'Ferozepur',
    'FIROZPUR': 'Ferozepur',
    'FIROZEPUR': 'Ferozepur',
    # Tarn Taran (trailing space in source)
    'TARN TARAN': 'Tarn Taran',
    'TARNTARAN': 'Tarn Taran',
    # Fatehgarh Sahib
    'FATEHGARH SAHIB': 'Fatehgarh Sahib',
    # Identity entries (case-insensitive match through alias map)
    'AMRITSAR': 'Amritsar',
    'BARNALA': 'Barnala',
    'BATHINDA': 'Bathinda',
    'FARIDKOT': 'Faridkot',
    'FAZILKA': 'Fazilka',
    'GURDASPUR': 'Gurdaspur',
    'HOSHIARPUR': 'Hoshiarpur',
    'JALANDHAR': 'Jalandhar',
    'KAPURTHALA': 'Kapurthala',
    'LUDHIANA': 'Ludhiana',
    'MALERKOTLA': 'Malerkotla',
    'MANSA': 'Mansa',
    'MOGA': 'Moga',
    'PATHANKOT': 'Pathankot',
    'PATIALA': 'Patiala',
    'SANGRUR': 'Sangrur',
}


def normalize_district(name):
    if name is None:
        return None
    s = str(name).strip()
    if not s:
        return None
    s = s.rstrip('.').strip()
    m = re.match(r'^\d+[\.\)]?\s+(.+)$', s)
    if m:
        s = m.group(1).strip()
    s_up = s.upper()
    s_up_strip = re.sub(r'\s*\([^)]*\)\s*$', '', s_up).strip()
    for cand in (s_up, s_up_strip):
        if cand in DISTRICT_ALIASES_UPPER:
            return DISTRICT_ALIASES_UPPER[cand]
    return None


# ---------------------------------------------------------------------------
# Unit conversion: ACP source amounts are in Crores. FINER unit is Lakhs.
# Multiply monetary fields *100; leave percentage fields unchanged.
# ---------------------------------------------------------------------------

MONETARY_TOKENS = ('target_amt', 'achievement_amt')


def is_monetary_field(field_name):
    fl = field_name.lower()
    for tok in MONETARY_TOKENS:
        if tok in fl:
            return True
    return False


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('₹', '').replace('%', '')
    if not s or s in ('-', '#DIV/0!', 'N.A.', 'NA', '#REF!', '#VALUE!'):
        return None
    try:
        return float(s)
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Per-file parser
# ---------------------------------------------------------------------------

# Column layout (per file inspection, consistent across all 19 ACP files):
#   col B (1): District name
#   col C-E (2-4): Agri (Target, Achievement, % Achievement)
#   col F-H (5-7): MSME (Target, Achievement, % Achievement)
#   col I-K (8-10): Other PS (Target, Achievement, % Achievement)
#   col L-N (11-13): Total PS (Target, Achievement, % Achievement)
# Header rows: 1-8. Data rows: row 9 onwards until 'TOTAL'.
# Some files have a second table below in "Amount in thousand" Rupees -- ignored.

BUCKETS = [
    ('agri',    2),   # Agriculture & Allied
    ('msme',    5),   # MSME
    ('ops',     8),   # Other Priority Sector
    ('total_ps', 11), # Total Priority Sector
]

FIELDS_PER_BUCKET = ('target_amt', 'achievement_amt', 'achievement_pct')


def extract_acp_file(path):
    """Return list of dicts, one per district row in the first ACP table."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = {}
    seen_total = False
    for row in ws.iter_rows(values_only=True):
        if seen_total:
            break
        if row is None or all(c is None for c in row):
            continue
        # col B is index 1
        if len(row) < 14:
            continue
        cell_b = row[1]
        if cell_b is None:
            continue
        b_str = str(cell_b).strip().upper()
        # Stop when we hit the totals row, then skip everything after
        if b_str == 'TOTAL' or b_str.startswith('STATE TOTAL'):
            seen_total = True
            continue
        district = normalize_district(cell_b)
        if not district:
            continue
        rec = {'district': district}
        for bname, col in BUCKETS:
            t = to_float(row[col])
            a = to_float(row[col + 1])
            p = to_float(row[col + 2])
            f_target = f'priority_sector__{bname}_target_amt'
            f_ach = f'priority_sector__{bname}_achievement_amt'
            f_pct = f'priority_sector__{bname}_achievement_pct'
            if t is not None:
                rec[f_target] = t * 100  # Crore -> Lakh
            if a is not None:
                rec[f_ach] = a * 100  # Crore -> Lakh
            if p is not None:
                # Source stores percentages as 0.92 = 92% (decimal form).
                # Express as 0-100 scale to match FINER conventions.
                rec[f_pct] = p * 100 if abs(p) <= 5 else p
        # de-dup (same district appearing twice in a file - take first)
        if district not in records:
            records[district] = rec
    return list(records.values())


# ---------------------------------------------------------------------------
# Build aggregate outputs
# ---------------------------------------------------------------------------

def main():
    quarters = []  # for _fi_timeseries.json
    complete_quarters = {}  # for _complete.json

    samples = []  # for pre/post unit-conversion sanity

    for fname, (label, period_code, end_date) in ACP_FILES.items():
        path = os.path.join(DOWNLOADS_DIR, fname)
        if not os.path.exists(path):
            print(f"  SKIP missing {fname}")
            continue
        recs = extract_acp_file(path)
        if not recs:
            print(f"  WARN no rows in {fname}")
            continue
        # add 'period' to each record
        for r in recs:
            r['period'] = label
        quarters.append({'period': label, 'districts': recs})

        # complete.json shape
        fields_set = set()
        for r in recs:
            for k in r:
                if k not in ('district', 'period'):
                    fields_set.add(k)
        quarter_key = label.lower().replace(' ', '_')
        complete_quarters[quarter_key] = {
            'period': label,
            'tables': {
                'priority_sector': {
                    'fields': sorted(fields_set),
                    'districts': {r['district']: {k: r[k] for k in r if k not in ('district', 'period')} for r in recs},
                }
            }
        }

        n_d = len(recs)
        # sample for unit-conversion check (Ludhiana achievement_amt total_ps)
        for r in recs:
            if r['district'] == 'Ludhiana' and label in ('December 2025', 'December 2024'):
                samples.append((label, r.get('priority_sector__total_ps_achievement_amt')))
        print(f"  {fname}: {label}  {n_d} districts, {len(fields_set)} fields")

    # Sort quarters chronologically
    def keyfn(q):
        from datetime import datetime
        return datetime.strptime(q['period'], '%B %Y')
    quarters.sort(key=keyfn)

    timeseries = {'periods': quarters}
    complete = {'state': 'Punjab', 'quarters': complete_quarters}

    out_state_complete = os.path.join(BASE_DIR, 'punjab_complete.json')
    out_state_ts = os.path.join(BASE_DIR, 'punjab_fi_timeseries.json')
    with open(out_state_complete, 'w') as f:
        json.dump(complete, f, indent=2, ensure_ascii=False)
    with open(out_state_ts, 'w') as f:
        json.dump(timeseries, f, indent=2, ensure_ascii=False)

    # copy to public
    os.makedirs(PUBLIC_DIR, exist_ok=True)
    import shutil
    shutil.copy(out_state_complete, os.path.join(PUBLIC_DIR, 'punjab_complete.json'))
    shutil.copy(out_state_ts, os.path.join(PUBLIC_DIR, 'punjab_fi_timeseries.json'))

    # slim json — same as full ts since we only have priority_sector
    slim_path = os.path.join(PUBLIC_DIR, 'punjab_fi_slim.json')
    with open(slim_path, 'w') as f:
        json.dump(timeseries, f, indent=2, ensure_ascii=False)

    print(f"\nWrote {out_state_complete}")
    print(f"Wrote {out_state_ts}")
    print(f"Copied outputs to {PUBLIC_DIR}/")
    print(f"\nSample Ludhiana total_ps achievement values (after *100 Crore->Lakh):")
    for label, v in samples:
        print(f"  {label}: {v:,.2f} Lakhs")


if __name__ == '__main__':
    main()
