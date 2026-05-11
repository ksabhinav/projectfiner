#!/usr/bin/env python3
"""
Haryana SLBC Data Extractor v2
==============================

Extends extract_haryana.py to:
  1. Backfill older meetings 131st-162nd (Dec 2014 -> Sep 2022) downloaded
     from slbcharyana.pnb.bank.in archive page.
  2. Operate on already-extracted folders in raw/{N}/ (mixture of .zip+.rar
     archives; .xlsx + legacy .xls files).
  3. APPLY CRORE -> LAKH CONVERSION (*100) for all monetary fields
     (deposit, advance, amount, disbursement_amt, outstanding_amt, etc.).
     The Haryana source publishes all amounts in Crores; existing
     haryana_fi_timeseries.json incorrectly stored Crore values directly.
     v2 fixes this.
  4. Preserve all indicators the v1 extractor produces for meetings 163+:
     credit_deposit_ratio, pmjdy, kcc (saturation+AH+fishery), shg,
     sc_st_finance (HSFDC), branch_network, atm_network, aadhaar_authentication.
  5. For the older meetings, prioritise CD ratio (cleanest, most uniform
     format) plus PMJDY + SHG when available.

Output mirrors v1: haryana_complete.json + haryana_fi_timeseries.json
(periods -> districts), copied to public/slbc-data/haryana/.

Source: https://slbcharyana.pnb.bank.in/slbc-meeting-held/
Audit:  slbc-data/haryana/meetings_audit.txt
"""

import json
import os
import re
import shutil
import sys
import warnings
from collections import defaultdict
from glob import glob
from io import BytesIO

import openpyxl

try:
    import xlrd  # for legacy .xls files (older meetings)
    HAS_XLRD = True
except Exception:
    HAS_XLRD = False

warnings.filterwarnings('ignore')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, 'raw')
PUBLIC_DIR = os.path.abspath(os.path.join(BASE_DIR, '..', '..', 'public', 'slbc-data', 'haryana'))


# ----------------------------------------------------------------------
# Meeting -> Quarter mapping (extends v1 with 131-162)
# ----------------------------------------------------------------------

MEETING_QUARTERS = {
    # Backfill (131-162)
    131: ('December 2014', '2014-12', '31-12-2014'),
    132: ('March 2015',    '2015-03', '31-03-2015'),
    133: ('June 2015',     '2015-06', '30-06-2015'),
    134: ('September 2015','2015-09', '30-09-2015'),
    135: ('December 2015', '2015-12', '31-12-2015'),
    136: ('March 2016',    '2016-03', '31-03-2016'),
    137: ('June 2016',     '2016-06', '30-06-2016'),
    138: ('September 2016','2016-09', '30-09-2016'),
    139: ('December 2016', '2016-12', '31-12-2016'),
    140: ('March 2017',    '2017-03', '31-03-2017'),
    141: ('June 2017',     '2017-06', '30-06-2017'),
    142: ('September 2017','2017-09', '30-09-2017'),
    143: ('December 2017', '2017-12', '31-12-2017'),
    144: ('March 2018',    '2018-03', '31-03-2018'),
    145: ('June 2018',     '2018-06', '30-06-2018'),
    146: ('September 2018','2018-09', '30-09-2018'),
    147: ('December 2018', '2018-12', '31-12-2018'),
    148: ('March 2019',    '2019-03', '31-03-2019'),
    149: ('June 2019',     '2019-06', '30-06-2019'),
    150: ('September 2019','2019-09', '30-09-2019'),
    151: ('December 2019', '2019-12', '31-12-2019'),
    152: ('March 2020',    '2020-03', '31-03-2020'),
    153: ('June 2020',     '2020-06', '30-06-2020'),
    # 154 (Nov 2020) had no annexures — skipped
    155: ('December 2020', '2020-12', '31-12-2020'),
    156: ('March 2021',    '2021-03', '31-03-2021'),
    157: ('June 2021',     '2021-06', '30-06-2021'),
    158: ('September 2021','2021-09', '30-09-2021'),
    159: ('December 2021', '2021-12', '31-12-2021'),
    160: ('March 2022',    '2022-03', '31-03-2022'),
    161: ('June 2022',     '2022-06', '30-06-2022'),
    162: ('September 2022','2022-09', '30-09-2022'),
    # Currently in v1
    163: ('December 2022', '2022-12', '31-12-2022'),
    164: ('March 2023',    '2023-03', '31-03-2023'),
    165: ('June 2023',     '2023-06', '30-06-2023'),
    166: ('September 2023','2023-09', '30-09-2023'),
    167: ('December 2023', '2023-12', '31-12-2023'),
    168: ('March 2024',    '2024-03', '31-03-2024'),
    169: ('June 2024',     '2024-06', '30-06-2024'),
    170: ('September 2024','2024-09', '30-09-2024'),
    171: ('December 2024', '2024-12', '31-12-2024'),
    172: ('March 2025',    '2025-03', '31-03-2025'),
    173: ('June 2025',     '2025-06', '30-06-2025'),
    174: ('September 2025','2025-09', '30-09-2025'),
    175: ('December 2025', '2025-12', '31-12-2025'),
}


# ----------------------------------------------------------------------
# Districts + aliases
# ----------------------------------------------------------------------

# Canonical names match the FINER SQLite `districts.name` for Haryana.
# Charkhi Dadri came into existence 2016 (split from Bhiwani); meetings 131-137
# will lack it. Palwal split from Faridabad in 2008 (already a district).
# Mewat was renamed to Nuh in April 2016; older meetings use MEWAT.
HARYANA_DISTRICTS = [
    "Ambala", "Bhiwani", "Charki Dadri", "Faridabad", "Fatehabad",
    "Gurugram", "Hisar", "Jhajjar", "Jind", "Kaithal",
    "Karnal", "Kurukshetra", "Mahendragarh", "Nuh", "Palwal",
    "Panchkula", "Panipat", "Rewari", "Rohtak", "Sirsa",
    "Sonipat", "Yamunanagar",
]
# Title-cased for output, looked-up via UPPER alias map below.

DISTRICT_ALIASES_UPPER = {
    # Charkhi Dadri spellings
    'CHARKHI DADRI': 'Charki Dadri',
    'CHARKHI DAD.': 'Charki Dadri',
    'CHARKI DADRI': 'Charki Dadri',
    'DADRI': 'Charki Dadri',  # rarer
    # Mahendragarh
    'M.GARH': 'Mahendragarh',
    'M GARH': 'Mahendragarh',
    'MGARH': 'Mahendragarh',
    'MAHENDERGARH': 'Mahendragarh',
    'MAHENDRAGARH': 'Mahendragarh',
    'NARNAUL': 'Mahendragarh',  # pre-2016 reports
    # Sonipat / Sonepat
    'SONEPAT': 'Sonipat',
    'SONIPAT': 'Sonipat',
    # Yamunanagar
    'YAMUNA NAGAR': 'Yamunanagar',
    'YAMUNA-NAGAR': 'Yamunanagar',
    'YAMUNANAGAR': 'Yamunanagar',
    'Y.NAGAR': 'Yamunanagar',
    'Y NAGAR': 'Yamunanagar',
    # Nuh / Mewat
    'MEWAT': 'Nuh',
    'NUH': 'Nuh',
    'NUH (MEWAT)': 'Nuh',
    # Gurugram / Gurgaon
    'GURGAON': 'Gurugram',
    'GURGOAN': 'Gurugram',
    'GURUGRAM': 'Gurugram',
    # Misc identity (allow case-insensitive match for the rest)
    'AMBALA': 'Ambala', 'BHIWANI': 'Bhiwani', 'FARIDABAD': 'Faridabad',
    'FATEHABAD': 'Fatehabad', 'HISAR': 'Hisar', 'HISSAR': 'Hisar',
    'JHAJJAR': 'Jhajjar', 'JIND': 'Jind', 'KAITHAL': 'Kaithal',
    'KARNAL': 'Karnal', 'KURUKSHETRA': 'Kurukshetra',
    'PALWAL': 'Palwal', 'PANCHKULA': 'Panchkula',
    'PANIPAT': 'Panipat', 'REWARI': 'Rewari', 'ROHTAK': 'Rohtak',
    'SIRSA': 'Sirsa',
}

DISTRICT_LOOKUP_UPPER = {d.upper(): d for d in HARYANA_DISTRICTS}
DISTRICT_LOOKUP_UPPER.update({k: v for k, v in DISTRICT_ALIASES_UPPER.items()})


def normalize_district(name):
    """Resolve raw cell text to canonical district name (or None)."""
    if name is None:
        return None
    s = str(name).strip()
    if not s:
        return None
    s = s.rstrip('.')
    # Strip leading serial numbers like "1. AMBALA" or "1 AMBALA"
    m = re.match(r'^\d+[\.\)]?\s+(.+)$', s)
    if m:
        s = m.group(1)
    s_up = s.upper()
    # Drop trailing parenthesised parts: "AMBALA (URBAN)" -> "AMBALA"
    s_up_strip = re.sub(r'\s*\([^)]*\)\s*$', '', s_up).strip()
    for cand in (s_up, s_up_strip):
        if cand in DISTRICT_LOOKUP_UPPER:
            return DISTRICT_LOOKUP_UPPER[cand]
    return None


def is_district_name(val):
    if val is None:
        return False
    s = str(val).strip().upper()
    if s in ('', 'TOTAL', 'SUM:', 'GRAND TOTAL', 'HARYANA STATE', 'HARYANA',
             'NAME OF DISTRICT', 'NAME OF THE DISTRICT', 'DISTRICT',
             'DISTRICTS', 'SOURCE:LDMS', 'SOURCE: LDMS', 'STATE TOTAL'):
        return False
    return normalize_district(val) is not None


# ----------------------------------------------------------------------
# Monetary field detection -> Crore->Lakh conversion (*100)
# ----------------------------------------------------------------------

# Field-name substrings that indicate a monetary amount in INR Crores.
# The data is published in Crores; FINER stores in Lakhs (1 Cr = 100 Lakhs).
# Multiply *100 before inserting into JSON / SQLite.
MONETARY_TOKENS = (
    '_deposit',           # total_deposit
    '_advance',           # total_advance
    '_amt',               # *_amt (lots)
    'savings_linked_amt', 'credit_linked_amt',
    'amount',             # generic
    'disbursement',       # disbursement_amt (sometimes bare)
    'outstanding',        # outstanding (loan o/s)
    'sanctioned_amt',
    'limit_sanctioned',
)
# Fields that *contain* monetary tokens but are NOT amounts (e.g. percentage
# coverage of deposit accounts, deposit account counts). Exclude these.
MONETARY_EXCLUDE_TOKENS = (
    '_pct', 'ratio', 'percentage', '_no', '_a_c', '_number',
    '_count', 'no_of',
)


def is_monetary_field(field_key: str) -> bool:
    """Heuristic: is this field name a monetary amount (Crores in source)?
    Only the part AFTER '__' is checked, so 'credit_deposit_ratio' as a
    category doesn't accidentally veto a monetary field via 'ratio'.
    """
    fl = field_key.lower()
    # Look only at the field portion (after the '__' category separator).
    if '__' in fl:
        fl = fl.split('__', 1)[1]
    if any(ex in fl for ex in MONETARY_EXCLUDE_TOKENS):
        # Exception: '*_amt' beats '_no' (e.g. there is no real-world field
        # combining both); a field key with '_amt' overrides.
        if '_amt' in fl or 'amount' in fl:
            return True
        return False
    return any(tok in fl for tok in MONETARY_TOKENS)


def cr_to_lakh(val):
    """Multiply a Crore-denominated value by 100 to convert to Lakhs."""
    if val is None:
        return None
    try:
        return round(float(val) * 100.0, 4)
    except (ValueError, TypeError):
        return None


# ----------------------------------------------------------------------
# XLS / XLSX iteration helpers
# ----------------------------------------------------------------------

def _safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('%', '').replace('₹', '')
    if s in ('', '-', 'NA', 'N/A', 'nil', 'Nil', 'NIL', '#REF!', '#N/A', '#VALUE!'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _iter_xlsx_rows(path):
    """Yield (sheet_name, list_of_rows) for an .xlsx file."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        yield sn, list(ws.iter_rows(values_only=True))
    wb.close()


def _iter_xls_rows(path):
    """Yield (sheet_name, list_of_rows) for an .xls file (requires xlrd)."""
    if not HAS_XLRD:
        return
    try:
        wb = xlrd.open_workbook(path)
    except Exception as e:
        print(f"    [xlrd] open failed: {e}")
        return
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        rows = []
        for i in range(ws.nrows):
            row = []
            for j in range(ws.ncols):
                v = ws.cell_value(i, j)
                row.append(v if v != '' else None)
            rows.append(tuple(row))
        yield sn, rows


def _iter_rows(path):
    fl = path.lower()
    if fl.endswith('.xlsx'):
        yield from _iter_xlsx_rows(path)
    elif fl.endswith('.xls'):
        yield from _iter_xls_rows(path)


# ----------------------------------------------------------------------
# CD ratio extractor (works for ALL meetings 131-175)
# ----------------------------------------------------------------------

def _find_cd_files(meeting_dir):
    """Return list of (path, expected_quarter_hint) for CD-ratio-like xlsx/xls."""
    out = []
    for path in glob(os.path.join(meeting_dir, '**', '*.*'), recursive=True):
        bn = os.path.basename(path).lower()
        if bn.startswith('~') or bn.startswith('.'):
            continue
        if not (bn.endswith('.xlsx') or bn.endswith('.xls')):
            continue
        if 'basic data' in bn or 'basic-data' in bn or 'national' in bn:
            # The BASIC DATA & National Goals annexure also contains a CD-ratio
            # mini-table on one of its many sheets; we accept it as a fallback.
            out.append((path, 'basic'))
            continue
        if ('cd ratio' in bn or 'cd-ratio' in bn or 'cdratio' in bn
                or 'disttwise cd' in bn or 'districtwise cd' in bn):
            out.append((path, 'cd'))
    # Prefer dedicated CD ratio annexures over basic-data fallback
    out.sort(key=lambda x: 0 if x[1] == 'cd' else 1)
    return out


def extract_cd_ratio(meeting_dir, meeting_num):
    """Extract CD ratio + deposit + advance from a meeting's XLSX/XLS files.
    Returns dict { canonical_district: { 'total_deposit_cr': float,
                                         'total_advance_cr': float,
                                         'overall_cd_ratio': float } }

    Values are left in the natural source unit. The Haryana CD-ratio annexures
    almost always label values as "AMT. RS. IN CRORES" — and all but a few
    quarters genuinely use Crores. Meetings 156 (Mar 2021) and 160 (Mar 2022)
    are exceptions: their source files print values already converted to
    Lakhs despite the Crores label. We auto-detect this by checking the
    median deposit magnitude:
        - if median >= 100_000 Cr (highly unlikely for any HR district),
          treat the data as already in Lakhs and skip the *100 conversion.
    """
    files = _find_cd_files(meeting_dir)
    if not files:
        return None

    for path, kind in files:
        for sheet_name, rows in _iter_rows(path):
            districts = _parse_cd_sheet(rows)
            if districts and len(districts) >= 15:  # sanity threshold
                # Detect unit: median deposit > 100,000 implies values are
                # already in Lakhs (not Crores as the header claims).
                deps = [v['total_deposit_cr'] for v in districts.values()
                        if v.get('total_deposit_cr') is not None]
                already_lakhs = False
                if deps:
                    deps_sorted = sorted(deps)
                    median = deps_sorted[len(deps_sorted) // 2]
                    if median >= 100_000:
                        already_lakhs = True
                        for v in districts.values():
                            # Mark by flipping the sign convention: we wrap
                            # values in a tuple-like marker so cr_to_lakh isn't
                            # applied later. Simpler: divide by 100 here so the
                            # downstream *100 lands the same Lakhs value.
                            if v.get('total_deposit_cr') is not None:
                                v['total_deposit_cr'] = v['total_deposit_cr'] / 100
                            if v.get('total_advance_cr') is not None:
                                v['total_advance_cr'] = v['total_advance_cr'] / 100
                print(f"    [CD] {meeting_num}: {len(districts)} districts "
                      f"from {os.path.basename(path)}::{sheet_name}"
                      f"{'  (source already in Lakhs; pre-divided)' if already_lakhs else ''}")
                return districts
    return None


def _parse_cd_sheet(rows):
    """Find the header row with district/deposits/advances/cd-ratio columns
    and parse out the district-wise data."""
    header_idx = None
    cols = {}
    for i, row in enumerate(rows):
        if not row:
            continue
        row_str = ' '.join(str(v or '').lower() for v in row)
        # Must mention all four: name/district, deposit, advance, ratio
        has_district = 'name of' in row_str or 'district' in row_str
        has_dep = 'deposit' in row_str
        has_adv = 'advance' in row_str
        has_ratio = 'cd ratio' in row_str or 'c.d. ratio' in row_str or 'cdratio' in row_str
        if has_district and has_dep and has_adv and has_ratio:
            header_idx = i
            break

    if header_idx is None:
        return None

    header = rows[header_idx]
    # Identify columns: prefer the FIRST occurrence (INCL. COOP. BANKS comes
    # first in the layouts we see). If the row has multiple "DEPOSITS" cells
    # (some files put incl-coop and excl-coop side by side), we take the
    # leftmost set.
    col_dist = col_dep = col_adv = col_ratio = None
    for idx, val in enumerate(header):
        if val is None:
            continue
        vl = str(val).strip().lower()
        if col_dist is None and ('district' in vl or vl == 'name'):
            col_dist = idx
        elif col_dep is None and 'deposit' in vl:
            col_dep = idx
        elif col_adv is None and 'advance' in vl:
            col_adv = idx
        elif col_ratio is None and ('ratio' in vl):
            col_ratio = idx

    if None in (col_dist, col_dep, col_adv, col_ratio):
        return None

    districts = {}
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= col_dist:
            continue
        dname = normalize_district(row[col_dist])
        if not dname:
            continue
        dep = _safe_float(row[col_dep]) if col_dep < len(row) else None
        adv = _safe_float(row[col_adv]) if col_adv < len(row) else None
        ratio = _safe_float(row[col_ratio]) if col_ratio < len(row) else None
        # CD ratio is often stored as a decimal (0.5391) in the source — convert
        # to percentage. Modern meetings already store as percent (e.g. 51.72).
        if ratio is not None and ratio < 5:
            ratio = ratio * 100
        if ratio is not None:
            ratio = round(ratio, 2)
        if dep is None and adv is None and ratio is None:
            continue
        # If a district appears multiple times in the sheet (e.g. incl + excl
        # coop columns share row position), keep the first.
        if dname in districts:
            continue
        districts[dname] = {
            'total_deposit_cr': dep,
            'total_advance_cr': adv,
            'overall_cd_ratio': ratio,
        }
    return districts if len(districts) >= 15 else None


# ----------------------------------------------------------------------
# Main: walk every meeting in raw/ and build the timeseries
# ----------------------------------------------------------------------

def _meeting_dir_for(meeting_num):
    """Return the actual directory containing the XLSX files for this meeting.
    Most archives extract to raw/{N}/<single subdir>/*.xlsx. Some are flat
    (raw/{N}/*.xlsx)."""
    base = os.path.join(RAW_DIR, str(meeting_num))
    if not os.path.isdir(base):
        return None
    # If there's already xlsx/xls at top level, use that.
    direct = [f for f in os.listdir(base) if f.lower().endswith(('.xlsx', '.xls'))]
    if direct:
        return base
    # Single nested folder?
    subs = [os.path.join(base, e) for e in os.listdir(base)
            if os.path.isdir(os.path.join(base, e)) and not e.startswith('.')]
    if len(subs) == 1:
        return subs[0]
    # Multiple subdirs — return base; downstream uses recursive glob.
    return base


def build_v1_data_for_meeting(meeting_num):
    """Run the v1 extractor's PMJDY/Aadhaar/KCC/SHG/HSFDC functions when this
    is a meeting v1 already supports (163-175). The v1 extractor opens ZIPs;
    we adapt it to read from raw/{N}/ directories instead."""
    # Lazy import to avoid loading v1 unless needed.
    from extract_haryana import (
        extract_pmjdy as _e_pmjdy,
        extract_aadhaar as _e_aadhaar,
        extract_kcc_saturation as _e_kcc,
        extract_kcc_ah as _e_kcc_ah,
        extract_hsfdc as _e_hsfdc,
        extract_shg as _e_shg,
    )
    # v1 expects a ZipFile-like object with .namelist() and .read(name).
    # Wrap the raw/{N}/ directory.
    mdir = _meeting_dir_for(meeting_num)
    if not mdir:
        return None

    class DirZip:
        def __init__(self, root):
            self.root = root
            self._files = []
            for p in glob(os.path.join(root, '**', '*.*'), recursive=True):
                rel = os.path.relpath(p, root).replace(os.sep, '/')
                bn = os.path.basename(rel)
                if bn.startswith('~') or bn.startswith('.'):
                    continue
                self._files.append(rel)

        def namelist(self):
            return self._files

        def read(self, name):
            with open(os.path.join(self.root, name), 'rb') as f:
                return f.read()

    z = DirZip(mdir)
    meeting_tag = str(meeting_num)  # v1 uses meeting tag like '163rd'; the
                                    # functions only use it for logging.

    out = {}
    try:
        pmjdy = _e_pmjdy(z, meeting_tag)
        if pmjdy:
            out['pmjdy'] = pmjdy
    except Exception as e:
        print(f"    [v1 pmjdy] {meeting_num}: {e}")
    try:
        aadhaar = _e_aadhaar(z, meeting_tag)
        if aadhaar:
            out['aadhaar_authentication'] = aadhaar
    except Exception as e:
        print(f"    [v1 aadhaar] {meeting_num}: {e}")
    try:
        kcc = _e_kcc(z, meeting_tag)
        if kcc:
            out['kcc'] = kcc
    except Exception as e:
        print(f"    [v1 kcc] {meeting_num}: {e}")
    try:
        kcc_ah = _e_kcc_ah(z, meeting_tag)
        if kcc_ah:
            out['kcc_ah'] = kcc_ah
    except Exception as e:
        print(f"    [v1 kcc_ah] {meeting_num}: {e}")
    try:
        hsfdc = _e_hsfdc(z, meeting_tag)
        if hsfdc:
            out['sc_st_finance'] = hsfdc
    except Exception as e:
        print(f"    [v1 hsfdc] {meeting_num}: {e}")
    try:
        shg = _e_shg(z, meeting_tag)
        if shg:
            out['shg'] = shg
    except Exception as e:
        print(f"    [v1 shg] {meeting_num}: {e}")
    return out


def _district_records_for_meeting(meeting_num, period_label):
    """Build a list of per-district record dicts for a meeting.
    Each record has keys: district, period, plus snake_case
    'category__field' entries.
    """
    mdir = _meeting_dir_for(meeting_num)
    if not mdir:
        return []

    # 1. CD ratio (universal).
    cd = extract_cd_ratio(mdir, meeting_num)

    # 2. Other indicators via v1 logic — only meaningful for meetings >= 163
    #    (v1's file-naming heuristics expect the 2022+ layout).
    v1 = build_v1_data_for_meeting(meeting_num) if meeting_num >= 163 else {}

    # Map v1's per-district dicts (key = UPPERCASE district name) to canonical
    # title-case names, AND remap field names to the FINER canonical
    # category__field convention.
    def remap(category, distmap, field_map):
        """Yield (canon_district, {field_key: value})."""
        if not distmap:
            return
        for raw_d, fields in distmap.items():
            cd_name = normalize_district(raw_d) if isinstance(raw_d, str) else None
            if not cd_name:
                continue
            row = {}
            for src_field, val in (fields or {}).items():
                tgt = field_map.get(src_field)
                if not tgt:
                    continue
                row[f"{category}__{tgt}"] = val
            yield cd_name, row

    # v1's PMJDY fields -> FINER field names (already roughly matching).
    PMJDY_MAP = {
        'target': 'target_no',
        'total': 'total_pmjdy_no',
        'rural': 'rural_no',
        'urban': 'urban_no',
        'male': 'male_no',
        'female': 'female_no',
        'rupay_card': 'no_of_rupay_card_issued',
        'aadhaar_seeded': 'no_of_aadhaar_seeded',
    }
    AADHAAR_MAP = {
        'operative_casa': 'no_of_operative_casa',
        'aadhaar_seeded': 'no_of_aadhaar_seeded_casa',
        'authenticated': 'no_of_authenticated_casa',
    }
    # KCC Saturation: 'total_kcc_accounts' is the canonical FINER 'total_no_of_kcc'
    # field. The other two keep their existing snake_case names.
    KCC_MAP = {
        'total_kcc_accounts': 'total_no_of_kcc',
        'insured': 'kcc_insured',
        'pending': 'kcc_pending',
    }
    KCC_AH_MAP = {
        'applications_received': 'ah_applications_received',
        'applications_accepted': 'ah_applications_accepted',
        'applications_sanctioned': 'ah_applications_sanctioned',
    }
    HSFDC_MAP = {
        'applications': 'sc_applications_no',
        'sponsored': 'sc_sponsored_no',
        'sanctioned': 'sc_disbursement_no',  # treat sanctioned as disbursement_no
    }
    SHG_MAP = {
        'savings_linked_no': 'savings_linked_no',
        'savings_linked_amt': 'savings_linked_amt',
        'credit_linked_no': 'credit_linked_no',
        'credit_linked_amt': 'credit_linked_amt',
    }

    by_district = defaultdict(dict)

    if cd:
        for d, vals in cd.items():
            row = {}
            if vals.get('total_deposit_cr') is not None:
                row['credit_deposit_ratio__total_deposit'] = vals['total_deposit_cr']
            if vals.get('total_advance_cr') is not None:
                row['credit_deposit_ratio__total_advance'] = vals['total_advance_cr']
            if vals.get('overall_cd_ratio') is not None:
                row['credit_deposit_ratio__overall_cd_ratio'] = vals['overall_cd_ratio']
            by_district[d].update(row)

    for cat, src, fmap in [
        ('pmjdy', v1.get('pmjdy'), PMJDY_MAP),
        ('aadhaar_authentication', v1.get('aadhaar_authentication'), AADHAAR_MAP),
        ('kcc', v1.get('kcc'), KCC_MAP),
        ('kcc', v1.get('kcc_ah'), KCC_AH_MAP),
        ('sc_st_finance', v1.get('sc_st_finance'), HSFDC_MAP),
        ('shg', v1.get('shg'), SHG_MAP),
    ]:
        for d, row in remap(cat, src, fmap):
            by_district[d].update(row)

    # ----- Apply Crore -> Lakh conversion on monetary fields -----
    records = []
    for d in HARYANA_DISTRICTS:
        if d not in by_district:
            continue
        rec = {'district': d, 'period': period_label}
        for k, v in by_district[d].items():
            if v is None:
                continue
            if is_monetary_field(k):
                rec[k] = cr_to_lakh(v)
            else:
                rec[k] = v
        records.append(rec)
    return records


def main():
    if not os.path.isdir(RAW_DIR):
        print(f"ERROR: {RAW_DIR} not found. Run download_haryana_v2.sh first.")
        sys.exit(1)

    print('=' * 70)
    print('HARYANA SLBC v2 EXTRACTOR (with Crore -> Lakh conversion)')
    print('=' * 70)

    periods_out = []  # list of {period, districts: [...]}
    meeting_summary = []  # for the print summary

    for mnum in sorted(MEETING_QUARTERS.keys()):
        period_label, period_code, as_on = MEETING_QUARTERS[mnum]
        mdir = _meeting_dir_for(mnum)
        if not mdir:
            print(f"\n[SKIP] {mnum}: no extracted folder at raw/{mnum}/")
            continue
        print(f"\n--- Meeting {mnum} ({period_label}) ---")
        records = _district_records_for_meeting(mnum, period_label)
        if not records:
            print(f"    no district-wise records extracted")
            meeting_summary.append((mnum, period_label, 0, 0))
            continue

        # Count distinct categories
        cats = set()
        for r in records:
            for k in r:
                if '__' in k:
                    cats.add(k.split('__', 1)[0])
        print(f"    -> {len(records)} districts, {len(cats)} categories: "
              f"{', '.join(sorted(cats))}")
        meeting_summary.append((mnum, period_label, len(records), len(cats)))

        periods_out.append({'period': period_label, 'districts': records})

    # Sort periods chronologically using their code.
    code_for = {label: code for label, code, _ in MEETING_QUARTERS.values()}
    periods_out.sort(key=lambda p: code_for.get(p['period'], p['period']))

    # ---------------- Output files ----------------
    out_ts = {
        'source': 'SLBC Haryana (v2: 131st-175th, Crore->Lakh applied)',
        'state': 'haryana',
        'periods': periods_out,
    }

    os.makedirs(PUBLIC_DIR, exist_ok=True)
    ts_path_local = os.path.join(BASE_DIR, 'haryana_fi_timeseries.json')
    with open(ts_path_local, 'w') as f:
        json.dump(out_ts, f, indent=2)
    shutil.copy2(ts_path_local, os.path.join(PUBLIC_DIR, 'haryana_fi_timeseries.json'))
    print(f"\nWrote {ts_path_local}")
    print(f"Copied to {PUBLIC_DIR}/haryana_fi_timeseries.json")

    # Build a 'complete' JSON: per-quarter -> per-category -> per-district.
    complete = {'state': 'Haryana', 'quarters': {}}
    for period in periods_out:
        qkey = period['period'].lower().replace(' ', '_')
        by_cat = defaultdict(lambda: {'fields': set(), 'districts': {}})
        for rec in period['districts']:
            d = rec['district']
            for k, v in rec.items():
                if '__' not in k:
                    continue
                cat, fld = k.split('__', 1)
                by_cat[cat]['fields'].add(fld)
                by_cat[cat]['districts'].setdefault(d, {})[fld] = v
        complete['quarters'][qkey] = {
            'period': period['period'],
            'tables': {
                cat: {
                    'fields': sorted(info['fields']),
                    'districts': info['districts'],
                }
                for cat, info in by_cat.items()
            },
        }
    complete_path = os.path.join(BASE_DIR, 'haryana_complete.json')
    with open(complete_path, 'w') as f:
        json.dump(complete, f, indent=2)
    shutil.copy2(complete_path, os.path.join(PUBLIC_DIR, 'haryana_complete.json'))
    print(f"Wrote {complete_path} -> public/")

    # ---------------- Summary ----------------
    print('\n' + '=' * 70)
    print('EXTRACTION SUMMARY')
    print('=' * 70)
    print(f"{'Meeting':>7}  {'Period':<18}  {'Districts':>9}  {'Categories':>10}")
    print('-' * 50)
    for mnum, period, n_d, n_c in meeting_summary:
        print(f"{mnum:>7}  {period:<18}  {n_d:>9}  {n_c:>10}")

    total_records = sum(len(p['districts']) for p in periods_out)
    print(f"\nTotal: {len(periods_out)} quarters, {total_records} district-records")


if __name__ == '__main__':
    main()
