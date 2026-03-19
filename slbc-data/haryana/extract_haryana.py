#!/usr/bin/env python3
"""
Comprehensive extractor for Haryana SLBC data.
Extracts district-wise indicators from 163rd-175th meeting ZIP files.
"""

import zipfile
import openpyxl
import json
import os
import re
import shutil
from io import BytesIO
from collections import defaultdict

BASE_DIR = "/Users/abhinav/Downloads/projectfiner/slbc-data/haryana"
EXTRACTED_DIR = os.path.join(BASE_DIR, "extracted")
OUTPUT_DIR = os.path.join(BASE_DIR)
PUBLIC_DIR = "/Users/abhinav/Downloads/projectfiner/public/slbc-data/haryana"

# Meeting to quarter mapping
MEETING_QUARTERS = {
    "163rd": {"quarter": "Dec 2022", "date": "2022-12-31"},
    "164th": {"quarter": "Mar 2023", "date": "2023-03-31"},
    "165th": {"quarter": "Jun 2023", "date": "2023-06-30"},
    "166th": {"quarter": "Sep 2023", "date": "2023-09-30"},
    "167th": {"quarter": "Dec 2023", "date": "2023-12-31"},
    "168th": {"quarter": "Mar 2024", "date": "2024-03-31"},
    "169th": {"quarter": "Jun 2024", "date": "2024-06-30"},
    "170th": {"quarter": "Sep 2024", "date": "2024-09-30"},
    "171st": {"quarter": "Dec 2024", "date": "2024-12-31"},
    "172nd": {"quarter": "Mar 2025", "date": "2025-03-31"},
    "173rd": {"quarter": "Jun 2025", "date": "2025-06-30"},
    "174th": {"quarter": "Sep 2025", "date": "2025-09-30"},
    "175th": {"quarter": "Dec 2025", "date": "2025-12-31"},
}

# District name aliases: source -> canonical GeoJSON name
DISTRICT_ALIASES = {
    "CHARKHI DADRI": "CHARKI DADRI",
    "CHARKHI DAD.": "CHARKI DADRI",
    "CHARKI DADRI": "CHARKI DADRI",  # already correct
    "M.GARH": "MAHENDRAGARH",
    "MAHENDERGARH": "MAHENDRAGARH",
    "MAHENDRAGARH": "MAHENDRAGARH",
    "SONEPAT": "SONIPAT",
    "SONIPAT": "SONIPAT",
    "YAMUNA NAGAR": "YAMUNANAGAR",
    "Y.NAGAR": "YAMUNANAGAR",
    "YAMUNANAGAR": "YAMUNANAGAR",
    "MEWAT": "NUH",
    "NARNAUL": "MAHENDRAGARH",
    "GURGAON": "GURUGRAM",
}

HARYANA_DISTRICTS = [
    "AMBALA", "BHIWANI", "CHARKI DADRI", "FARIDABAD", "FATEHABAD",
    "GURUGRAM", "HISAR", "JHAJJAR", "JIND", "KAITHAL",
    "KARNAL", "KURUKSHETRA", "MAHENDRAGARH", "NUH", "PALWAL",
    "PANCHKULA", "PANIPAT", "REWARI", "ROHTAK", "SIRSA",
    "SONIPAT", "YAMUNANAGAR"
]


def normalize_district(name):
    """Normalize district name to canonical GeoJSON form."""
    if name is None:
        return None
    name = str(name).strip().upper()
    # Remove trailing dots
    name = name.rstrip(".")
    # Check alias map
    if name in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[name]
    # Check if it's a known district
    if name in HARYANA_DISTRICTS:
        return name
    # Try partial matching
    for alias, canonical in DISTRICT_ALIASES.items():
        if alias in name or name in alias:
            return canonical
    return name


def safe_float(val):
    """Convert value to float, return None if not possible."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip().replace(",", "")
        if val in ("", "-", "#REF!", "#N/A", "#VALUE!", "NA", "N/A"):
            return None
        try:
            return float(val)
        except ValueError:
            return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def find_file_in_zip(z, patterns, exclude_patterns=None):
    """Find a file in ZIP matching any of the patterns."""
    exclude_patterns = exclude_patterns or []
    for name in z.namelist():
        fn = name.split("/")[-1]
        if fn.startswith("~") or "__MACOSX" in name or fn.startswith("."):
            continue
        if not (fn.endswith(".xlsx") or fn.endswith(".xls")):
            continue
        fn_lower = fn.lower()
        name_lower = name.lower()
        # Check excludes
        excluded = False
        for ep in exclude_patterns:
            if ep.lower() in fn_lower:
                excluded = True
                break
        if excluded:
            continue
        # Check includes
        for pat in patterns:
            if pat.lower() in name_lower:
                return name
    return None


def find_sheet(wb, candidates):
    """Find a sheet by trying candidate names (case-insensitive)."""
    for cand in candidates:
        for sn in wb.sheetnames:
            if sn.strip().lower() == cand.lower():
                return wb[sn]
    # Fuzzy match
    for cand in candidates:
        for sn in wb.sheetnames:
            if cand.lower() in sn.strip().lower():
                return wb[sn]
    return None


def is_district_name(val):
    """Check if a value looks like a district name."""
    if val is None:
        return False
    val = str(val).strip().upper()
    if val in ("", "TOTAL", "SUM:", "GRAND TOTAL", "HARYANA STATE", "HARYANA"):
        return False
    normalized = normalize_district(val)
    return normalized in HARYANA_DISTRICTS


def open_workbook_from_zip(z, filename):
    """Open an Excel workbook from inside a ZIP."""
    data = z.read(filename)
    try:
        return openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception:
        return None


# ============================================================
# EXTRACTOR FUNCTIONS
# ============================================================

def _parse_pmjdy_rows(rows, meeting):
    """Parse PMJDY district-wise data from a list of row tuples."""
    districts = {}

    # Find header row
    header_row_idx = None
    for i, row in enumerate(rows):
        row_str = " ".join(str(v or "").lower() for v in row)
        if "district" in row_str and ("total" in row_str or "rural" in row_str or "male" in row_str):
            header_row_idx = i
            break

    if header_row_idx is None:
        return None

    header = rows[header_row_idx]

    # Find column indices by header names
    def find_col(keywords, header):
        for idx, val in enumerate(header):
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if any(kw in val_lower for kw in keywords):
                return idx
        return None

    col_district = find_col(["district"], header)
    col_target = find_col(["target"], header)
    # "Total" might appear in "Target" column too; find the actual total accounts column
    col_total = None
    for idx, val in enumerate(header):
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        if val_lower in ("total", "achievement") or ("total" in val_lower and "target" not in val_lower and "deposit" not in val_lower and "%" not in val_lower):
            col_total = idx
            break
    col_rural = find_col(["rural"], header)
    col_urban = find_col(["urban"], header)
    col_male = find_col(["male"], header)
    col_female = find_col(["female"], header)
    col_rupay = find_col(["rupay card issued"], header)
    if col_rupay is None:
        col_rupay = find_col(["rupay card"], header)
    col_aadhaar_seeded = None
    for idx, val in enumerate(header):
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        if ("aadhar" in val_lower or "aadhaar" in val_lower) and "seeded" in val_lower and "%" not in val_lower and "age" not in val_lower:
            col_aadhaar_seeded = idx
            break
    # Fallback: if header just says "Aadhar Seeded" without %
    if col_aadhaar_seeded is None:
        for idx, val in enumerate(header):
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if ("aadhar" in val_lower or "aadhaar" in val_lower) and "%" not in val_lower and "age" not in val_lower:
                col_aadhaar_seeded = idx
                break

    if col_district is None:
        return None

    # Read data rows
    for row in rows[header_row_idx + 1:]:
        district_raw = row[col_district] if col_district < len(row) else None
        if not is_district_name(district_raw):
            continue
        district = normalize_district(district_raw)
        entry = {}
        if col_target is not None and col_target < len(row):
            entry["target"] = safe_float(row[col_target])
        if col_total is not None and col_total < len(row):
            entry["total"] = safe_float(row[col_total])
        if col_rural is not None and col_rural < len(row):
            entry["rural"] = safe_float(row[col_rural])
        if col_urban is not None and col_urban < len(row):
            entry["urban"] = safe_float(row[col_urban])
        if col_male is not None and col_male < len(row):
            entry["male"] = safe_float(row[col_male])
        if col_female is not None and col_female < len(row):
            entry["female"] = safe_float(row[col_female])
        if col_rupay is not None and col_rupay < len(row):
            entry["rupay_card"] = safe_float(row[col_rupay])
        if col_aadhaar_seeded is not None and col_aadhaar_seeded < len(row):
            entry["aadhaar_seeded"] = safe_float(row[col_aadhaar_seeded])
        districts[district] = entry

    return districts if districts else None


def extract_pmjdy(z, meeting):
    """Extract PMJDY district-wise data.
    Tries multiple sheet names: DISTRICT WISE, 1.2, Sheet1.
    Also supports .xls files via xlrd.
    """
    filename = find_file_in_zip(z, ["pmjdy"], exclude_patterns=["inoperative"])
    if not filename:
        print(f"  [PMJDY] No file found for {meeting}")
        return None

    fn_lower = filename.lower()

    # Handle .xls files via xlrd
    if fn_lower.endswith(".xls") and not fn_lower.endswith(".xlsx"):
        try:
            import xlrd
            data = z.read(filename)
            xls_wb = xlrd.open_workbook(file_contents=data)
            # Try DISTRICT WISE sheet first, then 1.2
            target_sheet = None
            for cand in ["DISTRICT WISE", "District Wise", "1.2"]:
                if cand in xls_wb.sheet_names():
                    target_sheet = cand
                    break
            if target_sheet is None:
                for sn in xls_wb.sheet_names():
                    if "district" in sn.lower():
                        target_sheet = sn
                        break
            if target_sheet is None:
                print(f"  [PMJDY] No district-wise sheet in .xls {filename} (sheets: {xls_wb.sheet_names()})")
                return None
            ws = xls_wb.sheet_by_name(target_sheet)
            rows = []
            for i in range(ws.nrows):
                rows.append(tuple(ws.cell_value(i, j) for j in range(ws.ncols)))
            print(f"  [PMJDY] Found (.xls): {filename} -> sheet: {target_sheet}")
            districts = _parse_pmjdy_rows(rows, meeting)
            if districts:
                print(f"  [PMJDY] Extracted {len(districts)} districts")
            else:
                print(f"  [PMJDY] No district data parsed from .xls")
            return districts
        except Exception as e:
            print(f"  [PMJDY] Cannot open .xls {filename}: {e}")
            return None

    wb = open_workbook_from_zip(z, filename)
    if wb is None:
        print(f"  [PMJDY] Cannot open {filename}")
        return None

    # Try multiple sheet names: DISTRICT WISE, 1.2, Sheet1
    ws = find_sheet(wb, ["DISTRICT WISE", "District Wise", "1.2", "Sheet1"])
    if ws is None:
        print(f"  [PMJDY] No district-wise sheet in {filename} (sheets: {wb.sheetnames})")
        return None

    print(f"  [PMJDY] Found: {filename} -> sheet: {ws.title}")

    rows = list(ws.iter_rows(values_only=True))
    districts = _parse_pmjdy_rows(rows, meeting)

    if districts:
        print(f"  [PMJDY] Extracted {len(districts)} districts")
    else:
        print(f"  [PMJDY] No district data parsed")
    return districts


def extract_aadhaar(z, meeting):
    """Extract Aadhaar district-wise data."""
    filename = find_file_in_zip(z, ["aadhaar", "aadhar"])
    if not filename:
        return None

    wb = open_workbook_from_zip(z, filename)
    if wb is None:
        print(f"  [AADHAAR] Cannot open {filename}")
        return None

    ws = find_sheet(wb, ["DISTRICT WISE", "District Wise"])
    if ws is None:
        print(f"  [AADHAAR] No district-wise sheet in {filename} (sheets: {wb.sheetnames})")
        return None

    print(f"  [AADHAAR] Found: {filename} -> sheet: {ws.title}")

    districts = {}
    rows = list(ws.iter_rows(values_only=True))

    # Find header row and district column
    header_row_idx = None
    for i, row in enumerate(rows):
        row_str = " ".join(str(v or "").lower() for v in row)
        if "district" in row_str and ("casa" in row_str or "aadhar" in row_str or "aadhaar" in row_str or "authenticated" in row_str):
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  [AADHAAR] Cannot find header row")
        return None

    header = rows[header_row_idx]

    def find_col(keywords, header):
        for idx, val in enumerate(header):
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if all(kw in val_lower for kw in keywords):
                return idx
        return None

    col_district = find_col(["district"], header)
    col_casa = find_col(["operative"], header) or find_col(["casa"], header)
    col_seeded = find_col(["seeded"], header)
    col_auth = find_col(["authenticated"], header)

    if col_district is None:
        # Try to find district column by checking which column has district names
        for ci in range(len(header)):
            for row in rows[header_row_idx + 1: header_row_idx + 5]:
                if ci < len(row) and is_district_name(row[ci]):
                    col_district = ci
                    break
            if col_district is not None:
                break

    if col_district is None:
        print(f"  [AADHAAR] Cannot find district column")
        return None

    for row in rows[header_row_idx + 1:]:
        district_raw = row[col_district] if col_district < len(row) else None
        if not is_district_name(district_raw):
            continue
        district = normalize_district(district_raw)
        entry = {}
        if col_casa is not None and col_casa < len(row):
            entry["operative_casa"] = safe_float(row[col_casa])
        if col_seeded is not None and col_seeded < len(row):
            entry["aadhaar_seeded"] = safe_float(row[col_seeded])
        if col_auth is not None and col_auth < len(row):
            entry["authenticated"] = safe_float(row[col_auth])
        districts[district] = entry

    if districts:
        print(f"  [AADHAAR] Extracted {len(districts)} districts")
    return districts if districts else None


def extract_kcc_saturation(z, meeting):
    """Extract KCC Saturation district-wise data."""
    filename = find_file_in_zip(z, ["kcc_saturation", "kcc_saturat", "kcc saturation", "kcc_saturatiom"])
    if not filename:
        return None

    wb = open_workbook_from_zip(z, filename)
    if wb is None:
        print(f"  [KCC] Cannot open {filename}")
        return None

    # Try various sheet names for district-wise data
    ws = find_sheet(wb, ["District Wise", "DISTRICT WISE", "District wise",
                          "DIst wise", "13.2", "Sheet2"])
    if ws is None:
        # Scan all sheets for one that has a header with "district"/"name" AND data rows
        # that are purely district names (not bank names containing district names)
        for sn in wb.sheetnames:
            test_ws = wb[sn]
            has_district_header = False
            district_count = 0
            for row in test_ws.iter_rows(min_row=1, max_row=35, values_only=True):
                row_str = " ".join(str(v or "").lower() for v in row)
                if "district" in row_str or ("name" in row_str and "sr" in row_str):
                    has_district_header = True
                    continue
                if has_district_header:
                    # Check first non-None value in row
                    for val in row:
                        if val is not None:
                            val_str = str(val).strip()
                            # Must be a clean district name (not containing "bank", "coop", etc.)
                            if is_district_name(val_str) and "bank" not in val_str.lower() and "coop" not in val_str.lower():
                                district_count += 1
                            break
            if has_district_header and district_count >= 10:
                ws = test_ws
                break

    if ws is None:
        print(f"  [KCC] No district-wise sheet in {filename} (sheets: {wb.sheetnames})")
        return None

    print(f"  [KCC] Found: {filename} -> sheet: {ws.title}")

    districts = {}
    rows = list(ws.iter_rows(values_only=True))

    # Find header row - must have individual cells matching column-header patterns
    # (not a long title like "District-wise Saturation of KCC...")
    header_row_idx = None
    for i, row in enumerate(rows):
        # Count how many cells look like column headers
        has_name_col = False
        has_data_col = False
        col_count = 0
        for val in row:
            if val is None:
                continue
            vl = str(val).strip().lower()
            # Short cell that says "name" or "district" (a column header, not a title)
            if len(vl) < 30 and ("name" in vl or "district" in vl):
                has_name_col = True
            if any(kw in vl for kw in ["kcc", "pending", "insured", "enrolled", "eligible", "saturated"]):
                has_data_col = True
            if vl:
                col_count += 1
        if has_name_col and has_data_col and col_count >= 3:
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  [KCC] Cannot find header row")
        return None

    header = rows[header_row_idx]

    def find_col(keywords, header):
        for idx, val in enumerate(header):
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if all(kw in val_lower for kw in keywords):
                return idx
        return None

    col_district = find_col(["district"], header) or find_col(["name"], header)
    col_total = find_col(["total kcc"], header) or find_col(["kcc account"], header) or find_col(["total"], header)
    col_insured = find_col(["insured"], header) or find_col(["enrolled"], header) or find_col(["policy"], header)
    col_pending = find_col(["pending"], header)

    # Find the pending count column (not percentage)
    # Look for "Pending Accounts" vs "Pending Percent"
    col_pending_count = None
    col_pending_pct = None
    for idx, val in enumerate(header):
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        if "pending" in val_lower:
            if "percent" in val_lower or "%" in val_lower:
                col_pending_pct = idx
            elif "account" in val_lower or col_pending_count is None:
                col_pending_count = idx

    if col_district is None:
        # District might be second column (after Sr.No.)
        for ci in range(len(header)):
            test_rows = rows[header_row_idx + 1: header_row_idx + 4]
            for row in test_rows:
                if ci < len(row) and is_district_name(row[ci]):
                    col_district = ci
                    break
            if col_district is not None:
                break

    if col_district is None:
        print(f"  [KCC] Cannot find district column")
        return None

    for row in rows[header_row_idx + 1:]:
        district_raw = row[col_district] if col_district < len(row) else None
        if not is_district_name(district_raw):
            continue
        district = normalize_district(district_raw)
        entry = {}
        if col_total is not None and col_total < len(row):
            entry["total_kcc_accounts"] = safe_float(row[col_total])
        if col_insured is not None and col_insured < len(row):
            entry["insured"] = safe_float(row[col_insured])
        if col_pending_count is not None and col_pending_count < len(row):
            entry["pending"] = safe_float(row[col_pending_count])
        districts[district] = entry

    if districts:
        print(f"  [KCC] Extracted {len(districts)} districts")
    return districts if districts else None


def extract_kcc_ah(z, meeting):
    """Extract KCC AH (Animal Husbandry) district-wise data.
    Data may be in:
    - Sheet1 with district-level pivot (Row Labels = district names)
    - Sheet1 with bank-level pivot (Row Labels = bank names) -- skip this
    - '2.1-2.2' sheet with pre-aggregated district data
    - Raw 'Weekly KCC camps' data with bank-level rows that need aggregation
    """
    filename = find_file_in_zip(z, ["kcc ah", "kcc_ah", "kcc ahd", "ahdf"],
                                 exclude_patterns=["kcc crop", "kcc fish", "kcc fishir", "kcc fisheries"])
    if not filename:
        return None

    wb = open_workbook_from_zip(z, filename)
    if wb is None:
        print(f"  [KCC_AH] Cannot open {filename}")
        return None

    print(f"  [KCC_AH] Found: {filename} (sheets: {wb.sheetnames})")

    districts = {}

    # Strategy 1: Try '2.1-2.2' sheet (pre-aggregated district data)
    ws_district = find_sheet(wb, ["2.1-2.2"])
    if ws_district:
        rows = list(ws_district.iter_rows(values_only=True))
        header_row_idx = None
        for i, row in enumerate(rows):
            row_str = " ".join(str(v or "").lower() for v in row)
            if "district" in row_str and ("cummulative" in row_str or "cumulative" in row_str or "received" in row_str):
                header_row_idx = i
                break
        if header_row_idx is not None:
            header = rows[header_row_idx]
            col_district = None
            col_received = None
            col_accepted = None
            col_sanctioned = None
            for idx, val in enumerate(header):
                if val is None:
                    continue
                vl = str(val).strip().lower()
                if "district" in vl:
                    col_district = idx
                if ("cummulative" in vl or "cumulative" in vl) and "received" in vl:
                    col_received = idx
                if ("cummulative" in vl or "cumulative" in vl) and "accepted" in vl:
                    col_accepted = idx
                if ("cummulative" in vl or "cumulative" in vl) and "sanctioned" in vl:
                    col_sanctioned = idx

            if col_district is not None:
                for row in rows[header_row_idx + 1:]:
                    dr = row[col_district] if col_district < len(row) else None
                    if not is_district_name(dr):
                        continue
                    d = normalize_district(dr)
                    entry = {}
                    if col_received is not None and col_received < len(row):
                        entry["applications_received"] = safe_float(row[col_received])
                    if col_accepted is not None and col_accepted < len(row):
                        entry["applications_accepted"] = safe_float(row[col_accepted])
                    if col_sanctioned is not None and col_sanctioned < len(row):
                        entry["applications_sanctioned"] = safe_float(row[col_sanctioned])
                    districts[d] = entry

    if districts:
        print(f"  [KCC_AH] Extracted {len(districts)} districts from 2.1-2.2 sheet")
        return districts

    # Strategy 2: Try Sheet1 -- check if it's district-level pivot (not bank-level)
    ws = find_sheet(wb, ["Sheet1"])
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = None
    is_district_pivot = False

    for i, row in enumerate(rows):
        row_str = " ".join(str(v or "").lower() for v in row)
        if "row labels" in row_str or ("cummulative" in row_str and "received" in row_str):
            header_row_idx = i
            # Check if next rows are districts or banks
            check_count = 0
            for check_row in rows[i + 1: i + 6]:
                if check_row and check_row[0]:
                    if is_district_name(check_row[0]):
                        check_count += 1
            is_district_pivot = check_count >= 3
            break

    if is_district_pivot and header_row_idx is not None:
        header = rows[header_row_idx]
        col_district = 0
        col_received = None
        col_accepted = None
        col_sanctioned = None

        for idx, val in enumerate(header):
            if val is None:
                continue
            vl = str(val).strip().lower()
            if "row label" in vl:
                col_district = idx
            if ("cummulative" in vl or "cumulative" in vl or "sum of cummulative" in vl):
                if "received" in vl and col_received is None:
                    col_received = idx
                elif "accepted" in vl and col_accepted is None:
                    col_accepted = idx
                elif "sanctioned" in vl and col_sanctioned is None:
                    col_sanctioned = idx

        for row in rows[header_row_idx + 1:]:
            dr = row[col_district] if col_district < len(row) else None
            if not is_district_name(dr):
                continue
            d = normalize_district(dr)
            entry = {}
            if col_received is not None and col_received < len(row):
                entry["applications_received"] = safe_float(row[col_received])
            if col_accepted is not None and col_accepted < len(row):
                entry["applications_accepted"] = safe_float(row[col_accepted])
            if col_sanctioned is not None and col_sanctioned < len(row):
                entry["applications_sanctioned"] = safe_float(row[col_sanctioned])
            districts[d] = entry

    if districts:
        print(f"  [KCC_AH] Extracted {len(districts)} districts from Sheet1 pivot")
        return districts

    # Strategy 3: Use raw data (Weekly KCC camps sheet) and aggregate by district
    raw_ws = None
    for sn in wb.sheetnames:
        if "weekly" in sn.lower() or "kcc camps" in sn.lower():
            raw_ws = wb[sn]
            break
    if raw_ws is None:
        # Try first sheet if it has LDM data
        test_ws = wb[wb.sheetnames[0]]
        test_rows = list(test_ws.iter_rows(max_row=5, values_only=True))
        for row in test_rows:
            if row and any(str(v or "").lower().startswith("ldm") for v in row if v):
                raw_ws = test_ws
                break

    if raw_ws is not None:
        rows = list(raw_ws.iter_rows(values_only=True))
        # Find header row
        header_row_idx = None
        for i, row in enumerate(rows):
            row_str = " ".join(str(v or "").lower() for v in row)
            if "district" in row_str and "bank" in row_str:
                header_row_idx = i
                break

        if header_row_idx is None:
            # Check if first data row has LDM
            for i, row in enumerate(rows):
                if row and any(str(v or "").lower().startswith("ldm") for v in row if v):
                    header_row_idx = i - 1
                    break

        if header_row_idx is not None:
            header = rows[header_row_idx] if header_row_idx >= 0 else [None] * 35
            district_col = 2  # Default
            cum_received_col = None
            cum_accepted_col = None
            cum_sanctioned_col = None

            for idx, val in enumerate(header):
                if val is None:
                    continue
                vl = str(val).strip().lower()
                if "district" in vl:
                    district_col = idx
                if ("cummulative" in vl or "cumulative" in vl):
                    if "received" in vl and cum_received_col is None:
                        cum_received_col = idx
                    elif "accepted" in vl and cum_accepted_col is None:
                        cum_accepted_col = idx
                    elif "sanctioned" in vl and cum_sanctioned_col is None:
                        cum_sanctioned_col = idx

            if cum_received_col is None:
                cum_received_col = 19
            if cum_accepted_col is None:
                cum_accepted_col = 20
            if cum_sanctioned_col is None:
                cum_sanctioned_col = 21

            district_data = defaultdict(lambda: {"received": 0, "accepted": 0, "sanctioned": 0})
            start = max(header_row_idx + 1, 0)

            for row in rows[start:]:
                if len(row) <= district_col:
                    continue
                dr = row[district_col]
                if dr is None:
                    continue
                ds = str(dr).strip()
                if ds == "" or ds.lower() in ("total", "grand total"):
                    continue
                norm = normalize_district(ds)
                if norm not in HARYANA_DISTRICTS:
                    continue
                rcv = safe_float(row[cum_received_col]) if cum_received_col < len(row) else None
                acc = safe_float(row[cum_accepted_col]) if cum_accepted_col < len(row) else None
                san = safe_float(row[cum_sanctioned_col]) if cum_sanctioned_col < len(row) else None
                if rcv is not None:
                    district_data[norm]["received"] += rcv
                if acc is not None:
                    district_data[norm]["accepted"] += acc
                if san is not None:
                    district_data[norm]["sanctioned"] += san

            for dist, data in district_data.items():
                districts[dist] = {
                    "applications_received": data["received"],
                    "applications_accepted": data["accepted"],
                    "applications_sanctioned": data["sanctioned"],
                }

    if districts:
        print(f"  [KCC_AH] Extracted {len(districts)} districts from raw data")
    else:
        print(f"  [KCC_AH] No district data extracted from {filename}")
    return districts if districts else None


def extract_hsfdc(z, meeting):
    """Extract HSFDC (SC/ST) district-wise data."""
    filename = find_file_in_zip(z, ["hsfdc"])
    if not filename:
        return None

    wb = open_workbook_from_zip(z, filename)
    if wb is None:
        print(f"  [HSFDC] Cannot open {filename}")
        return None

    # Try various sheet names for district-wise data
    ws = find_sheet(wb, ["District wise", "District Wise", "DISTRICT WISE",
                          "Overall", "Annexure 23.2"])
    if ws is None:
        # Try numbered district-wise sheets, preferring "Revised" versions
        for candidate in ["Revised 2.2", "2.2", "24.2", "23.2"]:
            if candidate in wb.sheetnames:
                ws = wb[candidate]
                break

    if ws is None:
        # Look for any sheet that has district data and "application"/"sponsored" header
        for sn in wb.sheetnames:
            test_ws = wb[sn]
            for row in test_ws.iter_rows(max_row=10, values_only=True):
                row_str = " ".join(str(v or "").lower() for v in row)
                if "district" in row_str and ("application" in row_str or "sponsored" in row_str):
                    ws = test_ws
                    break
            if ws is not None:
                break

    if ws is None:
        print(f"  [HSFDC] No district-wise sheet in {filename} (sheets: {wb.sheetnames})")
        return None

    print(f"  [HSFDC] Found: {filename} -> sheet: {ws.title}")

    districts = {}
    rows = list(ws.iter_rows(values_only=True))

    # Find header row - must have "district" as a separate column header (not in a title)
    header_row_idx = None
    for i, row in enumerate(rows):
        # Check if any individual cell contains "district" as a column header
        has_district_col = False
        has_app_col = False
        non_none_count = 0
        for val in row:
            if val is None:
                continue
            non_none_count += 1
            vl = str(val).strip().lower()
            if "district" in vl and len(vl) < 30:  # Short header, not a title
                has_district_col = True
            if "sponsored" in vl or "sanctioned" in vl or "application" in vl:
                has_app_col = True
        if has_district_col and has_app_col and non_none_count >= 3:
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  [HSFDC] Cannot find header row")
        return None

    header = rows[header_row_idx]

    def find_col(keywords, header):
        for idx, val in enumerate(header):
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if all(kw in val_lower for kw in keywords):
                return idx
        return None

    col_district = find_col(["district"], header) or find_col(["name"], header)
    # Try "Total no. of applications" (overall count)
    col_applications = find_col(["total"], header)
    if col_applications is None:
        col_applications = find_col(["application"], header)
    col_sponsored = find_col(["sponsored"], header)
    col_sanctioned = find_col(["sanctioned"], header)

    # Also try for the older format with different columns
    if col_applications is None:
        # Look for "Application Received" or "Total Application Received"
        col_applications = find_col(["received"], header)

    if col_district is None:
        # District might be second column
        for ci in range(len(header)):
            test_rows = rows[header_row_idx + 1: header_row_idx + 4]
            for row in test_rows:
                if ci < len(row) and is_district_name(row[ci]):
                    col_district = ci
                    break
            if col_district is not None:
                break

    if col_district is None:
        print(f"  [HSFDC] Cannot find district column")
        return None

    for row in rows[header_row_idx + 1:]:
        district_raw = row[col_district] if col_district < len(row) else None
        if not is_district_name(district_raw):
            continue
        district = normalize_district(district_raw)
        entry = {}
        if col_applications is not None and col_applications < len(row):
            v = safe_float(row[col_applications])
            if v is not None:
                entry["applications"] = v
        if col_sponsored is not None and col_sponsored < len(row):
            v = safe_float(row[col_sponsored])
            if v is not None:
                entry["sponsored"] = v
        if col_sanctioned is not None and col_sanctioned < len(row):
            v = safe_float(row[col_sanctioned])
            if v is not None:
                entry["sanctioned"] = v
        # Skip entries where all values are None (e.g. #REF! errors)
        if entry:
            districts[district] = entry

    if districts:
        print(f"  [HSFDC] Extracted {len(districts)} districts")
    return districts if districts else None


def extract_shg(z, meeting):
    """Extract SHG district-wise data."""
    filename = find_file_in_zip(z, ["shg"], exclude_patterns=["hsrlm"])
    if not filename:
        return None

    wb = open_workbook_from_zip(z, filename)
    if wb is None:
        print(f"  [SHG] Cannot open {filename}")
        return None

    ws = find_sheet(wb, ["DISTRICT WISE", "District Wise"])
    if ws is None:
        # Some meetings only have BANK WISE or numbered sheets
        # Check other sheets for district data (not bank names)
        for sn in wb.sheetnames:
            sn_lower = sn.strip().lower()
            if sn_lower in ("bank wise",):
                continue
            test_ws = wb[sn]
            has_district_header = False
            has_district_data = False
            for row in test_ws.iter_rows(max_row=15, values_only=True):
                row_str = " ".join(str(v or "").lower() for v in row)
                if "district" in row_str and ("saving" in row_str or "credit" in row_str):
                    has_district_header = True
                # Check if data rows have district names (not bank names)
                for val in row:
                    if val is not None and is_district_name(val):
                        has_district_data = True
                        break
            if has_district_header and has_district_data:
                ws = test_ws
                break

    if ws is None:
        print(f"  [SHG] No district-wise sheet in {filename} (sheets: {wb.sheetnames})")
        return None

    print(f"  [SHG] Found: {filename} -> sheet: {ws.title}")

    districts = {}
    rows = list(ws.iter_rows(values_only=True))

    # Find header row with district and savings/credit
    header_row_idx = None
    sub_header_idx = None
    for i, row in enumerate(rows):
        row_str = " ".join(str(v or "").lower() for v in row)
        if "district" in row_str and ("saving" in row_str or "credit" in row_str):
            header_row_idx = i
            # Check if there's a sub-header row with "No." and "Amt."
            if i + 1 < len(rows):
                next_str = " ".join(str(v or "").lower() for v in rows[i + 1])
                if "no" in next_str or "amt" in next_str:
                    sub_header_idx = i + 1
            break

    if header_row_idx is None:
        print(f"  [SHG] Cannot find header row")
        return None

    header = rows[header_row_idx]

    # Find district column
    col_district = None
    for idx, val in enumerate(header):
        if val is None:
            continue
        if "district" in str(val).strip().lower():
            col_district = idx
            break

    if col_district is None:
        # Try to find by data
        for ci in range(len(header)):
            start = (sub_header_idx + 1) if sub_header_idx else (header_row_idx + 1)
            for row in rows[start:start + 3]:
                if ci < len(row) and is_district_name(row[ci]):
                    col_district = ci
                    break
            if col_district is not None:
                break

    if col_district is None:
        print(f"  [SHG] Cannot find district column")
        return None

    # Find savings and credit linked columns
    # Header typically has: District Name, Savings Linked during quarter (No., Amt.), Credit Linked (No., Amt.), ...
    col_savings_no = None
    col_savings_amt = None
    col_credit_no = None
    col_credit_amt = None

    # The columns after district are typically: savings_no, savings_amt, credit_no, credit_amt, ...
    savings_found = False
    credit_found = False
    for idx, val in enumerate(header):
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        if "saving" in val_lower:
            savings_found = True
            credit_found = False
        if "credit" in val_lower:
            credit_found = True
            savings_found = False

    # Use positional approach: after district col, savings columns come first, then credit
    # Look at sub_header to identify No. and Amt. columns
    data_start = (sub_header_idx + 1) if sub_header_idx else (header_row_idx + 1)

    # Find the column positions from the header layout
    # The savings and credit columns follow the district column
    savings_cols = []
    credit_cols = []
    current_section = None

    for idx, val in enumerate(header):
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        if "saving" in val_lower:
            current_section = "savings"
        elif "credit" in val_lower:
            current_section = "credit"

    # Simpler approach: read the numeric columns after district
    # Pattern: district, savings_no, savings_amt, credit_no, credit_amt, ...
    # Or for FY columns too

    # Let's just find by position relative to district
    numeric_cols = []
    for row in rows[data_start:data_start + 3]:
        for idx in range(col_district + 1, len(row)):
            val = safe_float(row[idx])
            if val is not None:
                numeric_cols.append(idx)
        if numeric_cols:
            break

    # First two numeric cols after district = savings (no, amt)
    # Next two = credit (no, amt)
    if len(numeric_cols) >= 4:
        col_savings_no = numeric_cols[0]
        col_savings_amt = numeric_cols[1]
        col_credit_no = numeric_cols[2]
        col_credit_amt = numeric_cols[3]

    for row in rows[data_start:]:
        district_raw = row[col_district] if col_district < len(row) else None
        if not is_district_name(district_raw):
            continue
        district = normalize_district(district_raw)
        entry = {}
        if col_savings_no is not None and col_savings_no < len(row):
            entry["savings_linked_no"] = safe_float(row[col_savings_no])
        if col_savings_amt is not None and col_savings_amt < len(row):
            entry["savings_linked_amt"] = safe_float(row[col_savings_amt])
        if col_credit_no is not None and col_credit_no < len(row):
            entry["credit_linked_no"] = safe_float(row[col_credit_no])
        if col_credit_amt is not None and col_credit_amt < len(row):
            entry["credit_linked_amt"] = safe_float(row[col_credit_amt])
        districts[district] = entry

    if districts:
        print(f"  [SHG] Extracted {len(districts)} districts")
    return districts if districts else None


def extract_cd_ratio_from_extracted():
    """Extract CD ratio from already-extracted files."""
    cd_data = {}

    # Map extracted files to meetings by looking at their content
    file_to_meeting = {
        "Dec 2022": "163rd",
        "March 2023": "164th",
        "June 2023": "165th",
        "September 2023": "166th",
        "Dec 2023": "167th",
        "March 2024": "168th",
        "June 2024": "169th",
        "Sept 2024": "170th",
        "Dec 2024": "171st",
        "March 2025": "172nd",
        "June 2025": "173rd",
        "September 2025": "174th",
        "December 2025": "175th",
        "Sept 2025": "174th",
        "Dec 2025": "175th",
    }

    for fname in os.listdir(EXTRACTED_DIR):
        if not fname.endswith(".xlsx") or fname.startswith("~"):
            continue
        if "CD Ratio" not in fname and "CD ratio" not in fname:
            continue
        if "BASIC DATA" in fname:
            continue

        # Determine meeting
        meeting = None
        for date_str, mtg in file_to_meeting.items():
            if date_str.lower() in fname.lower():
                meeting = mtg
                break

        if meeting is None:
            print(f"  [CD_RATIO] Cannot map file to meeting: {fname}")
            continue

        filepath = os.path.join(EXTRACTED_DIR, fname)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"  [CD_RATIO] Cannot open {fname}: {e}")
            continue

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Find header row with DEPOSITS, ADVANCES, CD RATIO
        header_row_idx = None
        for i, row in enumerate(rows):
            row_str = " ".join(str(v or "").lower() for v in row)
            if "deposit" in row_str and "advance" in row_str and ("cd ratio" in row_str or "cd_ratio" in row_str or "ratio" in row_str):
                header_row_idx = i
                break

        if header_row_idx is None:
            print(f"  [CD_RATIO] Cannot find header in {fname}")
            continue

        header = rows[header_row_idx]

        col_district = None
        col_deposits = None
        col_advances = None
        col_cd_ratio = None

        for idx, val in enumerate(header):
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if "district" in val_lower or "name" in val_lower:
                col_district = idx
            elif "deposit" in val_lower:
                col_deposits = idx
            elif "advance" in val_lower:
                col_advances = idx
            elif "ratio" in val_lower:
                col_cd_ratio = idx

        if col_district is None:
            col_district = 0

        districts = {}
        for row in rows[header_row_idx + 1:]:
            district_raw = row[col_district] if col_district < len(row) else None
            if not is_district_name(district_raw):
                continue
            district = normalize_district(district_raw)
            entry = {}
            if col_deposits is not None and col_deposits < len(row):
                entry["deposits"] = safe_float(row[col_deposits])
            if col_advances is not None and col_advances < len(row):
                entry["advances"] = safe_float(row[col_advances])
            if col_cd_ratio is not None and col_cd_ratio < len(row):
                v = safe_float(row[col_cd_ratio])
                if v is not None:
                    # CD ratio might be stored as decimal (0.51) or percentage (51.7)
                    # Normalize to percentage
                    if v < 5:  # Likely decimal format
                        v = v * 100
                    entry["cd_ratio"] = round(v, 2)
            districts[district] = entry

        if districts:
            cd_data[meeting] = districts
            print(f"  [CD_RATIO] {meeting} ({fname}): {len(districts)} districts")

    return cd_data


# ============================================================
# MAIN EXTRACTION PIPELINE
# ============================================================

def main():
    print("=" * 70)
    print("HARYANA SLBC DATA EXTRACTOR")
    print("=" * 70)

    all_data = {}
    meetings = list(MEETING_QUARTERS.keys())

    # First extract CD ratio from already-extracted files
    print("\n--- Extracting CD Ratio from pre-extracted files ---")
    cd_ratio_data = extract_cd_ratio_from_extracted()

    # Process each meeting ZIP
    for meeting in meetings:
        zip_path = os.path.join(BASE_DIR, f"{meeting}.zip")
        if not os.path.exists(zip_path):
            print(f"\n[SKIP] {meeting}.zip not found")
            continue

        print(f"\n{'='*50}")
        print(f"Processing {meeting} ({MEETING_QUARTERS[meeting]['quarter']})")
        print(f"{'='*50}")

        z = zipfile.ZipFile(zip_path)
        meeting_data = {
            "meeting": meeting,
            "quarter": MEETING_QUARTERS[meeting]["quarter"],
            "date": MEETING_QUARTERS[meeting]["date"],
            "districts": {},
        }

        # Initialize district entries
        for dist in HARYANA_DISTRICTS:
            meeting_data["districts"][dist] = {"district": dist}

        # Extract each indicator
        pmjdy = extract_pmjdy(z, meeting)
        aadhaar = extract_aadhaar(z, meeting)
        kcc = extract_kcc_saturation(z, meeting)
        kcc_ah = extract_kcc_ah(z, meeting)
        hsfdc = extract_hsfdc(z, meeting)
        shg = extract_shg(z, meeting)

        # Merge indicators into district data
        for dist in HARYANA_DISTRICTS:
            d = meeting_data["districts"][dist]

            if pmjdy and dist in pmjdy:
                d["pmjdy"] = pmjdy[dist]

            if aadhaar and dist in aadhaar:
                d["aadhaar"] = aadhaar[dist]

            if kcc and dist in kcc:
                d["kcc"] = kcc[dist]

            if kcc_ah and dist in kcc_ah:
                d["kcc_ah"] = kcc_ah[dist]

            if hsfdc and dist in hsfdc:
                d["hsfdc"] = hsfdc[dist]

            if shg and dist in shg:
                d["shg"] = shg[dist]

            if meeting in cd_ratio_data and dist in cd_ratio_data[meeting]:
                d["cd_ratio"] = cd_ratio_data[meeting][dist]

        z.close()
        all_data[meeting] = meeting_data

    # ============================================================
    # BUILD OUTPUT FILES
    # ============================================================

    print("\n" + "=" * 70)
    print("BUILDING OUTPUT FILES")
    print("=" * 70)

    # 1. haryana_complete.json - all data
    complete_output = {
        "state": "Haryana",
        "districts": HARYANA_DISTRICTS,
        "meetings": {},
    }
    for meeting, mdata in all_data.items():
        complete_output["meetings"][meeting] = {
            "quarter": mdata["quarter"],
            "date": mdata["date"],
            "districts": mdata["districts"],
        }

    complete_path = os.path.join(OUTPUT_DIR, "haryana_complete.json")
    with open(complete_path, "w") as f:
        json.dump(complete_output, f, indent=2)
    print(f"Wrote {complete_path}")

    # 2. haryana_fi_timeseries.json - district time series
    timeseries = {}
    for dist in HARYANA_DISTRICTS:
        timeseries[dist] = []
        for meeting in meetings:
            if meeting not in all_data:
                continue
            mdata = all_data[meeting]
            d = mdata["districts"].get(dist, {})
            entry = {
                "meeting": meeting,
                "quarter": mdata["quarter"],
                "date": mdata["date"],
            }
            # Flatten all indicators
            for key in ["pmjdy", "aadhaar", "kcc", "kcc_ah", "hsfdc", "shg", "cd_ratio"]:
                if key in d and d[key]:
                    for field, val in d[key].items():
                        entry[f"{key}_{field}"] = val
            timeseries[dist].append(entry)

    timeseries_path = os.path.join(OUTPUT_DIR, "haryana_fi_timeseries.json")
    with open(timeseries_path, "w") as f:
        json.dump(timeseries, f, indent=2)
    print(f"Wrote {timeseries_path}")

    # 3. Copy to public directory
    os.makedirs(PUBLIC_DIR, exist_ok=True)
    shutil.copy2(complete_path, os.path.join(PUBLIC_DIR, "haryana_complete.json"))
    shutil.copy2(timeseries_path, os.path.join(PUBLIC_DIR, "haryana_fi_timeseries.json"))
    print(f"Copied to {PUBLIC_DIR}")

    # ============================================================
    # SUMMARY
    # ============================================================
    print("\n" + "=" * 70)
    print("EXTRACTION SUMMARY")
    print("=" * 70)
    indicators = ["pmjdy", "aadhaar", "kcc", "kcc_ah", "hsfdc", "shg", "cd_ratio"]
    header = f"{'Meeting':<10} " + " ".join(f"{ind:>8}" for ind in indicators)
    print(header)
    print("-" * len(header))
    for meeting in meetings:
        if meeting not in all_data:
            continue
        counts = {}
        for ind in indicators:
            count = sum(1 for d in HARYANA_DISTRICTS
                        if ind in all_data[meeting]["districts"].get(d, {}))
            counts[ind] = count
        row = f"{meeting:<10} " + " ".join(f"{counts[ind]:>8}" for ind in indicators)
        print(row)


if __name__ == "__main__":
    main()
