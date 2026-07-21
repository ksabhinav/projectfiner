#!/usr/bin/env python3
"""
Add one live MP quarter from the SLBC data-table XLSX.
=====================================================

Madhya Pradesh publishes a per-meeting XLSX at slbcmadhyapradesh.in whose
`CD Ratio_3(ii)Dist` sheet is the only district-wise table in the workbook
(everything else is bank-wise — see CLAUDE.md "Madhya Pradesh Data Pipeline").
This script ingests one such workbook as one quarter.

Usage:
    python3 db/add_mp_quarter_xlsx.py <xlsx> <YYYY-MM> [--dry-run]

    python3 db/add_mp_quarter_xlsx.py \
        slbc-data/madhya-pradesh/raw/Slbc-data-Mar26-Final.xlsx 2026-03

Parsing/writing primitives are imported from normalize_wayback_madhya_pradesh
(the Wayback backfill) so the two paths share one district resolver, one CD
sanity check, and one set of output writers — they cannot drift apart.

Header-row detection looks for a 'district name' substring but SKIPS row 0,
because the sheet's title row ("CREDIT DEPOSIT RATIO (DISTRICT WISE)") also
contains those words. Columns are the clean modern layout the normalizer's
2023-03-onward batches already use: district=1, deposits=2, advances=3, cd=4.

Source is in Rs. Lakhs ("Amount in lakh" in row 2) — no Crore conversion.

After this runs:
  python3 db/regenerate_indicator_files_from_states.py credit_deposit_ratio
  python3 validate_data.py --state madhya-pradesh
"""
from __future__ import annotations
import argparse
import sys
import warnings
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from normalize_wayback_madhya_pradesh import (   # noqa: E402
    LAC, UNRESOLVED, WARN, map_cd, period_label, write_outputs,
)

SHEET = 'CD Ratio_3(ii)Dist'
# Column layout of the modern sheet: 0=Sr. 1=District Name 2=Deposits
# 3=Advances 4=cd ratio.
DCOL, DEP, ADV, CD = 1, 2, 3, 4


def load_rows(xlsx: Path, sheet: str) -> list[list]:
    """Return the sheet's data rows, starting just after the header row."""
    warnings.filterwarnings('ignore')
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f'ERROR: sheet {sheet!r} not in {xlsx.name}\n'
                 f'  sheets: {wb.sheetnames}')
    rows = list(wb[sheet].iter_rows(values_only=True))

    hdr = None
    for i, r in enumerate(rows):
        if i == 0:
            continue          # title row also contains "DISTRICT ... WISE"
        joined = ' '.join(str(c).lower() for c in r if c is not None)
        if 'district name' in joined:
            hdr = i
            break
    if hdr is None:
        sys.exit(f'ERROR: no header row with "district name" in {xlsx.name}')
    print(f'  header row {hdr}: {rows[hdr]}')
    return [list(r) for r in rows[hdr + 1:]]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', type=Path)
    ap.add_argument('period', help='quarter key, YYYY-MM (e.g. 2026-03)')
    ap.add_argument('--sheet', default=SHEET)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f'ERROR: {args.xlsx} not found')
    pk = args.period
    if len(pk) != 7 or pk[4] != '-':
        sys.exit(f'ERROR: period must be YYYY-MM, got {pk!r}')

    print(f'{args.xlsx.name} -> {period_label(pk)}')
    rows = load_rows(args.xlsx, args.sheet)
    res = map_cd({'rows': rows}, pk, dcol=DCOL, dep_idx=DEP, adv_idx=ADV,
                 cd_idx=CD, factor=LAC, source=args.xlsx.name)
    print(f'  resolved {len(res)} districts [Lakh]')

    if WARN:
        print(f'\n--- {len(WARN)} CD sanity warnings ---')
        for w in WARN:
            print('  ' + w)
    if UNRESOLVED:
        print('\n--- unresolved district names (dropped) ---')
        for k, names in sorted(UNRESOLVED.items()):
            print(f'  {k}: {sorted(names)}')

    # MP has had 55 districts since the 2023-24 reorg; a short table means the
    # sheet layout moved and we'd be silently writing a partial quarter.
    if len(res) < 50:
        sys.exit(f'ERROR: only {len(res)} districts resolved — expected ~55. '
                 f'Check the sheet layout before writing.')

    if args.dry_run:
        for d in list(res)[:3]:
            print(f'  (dry) {d}: {res[d]["credit_deposit_ratio"]}')
        return

    write_outputs({pk: res})


if __name__ == '__main__':
    main()
