#!/usr/bin/env python3
"""
Append a Karnataka CD-ratio quarter from a slbckarnataka.com CDRatio_*.xlsx.

slbckarnataka.com publishes a per-quarter workbook with a `DW_CDRatio`
(district-wise) sheet:

  SR | Name of District | Branch | Deposits[Rural,Semi-Urban,Urban,Total] |
  Advances[Rural,Semi-Urban,Urban,Total] | CD Ratio

(title: "...Amount in Rs.Crore" — deposits/advances are Crores, matching the
live Karnataka data; cd_ratio is a %, unitless). The site overwrites the
CDRatio link each quarter (cf. CLAUDE.md gotcha #73), so quarters are caught
as they publish.

Unlike the legacy KA quarters (whose ad-hoc dump reused the Rural/Total keys
for both deposits and advances, losing deposits and mis-labelling advances as
total_branch), this writes CLEAN canonical fields:
  credit_deposit_ratio__cd_ratio / __total_deposit / __total_advance /
  __total_branch
so the homepage choropleth (overall_cd_ratio) and analysis pages pick it up
via the standard fallback chain.

Appends the period to the four public KA files (live wins on conflict):
  karnataka_fi_timeseries.json / _complete.json / _fi_slim.json /
  _fi_timeseries.csv

Usage:
  python3 db/extract_karnataka_cdratio_xlsx.py <xlsx> <YYYY-MM> [--dry-run]
Then:
  python3 db/regenerate_indicator_files_from_states.py credit_deposit_ratio
  python3 validate_data.py --state karnataka
"""
from __future__ import annotations
import argparse
import calendar
import json
import re
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings('ignore')

ROOT = Path(__file__).resolve().parent.parent
STATE = 'karnataka'
PUB = ROOT / 'public/slbc-data' / STATE
TIMESERIES_PATH = PUB / f'{STATE}_fi_timeseries.json'
COMPLETE_PATH = PUB / f'{STATE}_complete.json'
CSV_PATH = PUB / f'{STATE}_fi_timeseries.csv'
SLIM_PATH = PUB / f'{STATE}_fi_slim.json'

# XLSX (UPPER) -> live canonical district name. Title-case handles most;
# only genuine spelling differences from the live data need an entry.
ALIASES = {
    'BELAGAVI': 'Belgaum',            # live uses the old name
    'CHAMARAJANAGARA': 'Chamarajanagar',
}
SKIP = {'total', 'grand total', 'state total', ''}

MONTHS = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May',
          6: 'June', 7: 'July', 8: 'August', 9: 'September', 10: 'October',
          11: 'November', 12: 'December'}


def canon_district(name):
    if name is None:
        return None
    s = str(name).strip()
    if not s or s.lower() in SKIP:
        return None
    if s.upper() in ALIASES:
        return ALIASES[s.upper()]
    return ' '.join(w.capitalize() for w in s.split())


def num(v):
    if v is None:
        return None
    s = str(v).strip().replace(',', '').replace('%', '')
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def fmt(x):
    return f'{x:.2f}'.rstrip('0').rstrip('.')


def parse_dw_cdratio(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'DW_CDRatio' not in wb.sheetnames:
        sys.exit(f'ERROR: no DW_CDRatio sheet in {xlsx_path} ({wb.sheetnames})')
    ws = wb['DW_CDRatio']
    out = {}
    warn = []
    for r in ws.iter_rows(values_only=True):
        if not r or r[0] is None or not str(r[0]).strip().isdigit():
            continue                       # skip headers / Total row (sno blank)
        d = canon_district(r[1])
        if not d:
            continue
        branch = num(r[2]) if len(r) > 2 else None
        dep_total = num(r[6]) if len(r) > 6 else None
        adv_total = num(r[10]) if len(r) > 10 else None
        printed = num(r[11]) if len(r) > 11 else None
        if dep_total is None or adv_total is None or dep_total <= 0:
            continue
        derived = adv_total / dep_total * 100.0
        cd = printed
        if printed is None or abs(printed - derived) > max(0.5, 0.05 * derived):
            if printed is not None:
                warn.append(f'{d}: printed CD {printed:.2f} != derived '
                            f'{derived:.2f} -> using derived')
            cd = derived
        rec = {'cd_ratio': fmt(cd), 'total_deposit': fmt(dep_total),
               'total_advance': fmt(adv_total)}
        if branch is not None:
            rec['total_branch'] = fmt(branch)
        out[d] = rec
    return out, warn


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('period', help='YYYY-MM, e.g. 2026-03')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    pk = args.period
    y, m = int(pk[:4]), int(pk[5:7])
    label = f'{MONTHS[m]} {y}'
    data, warn = parse_dw_cdratio(args.xlsx)
    print(f'{label}: {len(data)} districts parsed from {Path(args.xlsx).name}')
    for w in warn:
        print('  WARN ' + w)
    if args.dry_run:
        for d in list(data)[:3]:
            print('  ', d, data[d])
        return

    # ---- timeseries (live wins) -----------------------------------------
    fi = json.loads(TIMESERIES_PATH.read_text())
    if any(p['period'] == label for p in fi['periods']):
        sys.exit(f'ERROR: {label} already present in timeseries — aborting.')
    rows = []
    for d in sorted(data):
        row = {'district': d, 'period': label}
        for f, v in data[d].items():
            row[f'credit_deposit_ratio__{f}'] = v
        rows.append(row)
    fi['periods'].append({'period': label, 'districts': rows})

    def sort_key(p):
        mm = re.match(r'([A-Za-z]+)\s+(\d{4})', p['period'])
        rev = {v: f'{k:02d}' for k, v in MONTHS.items()}
        return (mm.group(2), rev.get(mm.group(1), '00')) if mm else ('0', '0')
    fi['periods'].sort(key=sort_key)
    TIMESERIES_PATH.write_text(json.dumps(fi, ensure_ascii=False, indent=2))
    print(f'wrote {TIMESERIES_PATH.relative_to(ROOT)} ({len(fi["periods"])} periods)')

    # ---- complete -------------------------------------------------------
    comp = json.loads(COMPLETE_PATH.read_text())
    dists = {d: dict(data[d]) for d in sorted(data)}
    fields, seen = [], set()
    for row in dists.values():
        for k in row:
            if k not in seen:
                fields.append(k); seen.add(k)
    comp['quarters'][pk] = {
        'period': label,
        'as_on_date': f'{calendar.monthrange(y, m)[1]:02d}-{m:02d}-{y}',
        'fy': f'{(y if m >= 4 else y - 1)}-{str((y if m >= 4 else y - 1) + 1)[-2:]}',
        'tables': {'credit_deposit_ratio': {'fields': fields, 'districts': dists}},
    }
    comp['quarters'] = dict(sorted(comp['quarters'].items()))
    COMPLETE_PATH.write_text(json.dumps(comp, ensure_ascii=False, indent=2))
    print(f'wrote {COMPLETE_PATH.relative_to(ROOT)} ({len(comp["quarters"])} quarters)')

    # ---- slim (credit_deposit_ratio prefix is included) -----------------
    slim = json.loads(SLIM_PATH.read_text())
    slim_rows = [{'district': d, 'period': label,
                  'credit_deposit_ratio__cd_ratio': data[d]['cd_ratio']}
                 for d in sorted(data)]
    slim['periods'].append({'period': label, 'districts': slim_rows})
    slim['periods'].sort(key=sort_key)
    SLIM_PATH.write_text(json.dumps(slim, ensure_ascii=False, indent=2))
    print(f'wrote {SLIM_PATH.relative_to(ROOT)} ({len(slim["periods"])} periods)')

    # ---- CSV (rebuild wide) ---------------------------------------------
    allf, seen = [], set()
    for p in fi['periods']:
        for row in p['districts']:
            for k in row:
                if k not in ('district', 'period') and k not in seen:
                    allf.append(k); seen.add(k)
    allf.sort()
    lines = ['district,period,' + ','.join(allf)]
    for p in fi['periods']:
        for row in p['districts']:
            lines.append(f"{row['district']},{row['period']}," +
                         ','.join(str(row.get(f, '')) for f in allf))
    CSV_PATH.write_text('\n'.join(lines) + '\n')
    print(f'wrote {CSV_PATH.relative_to(ROOT)} ({len(lines) - 1} rows)')


if __name__ == '__main__':
    main()
