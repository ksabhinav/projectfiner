#!/usr/bin/env python3
"""
Extract district-wise data from Karnataka SLBC annexure PDFs.
Produces karnataka_complete.json in the project's standard format.
"""

import os, json, re, glob
import pdfplumber

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_DIR = os.path.join(BASE_DIR, "karnataka/pdfs")
OUT_JSON = os.path.join(BASE_DIR, "karnataka/karnataka_complete.json")

# Karnataka districts (31)
KARNATAKA_DISTRICTS = [
    "BAGALKOTE", "BALLARI", "BELGAUM", "BENGALURU RURAL", "BENGALURU URBAN",
    "BIDAR", "CHAMARAJANAGAR", "CHIKKABALLAPURA", "CHIKKAMAGALURU", "CHITRADURGA",
    "DAKSHINA KANNADA", "DAVANAGERE", "DHARWAD", "GADAG", "HASSAN",
    "HAVERI", "KALABURAGI", "KODAGU", "KOLAR", "KOPPAL",
    "MANDYA", "MYSURU", "RAICHUR", "RAMANAGARA", "SHIVAMOGGA",
    "TUMAKURU", "UDUPI", "UTTARA KANNADA", "VIJAYAPURA", "VIJAYANAGARA", "YADGIR",
    # Alternate spellings
    "BELLARY", "BELAGAVI", "GULBARGA", "SHIMOGA", "TUMKUR",
    "BANGALORE RURAL", "BANGALORE URBAN", "BELGAVI",
]

DIST_NORMALIZE = {
    "BELLARY": "BALLARI", "BELAGAVI": "BELGAUM", "BELGAVI": "BELGAUM",
    "GULBARGA": "KALABURAGI", "SHIMOGA": "SHIVAMOGGA", "TUMKUR": "TUMAKURU",
    "BANGALORE RURAL": "BENGALURU RURAL", "BANGALORE URBAN": "BENGALURU URBAN",
}


def quarter_metadata(qkey):
    year, month = qkey.split("-")
    year = int(year)
    month_num = int(month)
    month_names = {3: "March", 6: "June", 9: "September", 12: "December"}
    last_days = {3: 31, 6: 30, 9: 30, 12: 31}
    if month_num <= 3:
        fy = f"{year-1}-{str(year)[2:]}"
    else:
        fy = f"{year}-{str(year+1)[2:]}"
    return {
        "period": f"{month_names[month_num]} {year}",
        "as_on_date": f"{last_days[month_num]:02d}-{month_num:02d}-{year}",
        "fy": fy,
    }


def is_district_row(row):
    if not row or len(row) < 2:
        return False
    for cell in row[:3]:
        if cell:
            cell_upper = cell.strip().upper()
            for d in KARNATAKA_DISTRICTS:
                if d in cell_upper:
                    return True
    return False


def get_district_name(row):
    for cell in row[:3]:
        if cell:
            cell_upper = cell.strip().upper()
            for d in KARNATAKA_DISTRICTS:
                if d in cell_upper:
                    normalized = DIST_NORMALIZE.get(d, d)
                    return normalized.title()
    return None


def categorize_table(title):
    title_upper = title.upper()
    mappings = [
        ("PRIORITY SECTOR TARGET", "acp_target_achievement"),
        ("ACP OUTSTANDING", "acp_outstanding"),
        ("ACP.*NPA", "acp_npa"),
        ("CD RATIO", "cd_ratio"),
        ("DEAF", "deaf"),
        ("PMEGP.*OUTSTANDING", "pmegp_outstanding"),
        ("PMEGP.*NPA", "pmegp_npa"),
        ("PMEGP", "pmegp"),
        ("PMFME", "pmfme"),
        ("APY", "apy"),
        ("BRANCH NETWORK", "branch_network"),
        ("ATM NETWORK", "atm_network"),
        ("NPA", "npa"),
        ("RSETI", "rseti"),
        ("KCC", "kcc"),
        ("EDUCATION", "education_loan"),
        ("MUDRA", "pmmy_mudra"),
        ("PMMY", "pmmy_mudra"),
        ("PMJDY", "pmjdy"),
        ("PMAY", "pmay"),
        ("SHG", "shg"),
        ("HOUSING", "housing"),
        ("AGRICULTURE", "agriculture"),
        ("MSME", "msme"),
    ]
    for keyword, category in mappings:
        if re.search(keyword, title_upper):
            return category
    return "uncategorized"


def extract_district_tables(pdf_path):
    """Extract district-wise tables from a Karnataka PDF."""
    pdf = pdfplumber.open(pdf_path)
    results = []

    for page_num, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        first_line = text.split("\n")[0].strip() if text else ""

        # Only process pages with DISTRICT in title
        if "DISTRICT" not in first_line.upper() and "district" not in first_line.lower():
            # Check more lines
            has_district_title = False
            for line in text.split("\n")[:5]:
                if "district" in line.lower():
                    has_district_title = True
                    first_line = line.strip()
                    break
            if not has_district_title:
                continue

        tables = page.extract_tables()
        if not tables:
            continue

        for table in tables:
            if not table or len(table) < 3:
                continue

            has_district = any(is_district_row(row) for row in table)
            if not has_district:
                continue

            # Find header row
            header_idx = 0
            for i, row in enumerate(table):
                if is_district_row(row):
                    header_idx = max(0, i - 1)
                    break

            headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(table[header_idx])]

            for row in table[header_idx + 1:]:
                if not row:
                    continue
                clean_row = [str(c).strip() if c else "" for c in row]
                district = get_district_name(clean_row)
                if district:
                    results.append((first_line, headers, district, clean_row))

    pdf.close()
    return results


def process_pdf(pdf_path, quarter_key):
    print(f"  Processing {os.path.basename(pdf_path)} ({quarter_key})...")

    try:
        rows = extract_district_tables(pdf_path)
    except Exception as e:
        print(f"    ERROR: {e}")
        return {}

    if not rows:
        print(f"    No district-level data found")
        return {}

    categories = {}
    for title, headers, district, row_data in rows:
        category = categorize_table(title)

        if category not in categories:
            categories[category] = {"fields": [], "districts": {}}

        if not categories[category]["fields"]:
            categories[category]["fields"] = headers

        dist_dict = {}
        for h, v in zip(headers, row_data):
            if h:
                dist_dict[h] = v
        dist_dict["District"] = district
        categories[category]["districts"][district] = dist_dict

    for cat in categories:
        categories[cat]["num_districts"] = len(categories[cat]["districts"])
        if "District" not in categories[cat]["fields"]:
            categories[cat]["fields"] = ["District"] + categories[cat]["fields"]

    districts_found = set()
    for cat_data in categories.values():
        districts_found.update(cat_data["districts"].keys())

    print(f"    Found {len(categories)} categories, {len(districts_found)} districts")
    return categories


def process_excel(xlsx_path, quarter_key):
    """Process an Excel file for district-wise data."""
    try:
        import openpyxl
    except ImportError:
        print(f"    Skipping Excel (openpyxl not installed)")
        return {}

    print(f"  Processing {os.path.basename(xlsx_path)} ({quarter_key})...")

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"    ERROR: {e}")
        return {}

    categories = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        # Check if sheet has district data
        has_district = False
        for row in rows:
            for cell in (row[:3] if row else []):
                if cell and str(cell).strip().upper() in [d for d in KARNATAKA_DISTRICTS]:
                    has_district = True
                    break
            if has_district:
                break

        if not has_district:
            continue

        # Find header row
        header_idx = 0
        for i, row in enumerate(rows):
            for cell in (row[:3] if row else []):
                if cell and str(cell).strip().upper() in [d for d in KARNATAKA_DISTRICTS]:
                    header_idx = max(0, i - 1)
                    break

        headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(rows[header_idx])]
        category = sheet_name.lower().replace(" ", "_").replace("-", "_")

        cat_data = {"fields": headers, "districts": {}}
        for row in rows[header_idx + 1:]:
            clean_row = [str(c).strip() if c else "" for c in row]
            district = get_district_name(clean_row)
            if district:
                dist_dict = {"District": district}
                for h, v in zip(headers, clean_row):
                    if h:
                        dist_dict[h] = v
                cat_data["districts"][district] = dist_dict

        if cat_data["districts"]:
            cat_data["num_districts"] = len(cat_data["districts"])
            if "District" not in cat_data["fields"]:
                cat_data["fields"] = ["District"] + cat_data["fields"]
            categories[category] = cat_data

    districts_found = set()
    for cat_data in categories.values():
        districts_found.update(cat_data["districts"].keys())

    if categories:
        print(f"    Found {len(categories)} categories, {len(districts_found)} districts")
    else:
        print(f"    No district-level data found")
    return categories


def main():
    files = sorted(glob.glob(os.path.join(PDF_DIR, "*")))
    pdfs = [f for f in files if f.endswith('.pdf')]
    excels = [f for f in files if f.endswith(('.xlsx', '.xls'))]

    print(f"Found {len(pdfs)} PDFs, {len(excels)} Excel files")

    data = {
        "source": "SLBC Karnataka (State Level Bankers' Committee, Karnataka)",
        "state": "Karnataka",
        "amount_unit": "Rs. Crores",
        "quarters": {},
    }

    for file_path in pdfs + excels:
        fname = os.path.basename(file_path)
        match = re.match(r'(\d{4}-\d{2})_', fname)
        if not match:
            print(f"  Skipping {fname} (no quarter key)")
            continue
        qkey = match.group(1)

        if file_path.endswith('.pdf'):
            categories = process_pdf(file_path, qkey)
        else:
            categories = process_excel(file_path, qkey)

        if not categories:
            continue

        meta = quarter_metadata(qkey)
        if qkey in data["quarters"]:
            # Merge with existing
            data["quarters"][qkey]["tables"].update(categories)
        else:
            data["quarters"][qkey] = {
                "period": meta["period"],
                "as_on_date": meta["as_on_date"],
                "fy": meta["fy"],
                "tables": categories,
            }

    data["quarters"] = dict(sorted(data["quarters"].items()))

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    total_cats = sum(len(q["tables"]) for q in data["quarters"].values())
    print(f"\nDone! {len(data['quarters'])} quarters, {total_cats} category-quarter combos")
    print(f"Output: {OUT_JSON}")


if __name__ == "__main__":
    main()
